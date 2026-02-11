# -*- coding: utf-8 -*-
"""
Conciliação Multi-Bandeiras (MVP) - Alelo v1.3
Base pronta para futuras bandeiras (módulos), mantendo a lógica atual.

Atualizações v1.3 (sem mudar a lógica de conciliação):
1) Edição no duplo clique (explicita) + botão "Editar selecionado" nas telas de detalhe.
2) Importação OFX mais robusta + LOG de importação separado de LOG de erros.
3) Exportar para Excel em todas as etapas + pesquisa de banco.
4) Anti-duplicidade reforçada (inclusive quando autorização vem vazia) para reimportações duplicadas.
5) Exclusão em massa por período (mm/aaaa OU dd/mm/aaaa a dd/mm/aaaa) para ERP/Vendas/Receb/Banco.
6) Filtro por período (dd/mm/aaaa a dd/mm/aaaa) no cabeçalho, sem quebrar o mm/aaaa (se período preenchido, ele prevalece).
7) Fechamento do mês (Etapa 3): botão "Finalizar Fechamento" + nova aba "Fechamentos" com lista de meses fechados/pendentes
   desde a 1ª data importada e um "calendário" anual simples (12 meses) colorido.
8) Aba "Banco - Pesquisa" com filtro por data e termo (ALELO/TICKET/etc), total do período e exportar Excel.

Requisitos:
  pip install pandas openpyxl ofxparse

Rodar:
  python conciliador_multi_bandeiras_v1_alelo_v1_3.py
"""
from __future__ import annotations

import tkinter as tk
from step2_providers import get_provider
from tkinter import ttk, filedialog, messagebox, simpledialog
from dataclasses import dataclass
from pathlib import Path
import sqlite3
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

def _to_dec(v):
    """Converte valores (str/float/Decimal/None) para Decimal de forma segura.
    Aceita formato brasileiro (ex.: 'R$ 1.234,56')."""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        s = str(v).strip()
        if s == "":
            return Decimal("0")
        s = s.replace("R$", "").replace("\u00a0", " ").strip()
        # normaliza milhares e decimal pt-BR
        s = s.replace(".", "").replace(",", ".")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")

import re
import unicodedata
import json
import pandas as pd
import warnings
import numpy as np

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)

try:
    from ofxparse import OfxParser
except Exception:
    OfxParser = None

FEE_ICON_PNG_B64 = """iVBORw0KGgoAAAANSUhEUgAAAOUAAACsCAYAAACEnmuyAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAADkgSURBVHhe7Z15eBRV1offXtLpzlYJSZMYEnZkkQQBBRRFlE1UYMZtUBFQFGWCzKjgCp/OgDoIjiOLDDo4gqOjMjoDuLGqERRQEBI0hD2QhGyQVJZeKunu74/qztJLSCBJL+n3efKIdSvprqr7q3PvOeeeq7DZbDaCBAniMyidDwQJEsS7BEUZJIiPERRlkCA+RlCUQYL4GEFRBgniYwRFGSSIjxEUZZAgPkZQlEGC+BhBUQYJ4mMERRkkiI8RFGWQID5GUJRBgvgYQVEGCeJjBEUZJIiPERRlkCA+RlCUQYL4GEFRBgniYwRFGSSIjxEUZZAgPkZQlEGC+BhBUQYJ4mMoWrOancViwWq1YrPZsFqtzs0BjLX2fbdr7z7eWLWaTR9uge5dEMK1zif7DZLFitEkwYkSJt1+DY/9YRZXXZlCqEbjfGqQeiiVShQKBUqlEpVK5dzsQouL0mazUVNTg8VicW5qV2zcvI3fzfs/OJoDXbsjREXKYm2h2y1ZmveS06hacFCkUCBWGeH4MXoM6MsLz/6BsSNHIAiC85lB3KBSqVCr1SgUCucmaGlR1tTUUFNT43w4wHGIo84yjr7tQagQ0fXrhUZtfzNexG2WLFY0KqVsoaprwGIFBehCNUSa5ZdeSI3J+dcaUK2WLXORzQJ2YepC1ABo1CqkGsvFC1ahQDSa4ehJhJTeLH9+tl2ckfVOusi/3Q5Qq9Wo1fKzqE+LiNJms1FdXd3OhqjUDlPNksTR4ye5548LOJa+C7r3RIjQNVuIteKzC6eDQkmkzYJGF4ouIYa+3bpwRZ/exMfFkHhZIkJkBJGREWhDQ9FqQtDqdLV/y2Q0UlZRSUVFJWJFJfln8yksKeXUyRx+OpyNsaAUyWimQqHivE1+bsbqGnQh6uaLVKFCNBrh6BFSrh/G64ueCw5rm4hSqSQkJKSB1bxkUdpsNsxms/PhdoFZkigTy3nx1dd5d8VfoccQhHBdnXWwXXgI7xCiLkRNlE5LPFb0CXEMHj6Ywf36MeCKvnRJuqwVLI6VnNyzZJ84ScaBTNL3/kzhqTyqaszkVysuTqC1lnM302bO4dl5c0joqA+KswmEhobWCvOSRNl+BWlFFCvY8OV2HnngeYiOREiOAYWiSdaxvkVMjQpDnxBHnz49GXH9NQwddCXx+jjnX2kTRFFkz8FMtmz5msOHj1FUXsmxc+UYzRK6UE3TBapQIJ4theJCVv/zr4wbNcJr1+RPOIR5SaKUJKkdDVkbDlWfe+k1tn7yHbp+SWjUIU23igYT8bHR9E/oQJ8+PRk79kaGDkjxspOkzlvswCxJ/HQgk292fMMPP2VQVF5JZn4JKJUI2iZYPoVKfnll7GfUreP5y8Ln6NWjm91qun5eEHkoq9FoLl6U7dGpI4oiW75J59mFK8g7LyLENU1IokmC6hpSkjvSv0snRo4ayQ3XDaVLUifnU30Sx3Vv37GT7GOn2X2mACzWpoV3FArEc+WQl8+6D5cz4tpr7FYzKEx3qNXqixNl+xu2yvOv5X9/h5Vv/QcSwhF0UY0PVR3zK4OZYb27MHhgP8aOvZEbhg3x2zmWs/XcfvgUAEJYE8WZsZ/psx/mmT/8vt4LKShOZy5KlNXV1e0mDunoiIuXrWZr+kF0+qgLzq0kixVjcTkpfZKYdNO1XH3tUMaOuM75ND/Gyq69P/PNjm/Y9t1P7D6RCyHqCw9rFQrE0+cROnfgk+UvMXzIwKAg3XBRojSZGo+NBQoOx8cri1eyO+vkhYerCgVicRmCPpopY4YzduyNASbGhpgliW9372XLlq9J33uAzDNF6MK0jb+0FCrEyio4UcJH/3uDcTeO8NuRQ2vRbFFaLBaqq6udDwcQ8nBKFEU2fLmdlR9t5OjpvMY7mmPeWFzFPbcNY+Sokfzu9tvaTWcTRZFvd//IF19sZduuA+SZTBcc0koWK8Zfc1my4lnumnhLcJ5Zj2aLsj0MXQuLS1i/8Qve+2wHBoOBiiojJklyPq0W8Uwp8T3imX33Ldx1+0S/ceC0NDm5eWzavJ11H20g88RZhA71M3vcI2aeYv4LM5kyZbL9vgWF2WxRBnoYpLC4hPf+vZ71X+8BwGAwUFxW7urUUSiQaiwYf81n2szbuOWWMUwcN7rhOe2ULek7+fSTTaz94ht0UReYg9vnmWkzJvDYow+22xdafZotykCeTzoLUj52zq2VFKtMYLWycNa99axjwzzY9kfd9efk5rH+040s+MfHYLVecDgrlohMv2NUPc9s+7WYQVHaabIgFSrEs+dI6ZPEvLTpTBw/rl5AnHbbkeqouw+Oueai11bJw9nYaLnJTaKFZLFiFKuYduv1PDtvTru2mKoXX3zxReeDjRGICQOOOeSa/22hurqa6upqzpeVuxdkRg6Tbh7Mk0/MYuK4Mahr18cp7D/tnbr7oNVq6d2zO/169+Rs3hl+/eUk2sgwwNUOqJQKlJoQfso8RqVYQkq/PkRHRTmf1i5ox6K0AopaC/m39zdQbjBitVgpNxipcXZmKRSIGaeZNvM25j4xm2uuGtSwPYgHrHTu1ImUvpejUVbz3f92oU1wH1pyCPP04dOEING79+VEhIc5nxbwtGNRKhBFkc+3fMNrq97nrMGIVq1yFSMOC3matCcn8+QfHuXyHl2CVrHJyPepoz6OASlXEBGlZOvHX6BN0DufCHZhGpUKsrNOoFFZGTxwQL3RSPugHU6ArIDVHvj+kZUr1shxNV2o84m1iBnHWbh4dj3vYDu8bZeMlXh9HA/NmMbCxU8jZnwrr6pxg0alJE+S+Pu6//HRp5/ZjzrmqoHr+XfQznpXnRPi6PGTvL5sDRkFpfIaSHf+LoUCMTOXhYv/yP333NWunQ+XjtzV4vVxPDztXhYu/itixn45tGSxotVokCzW2jInglZDniSx4p332bh5m/3324dHtp15X+WHmpObx/OvLOWTL39A0Hdw6w2Uh6wnmf/CQzw0Y5pPrgcsLC7h2Mkc58NuESIj6N+3t/NhryGKIm+v/YAFT68g9eara48XFp+j3GiqjW2KVSbGDOzN03MeYfiQwfX+QuDSzkQpd4a/vrOWV19/HyHB7qJ3RqFCzDjulGnie2xJ38k7q9+lsKzKuakBBrGc60cOY+miBc5NXkUURVas+geL3vyY+M4didfHus2gEisMpN0xlnl/mOWTL8eWJvDHAvVwzCNfffN9dHoP7nZ72GP67Dt9WpAAJqOJE8Wl7D6Ry+4zBR5/Mk6XUGX2vZepIAjMnvUQE0YNAHv2FEBkuK5BtT5dmJZ3vkxn/cYv7EcCe17ZrkRZUFTMnIVvgC6srsqcE+K5MobeOpApd93u04Ksjy5Mi6DVePzBsw/L6wiCwPJXX6ZjmIZjx3MJC3MNgWhUSozVNaz7aANb0ncGfLcN7KurR2FxCcv//g6Fp4vkjupm1C6aJIQOkTw359HAmb8o3L98fIl4fRzr31lW+//FpaJLvqyg1ZB5LI9PP9lETm5eg7ZAox2IUg5/HMw6zMrX/uV5HglQcF6uXRrAayB9lS5Jndj2wXIyvvrIuakWIU5g7effsWnzduemgCLARSl7WwuKipk6dyH0SnRvORQKxMzTzH9iKmNHjnBu9W/ceZZ9lKuuTJFDJZnHPcYwCQ3hP19uYdfefc4tAUOAi1JOil7/6UbEzFyP5SrE4jLG3HE9kybd5uWqcs1Dq5NXXjjKVbpFoQCbigq7E8U9ckKFtwnVaHh42r2kXD9ALrblRphChI496XKdIFEUnZsDgoAPieTk5tGnV0+EFPcWULJYMZolPnptgX09pG8HqA9lZfPN93vY8vVOth7IblrJR4UCsbwKTv1EzxETuP3m6xg+eJDPFvGqfWapN7q19I66uesWPsFdkyY4N/s9vtv7Lgn5rS+KIn95403QpzqfUIvxxHkWzpzMuBsdovW9W7Jr7z6mpD2OTteTqwddz7y5b7D1+wx0oXbv6oWw2RAiwxBSbyCvpJRXX/0Xk8ZNIVoQGHnnFLtH0xnvWc6EjnpW//N9xIyTzk3g2KyozMCeHw8EpNMnQC2l3KEOZR3l6kFXenzjihUGxgzuWy9bxHespFmS2PjlZv70yhscP5gBl3VHFxMph3Ka98gaRTxfAXn5QDhLVjzN/XdO8okhfGFxCRPue4TMvBK39WUlixVjaQXrXn/ebi1959ldKoFxFS4oEcUK/rLi741aSaot3HnbzfUE6X3MksSW9J1cf9s9TJ38KMcNFoSUgQhxgmwhWlCQAEJsFEJqP3T9Epk3ezG9Ukey7sNPMUuSV+9JvL4Dry96Dk64t4QadQjU2Ni+Y6fdWgZOVw6cK3HiTH4Bn7zzdzkE4s5KlohM/+2N3HDdUPsRpZdvh1XOyf3zYiaNmyJbiNSUpg1PLwWbDWzydnhCShI1cXE88sBT3HHvw+za+7Pz2W2Ikp7dupD25DTEgjLnRrBZEBKiWbvpW/bu2+/c6td4sxe2GmZJYuU76+CyAW49eKJJIj5Rz03XDfOZrJ1DWUd57Jk/ybHU1H7ykK2FrWJT0KiUCKmXs/3AUUbf93tWrFnnNS9nvD6O6ffdDZ42yVUowGgNuLllQIqyoKiYd1e8bV8B4qZjG8zcOXKID8Qk5Y62a+8+Hpv7Ilu3H0BI7ebWsrcpNgtCh0h0UVHMm72YZxa9au/0bR86SU5MYOHc6RjzRdcYs82GLlFg5f+2cfCXrIZtfkzAidIsSaz/dCNok52bwP7GjU+IZejVV9odGm3byRqi5FBWNouXrWb3/uMInePcv0S8hGNI++6K//HKkmUcyjqKWWokJtriWBEEgZvH3ASxOqQa1yLgGnUI5FWS+fNBr1n0libgRGkyGlnwyj/RdXezTlKhwFhezs2D+zHi2msatnmBwuIS3n3/Y7Z+8h1CZzff10cQUruw9uOvWbRoCUePuw9TtCb6uFiemnEHxjNFrtMRmwWSo9iw43sOZR9r2OanBJwov939I5TluV0FItVYEDrEcO01Q+uty/POLTBLEunf/8DK19YgpHT2KQvpihX04Rw6foqsI0fsntm2QH428fo4xg4fDlHhSDWuLy5BF0rmd4c5fiKnDb9b6+GdHtmK/OO9j6Frd7ed3GgwcUvq5fU8rt6joKiYN9/6AHp1c266eOpbEWeLcgmIRjOdNBoefVjODfZGFlBSYgLTf3sjxuJy5yaZTgI/7NtHQVGxc4vfESCilOeFhcUlbP3kc4SocOcTZFRK+qX0dapm3vaYJYm9+/aze8cPxEdFXLyA7L8nmiTEKhPiuXLE8xXyz7ly+ZjBVOe5vIjPkSxWWZBTf8NdE2/xWmJBQkc940ffCO6SV2w2dEI4H+3YS25+gXOr3xEgopTZvD0dcL80SzRJjOrTlWuGDXFuanPKxHLeW78JOiViqq5GGxJyUYIRz5UjloikJHTgnmsHMO3W60mbPJ60yeOZduv13HPtAMZc2ZsonRaxoAzxbKm8OxhNE6gjL/g3N1/HXbdP9GopjlCNhu6dkxk6IkXeMsIJjUqJ8df8ekPYtvcUtxQBkmYnp1iNvHMKe7LOuE3LEisMTL9tBH976QWvDL/qcygrm6sH3YqQ0kM+oFCgDQnBVF3tdtjtTP0S//1Tr6BzUiLdOycTGRmBNlQuM2Aym6moqESsqCT/bD6FJaWcOpnD4cPH2L4rC8LV6ITwRtP2xAoD08Zey+xHH/SJolu1xbZe+AdCL6cXhH0b9+m/vdFpp2j/I0BEKT+whIR+CKkpLp1MsliJ0mmZ+8BdzJ4xtUFbW2OWJDZ/nc7vfvMoQmrPuu/aDGGKBhPzp0xk0qTbmiUWURQ5lH2M4ydy+P6HPaz9/DsA1y3rFArESiNjBvbm5eeftH+Gb+SWbknfyaTpc+U8YKfqBJLFSgeFkv+9t6xZ98XX8P5dbiFkd3iI2w5trK6hf0IHBqZc4dzU5piMRk7n5oNWaPhdbbYmDWXFgjLSfjOah2ZMa3bHEwSB4UMGM3Xy7cx+9EHWvfYs0269Xh7aOoaE9mVeo/p25ek5j9T7DN/oKol6PWOuTcXoGIbXQ6NSkvfL8Tb2ELc8vnGnW4D/bvwC9PHOh2UMZjrGCFx1ZYpzi1c4dTIHItWu4ruAMCWLFULht7U7H188/fv25q5JE3h23hxWv/o0Q/smI54qQKwykhobxbT772T4kIHOv+Z1khMTGDFkIJw3OjfJ6GM5knWEMtGDl9YPCBhRrvzoM7dlIyWLFZ0QTr+Uvk5b1nmPRss9NiJMo1kipVsSQmREg+OXQpekTkydfDsvPfU4aTPvRAjTkjZ7BhPHj7N3D+/fr/oIQiRduiSByeJyfwCI0rLnUDYms9m5xW8ICFGKogi5GW4TBgB6xkZxee9ezoe9RoXBAGo3HcqBJ2EqW+9xDR8ymHl/mMUnK15m0vhR9ZxhrfeZF4eSxMsSSbm6K6LRVXi6UA1b9xykoqLSuclv8LU7flHI88nObi/HWF1Dx6gIBlzR137E9Zy2JtJNbVMX3AjTUWaxteZM8fo4hg8Z7LVYZFNJSkxg8IA+UGl0SVLXqEMgt4z84mKfs/JNxfs9tAX4OfMX0Ea4zx2trqFjjECXpMucW7xGUmwHMDahw9hsmCSpocWM0PHmWx+w+ev0eif6b0zuYoiOjKBHcpJ8D52fuc0C2giOHD2BKFY0bPMTAkKU6Tt3Qmyo+zlGiJoO8XqfuVStTkdy125wgf0/6lNfmIIulN0ncnl92RqWLltVb9W9b1xfWyAIgjyvLKtyfeYKBUSqOXUyB5PkuqrEHwiIJ7np2wx0QrhLOESyWImPiiC1v/dDIQ5CNRp6dO8CnHBuapRaYQJCZBi7T+Wz9P1NPP30C6xYsy6gFvk2BSEmBl2/RJdnjs0GUVqOnDrjt86egBAlZw+6BJJxzCfDHCLwHZISE0i5/ha3jorGqD/HFMK1SDXVbNifzdJ/riftj88xd/5CN5XpAnNoG67Vkdotvi5tsB66UA07s076rbPHtSf7GbULW52HMXY6RkWQlJjgfNirREdGcPdto+FMM2NpTnNMjUqJEBlGudHE9sOneOfLdOYvep3b7pzG0mWrOJSVHbBDWyEygqR4PVTXuDx7OQ/2dINj/oTfPy3Z8xrvOowBsK+OiG7BuF5LULuaXus+hHMhnJ0/GpUSIUyLRh3CsXPlbD98ipfXfsqMh59g5J1TWPfhpwGzKt+BPi6WhJgOYPaUlmi2e2D9D78XZf7ZfEDnfLiWjjGCj7n45eFkcmICSxbNtu+b0XxxOgsTZM+jQ6AAGeUG9hzJ46m/rmbsrfcyJe1xN8Nb/0SrCSGhYxzUuBMkgA6xtLRVQketjd+LsrCkFLQe4n4qJZowz4L1DvItFwSBCeNGMerW4YjnylpOmHY0Knk7A0eJyqNGiU++3c/Dv/8/hl53q5sqdf4179TqdGjDw92vrwSI1FBYUorJ6CEdz4fxe1EWl5c1OgyM8LGhq4x827skdeIvC5+D8ipEo2sgvCk0Jsz6OCxoRaiKjHID8/6ykv4DxjJ3/kK/LGYcqtEQHxcDFR4socqHd8q9AP71JNxQlF8IIa4dUrJY0YWoCY/1paGrK/379mbbZ+/A0TOIlW7ibk3AJEke5lWu1FrQmGiqYqNY+eGX9Ok1mlnznqewuMT5dD/AQxcOUWCqanos2JfwcEX+Q4XBABqV286sUavQR7mvROBLDB8ymG3ffgQKuZqAu2tpDTQqpVzftV8S7679jK5XXMuKNev8ax6m9zwSKigq8csEAr8XZblYBaHu11E6gu3+wPAhg/lx/VuMurIX4smz7iuCtxIalRKhawK65C7Mm72YO+592B5O8W1CtTqIdK0y4e/4vyhNjW2GClHhnt+kvkb/vr15f80yFs5/FGN5OaJJantxpiSxfX82v/ntA6zfsMn5FN/DTdKIg0aXyPkwnq/ITzBa1NA2o702QK4IPnfOLH7cuI5JKT2INFvaXJxCnECeJpSpjy1i6bJVPj+cbct70xb4vSh1qhpwHbnWUl7lb6lWcgfr37c3H77/Fstfe45JKT3ooFDKpSTdpJW1BoIuFJ0+igVL32X539f4tDDdpVg2ul7Vx3FzNf5Fx8gwsAbKm1Lp9EiUTBw3mrX/fJNXFszmnqEppEaFIVmsrW89bTY0KqUszOXv8tGnnzmf4XXMJqO8ptIdNTbCQ/1zvun3ouwQFQPV7ktDmCTJb93i9QnVaLhr0gTeeXsZf/3rn3lw/AiGJSegUasQDaZWtZ4atQrUGla8875vZgN5EiUQE+EhqcTH8XtRRkRGgGRx8b5qVEoki5WCIn+MvXlm+JDBLF20gHdXL2X+zHu555oBDEtOkIe2FYaWt542G0JkGJkHjrFly9e+t0SswsMuYNU2QsPC0Gr8xwPvwO9FmdAxzuODMVbX+K0H7kJ0SerE7BlTeeftZSx6YS7z772Ne64bSM/YKHnbAkPLXrcuuSMrP9niU7smG42NxyC14R62r/Bx/F6UcmUzDx3QapWTCwKc4UMG8/wzT/LKy//H/CdnkTZ5PGOu7I1YYUA8X9Ei1lOjUkKZmSNZR3wi88csSbITL1Lt3CRTVkV8XAxana/lPl8YvxelEBMDuLeUADVlFT7tOWxJ4vVxTBw3mqWLFvDy80+ycNa93HPjVRiraxBLyxqe7GYOfkH04WzY8T3HTuY4t7Q5JqMRsbgYtJ6cOWY5ucAP8XtRJur1gIcV/EolleZqvy7Me7H079ubuXNm8cILT7PkDw9wz8ihiGdKa61mU5LYnXHsAykvl7t063splFVUUlrZeKnO2JgYr+8bczH4vSiTExOAQo8drMooUVxyzvlwO8FaO/d84YWnWfjsA7LVNElNXl3SAJsNosPJycn1eqW4iopKCkrPyymWTkgWK0R3aNGi1W2Jn4tSzoABe4dxQheipqpa4sTpM85N7Y4uSZ2YO2cWq597DIyyl9a5rmyTCFFQUFRCmZfr34gVlRw8lgchrnNKY3UNPVO71jviXaveXPxclPLXT7pqhMdY3WmxSt5Qx0cxSxI5uXmtNO91fbxTJ9/O/Bl3YzxT5Lbgc6MoFKBR+YRH+1xpKceyctA5i1KhgOoahvfpRmStpXS9D76Mf31bD0y64Toodl2LqFEpESuN/HLYd1c8HD1+kleWLOPt9z5ssxjglCmToUKUF1W7K/jsCTejEW9gliTKSsuhokauiF4fmw3KzCRf1rF2r05/IyBEmdr/CiircH85Viul+YU+WThKFEW+2rqDtW99xLzX1/DKkmWs37Cp1UMOCR316Pr1Qqqpi/M1SZiNtbUhZWI5Z06dbCQcUsHlfS8nWnDd8MkfcNOL/Y9BA64ATjkflglRU1hWZa965zuYJYk9BzNZsOoDdP16oQvVsHbTt/zplTdY8sYqNm7e1qovEuOJ886HaoWp9eSxtNlAsng9p7S45Bx7DmVDuJvvqVAAVhIvS/RLzyuBIspePbrZ/+U6odeFqDkpVnL8hPdja/UpKCpmxVvroNqCRq2S1zLqo8mvVrDyg89Z9NoqFraS5fzpQCaYRLerK0yOua0nq1htI6FjnFfLduYXF7P152x0YVqXvUSkGgsp1/f3W88rgSLKUI2GobfegVjlmpysUSkpLC0n49Avzk1eQxRF1n+6ka3bD8hbmzvmavaVGUKcQGbBeVZ+8DlLVr7LC4uWsGLNuhapBpCTm8f8Py2FHp08Cq/RoWxZFV26JHmtbKdZkijIL4L8SteXikKBUaxi8IA+9Zw8/kdAiBLgzvFjIc+Dm95qJfdkTps5Ui7EoexjLPjbWoTkGOcmGYWCnpfpSR3QB6NCxdptu1m0+n2eWfAyc+cvZMWadezau69Zw1tRFFm/Qd57ZHfWSYRwXaOOG3dDWdEkET+4uz2LyjsUFBXLL9gYNy8Nmw3yRK4ZPJiEjvqGbX6EwmZr5Mm4weQpz9TL5OTm0adXT4TUG1w6m2iSSEnowPwnZzFx3OgGbW2NKIpMmpHGnv3HEPTRLt8VQB8jEBYWRmHxOSLD5VSx4rJyxEojmKsRYqO5KllPxxiB0OhIkmI7kNy1G9ExUWh18nwvXKvjXGkpZfZRwvnCYg7l5JF54qxsnZuIQ5QmSUIsEZk24QaenTeHLkmdnE9tE3bt3cd9T7xIudHk1lKKGSfZsHkNY0dc17DNjwgYUZoliegrrkcXFeX6sADRVMH8e+/g+WeedG5qM8ySxNvvfci8+Ss8WkmtRkNkuI6KKiOR4TrCwsIw2JPqi8vKwWZDslgxVtfI+2ggO7PioyLoGKYhPKTOslVVSxQZJArLysFiRRemlddHNu+R1wqzcN8JVv9zIVMn3+58SptgliQ2frmZqZOfQUhJcm5GNEkMvbwTKxYuoH/f3s7NfoNr7/VTQjUa0n53G8ZiDyUaq1UcP+7dIezR4yeZN38FukT387H6qzkcgiwsrpciaBdTbe3WyDCEyDB0IWrKjSaOnSvnYEEJu88UsPtMAZkF5yk3mhDCwxGiwuWXVTMFiX23r8LySsbcMdzu6fYOBUXF7PnxAHQS3D/j4irGXDUQfVysc4tfETCiBJh+391QnOv2snRhWnZknfTaekCzJDF7wUIIV7tYcocYo3RaTNVy7LC+IMPCwqhw48RyoFEpG/w4tisQtBq7EF0XgTcHqcYClUbuvO1mL1kg+f7k5hew8rUPISzU9RkrVFAmkjJwAPH6uIZtfoZr7/VjevXoBr3c70WpUSkpLDgnv2nbGLMksfzva9jz+U6EWKeAtkJBp7gYunS6jHh9LF3s2/YZDAbi9bHE62MxGAx1oYq2RqHAeOI8affeyrhRI5xb2wilnOccGcEdD46Fo0cQM3Lk1EqFAhQqxMoqht56Nd07Jzv/st8RUKIM1Wh46oHJiGc9rAoJDeHw4WPs2rvPuaVVCdVo5K3vkuIRT59vsGeINiSEsLAwcvLO1s4dASqqjOTkF5CTX0BxadO9rC2NeLaUUaP689uJt3jZAinp37c3a15fzOGje1my4ml6hKkQM/YjlldAcRm3jRhuXzXk3wSUKAEevOsOKM5wO+cQwrVs35/ND7v3Oje1InUlI099/zl33D4cMeNreciqUBAZrsNgMKDVaBo4dSLDdfKw8RKGnZeKWCKS0ieJOXMeZviQwW6TM9qaUI2mdjnars2fsGHzp0wY1hcqjnB5715ei5+2JAHjfXVgliSuv+0eMgvO124DV4tCgVhl5J5rBvDEE2lemR/VehDnLgKbip49kmrF6Jg3xutjKSw+58Uhqwrx7DlS+iSxaP7jfhFeKCwuQasJCQhRBpiltBKq0TAvbTocdbOG0mZDCNPy769/4qutO+wH2/bt7ygXWfDTFtImj+dY+iYMBgNhYWEN5pDlRi+8/BQKuaZsxhFGDenDO2+87BeCxF4KRRAi7c+zbZ9pSxNgopSZOH4c6GM97/cYouKnPfvsaWveuQWCILB00QIOHz3GgJ6dyNj1M4XF58jJLyCvpNTFQ9vqKFSI5VUYzxSxcPGTfPaftfaRhD928Da+dy2Mf397F+ouZ8kLaYgZ7pPQhcgwNnyXWc9aegu5XMe/Vr7O4f1fcOfIIYjFZRjFqtoK6C1Ric4Tjr8vVhgQz57jjpuu4vD+L5g7Z1a9s/ypiyj97Pu6J+DmlA5EUSSh1wh0yR3dWh2xysSovl159qk5Tk4M13Nbj/qCkz/XLEls/jqdD9Z9yE/7fqVaraVIaWuwwt7d9TQFORPIUTE+lI5WBR1ClYwafwOPPfpgK6fOuV5ry2KtDZ3U0Rqf0/oErChrU9rmvoHQy70rXzxVwPwnpjJ71kN2B4HjwfoGhcUlpH//Axu27ODowSNIRjMVChXnbXKanS7ENRHBLfZhfIRFXuYUq1Uz6Jor+e2E8QwdkNIGzhH5vpolqVXXODqXVGnNz2pNAlaUOKzl4PHoBHuKWX0UKqSaaqJ0WpbMm8ldkyY0bPcqrlbbLEkcPX6SrCNHOJJ1hKwzuRw5eprTotFlTaEzksXKgIQ4nn06jd7du7WyRXRgra14dya/gPziYsTSUkK1OmJjYujfuycAWp3uEsUj36tde3/m+IkcdDq5PIjRWM0N1w1to2ttWQJalGZJ4qNPP+ORBxYgpCa7jfmJBWXcM26o10Ikl8KW9J3MX/S6+/BPPcTSMqb/ZjSrlrzk3NQqiKLIoexj/HfjF6x87V9AKRBTb85XDFQwYfL9PPTA1Eu21oeyslm0aAkbPt0I6IETjLnjfpb/5QW/FGUTxj7+S6hGw7hRIxh660B52ZO7hILLYvn3Z7v5ausOl+GPr2Oyh01cKrrVx5MHupUQRZEVq/7B6BumsvLDLxFSuyGkDkJI6YqQ0hkhtQtC6lUIqTey6cdsJo37HZNmpF30Am6zJPHN93vY8F0mQspAefXIZQN4es4jfilIAl2UANFCFM/NeRROlMiW0rmT2izoEgWWvr+BjV9u9jth+gbyEDInN4/70+ax6M2P0fVLROgQKVcwsCN7kq32BHkLQpgWIaU/e/YeZcSU2RclzJ8OZPKfL7dApLyOVDxTyvyZd9Gzm/scaH8gwEUpJxMMHZBC2pOTEc+Uup1/aVRKxHPlvLd+k1y/JpBwc70tjxJRFHn+laVs/T4DISG6dg5vkiSw2RArDHJ19uKyukRyO0JCNEaDiXv+uKBZ9YhEUeSH3XvZs+sIQoQO0STRqVMMkybd5uU83UsjwEUpu8gFQWD6fXeT2kPvsWiz0CGSrVv28N+NXzSrY/g8ziODVuLttR+w7cdf0cU0rGoglpZxVecE5k/7DQsfupv5M+8iJaGDSz0lISqSY7mFLHljVYPjjbHnYCYff7YN9OHyKKjMwB9n3ef3SekBLkpqL7FXj26kzZ4BZZ43VtUld2Tlus9479/rm1X/pr1zKCubb9J3I1YYGni5RYOJYZd359mn5vD8M08yd84snn/mSeY/OQucEyNsFlAp+e6b3U0axhYWl7Bly9dkHjiGoAtFrDAwdEgvRl479JKcRr5AOxCljMPpM/13oz1WJ9CoQ9Dpo1jwj4/Z8OX2wJhftsHw9autO/j+ZL5c8tGBQgXHi3j26TR7coYDKxPHjeapmXfJtWfrWXJdiJqjRokNGz6rd3597HNSIP37H1j5v23okjvKVtJczYN33l6v3Kj/0m5EiT1pOe3BqYwa0gfxnBth2izym766hpUr1rD563SfF6ahXMSYLyKeKXX/U1AGuYWttnluYXEJx46fwihWNbSSRiOdrujM0AEp9iMNRye/n3IvmBouGtColBjNklxo2SNKcnLz2L5jJ5Sb0KhViCUid4y/hhuuG3qJMU/fIKDjlO5wpLE98fwS8iQJQRfqPn55voJRV/ZizpyHuWHYEJ982Lv27uOH3XsxN0FwyV27tUrBq11797F42Wq2/pyNEG63lAoF4unzpM2YwEv/97T93tVPg5PFmzD0ZkSjuUGMVTRJjLq8Myv/9rLbkEZt7PmpxQgJ0UgWKx0USl5ZMJu7Jt0aEHZG9eKLL77ofLAxamo875rs+1hRq9Qk6GMJC9ew9es9KDVqVErXoaw2TEtWdi7Fuafp0aMbnTslOp/ideL1cQy4oi9Drh50wZ8+vbqj9bjr8cVz4NCvfLEtnSKDCa3aPhRVKDDnlPDQzN/Rv+/lqFUqwHGP6+718SNZHDiSg1ZTF2c111iID9Nx1eBUt/f88JFjLF/zHidKStGGhlBZJDL1jtHcPvEWoqP8ey7pwP9fK82izht718RbSJs8Xp5fusNmQ0iIZvvewyx/YxVb0nf63FA2VKNBEIQm/7QGZaXlFBmkhgkMNhtQRUJix0ZHGOGhWpBMDQo+O/YUdYcoinzz/R62btmDEBkm1/Ptk8TYsTe6tar+SjsTJbWXHK+PY/p9dzPttyPl+KWH0IGQEM2G7zJZtuxtv5hjtjXlVZUXXSEhPFYAs/xvhzCNjlq2bjiTX8C6jzZAQgf5QMF5Jt10bb15a2DQDkVZR/++vbj/vruZdFOq52JbgHBZDNv3Z/P6sjXBrB83iCbPW603dq+qzslhJ0dZTY+7fdkdShs2fEZm5hk5BFJlYtSIKxl508hWGwV4i3YtSlBy1ZUpPPjIdNkjW+IhNmmzIcQJ7M7OYcnKd/no088a7Wztifi4GOKjnIofKxSADrG0tOFxJwpKz4NGHqE4rK2glau8JzklABw7mcM//7OlrpB1hYmRI4Zx1ZWBZSUJilKel90wbAgzH76fYX27IZ73/NYXOkSSmVfCn5euZvnf1wQTDAAhJoZuQoTrsDM6kpyc3IbHoEGscU/Gr3JhZbv32yRJiCaJcJ2mwQY9hcUl/Gv9p+QVlckhkHPljBkxgJvH3FRvzlr3d/2ddi9KHIkFN47g8TkzSE2IaVyYETryJIkFb33IM4tebVL2SSCTqNfTLVFft68JdkePPpz1X+/BZHRf2d0sSeT+lI6gkzcwqk9Sp4TaMIpZktiz/wDvrv1MDoHUWCA0hLE3Xkf/vr3q/ZaytjuLosihrGx27d3Hrr1yLSZ/Sp1sZyERz6hVKrp2TiKxcyK5x09wPLcIrS7U+TQAtCEatCEq9vx6ghOZmZyvqGDIoAHOp7UDrHTUx/Ld3n38eOAwWvsOYQBatYrT3x/m5gk3OIU2bICSD9Zv4LMN+9HGx9iPyatIumgVPDD1bq7o0xtQkHe2gKdfXkpeaRXaMC2VhSLTbr2OGVPvsYdArIBCFu++A7z33r/590f/ZcMX29i87Tt2fPs96d99z45tX7MtfRdGyUTP7t3sYRrfpN0lD1wIR3LB68vWsDvzBEJCtPMpDRCrTHTSaRk2tB8vPTvX7pqvC5C3B9Zv2MSSle+6LLYWq0ykdIpj/TvLGoQsDmVlM378VKpiG+6QJpoqGJbcha/+u5ZQjaYuUeCBBQgpSUgWK1E6LX+e/UCDRIic3DyW//0d0vceIDO/BCxWqB+isVrlY0B8bDTDusTz4CPTfbZ8ZtBSOuGwmH369CTv1HGysnPRRngOums1aiqtVjKOnOanXbupNBkYMuhK59MCmp7du3Hw54McPHQMbb38V61Gzenz5fz03Q9UmgwUFBez6YvNvLFqLUcMRsJDG661jFdr+eOsqQwe0B+AvLMFTHtqEVURIWjVKipPFzL9jjHMnHZPbSLErr37eP6l1/hw+x6KjGYEXSharQatWoU2RF3339AQtKEhGKRqDuUVc2Dbd8QmxNgtsm8RFKUb1CoVnTslkprSjxNHsmRhRrrOfRyolAq0oSEcLynl1wNZbEv/js5JifWGbfIQK1BRq1TExHbgyC9Z8rA/LKx2SKoNUXO8pJS9GVlk7NrLlz//yvGS0rqUPDuV58sZO3IQf3r6cbDPC9/78D9s2rgTIS4S0SQxrG83Jt89kQFX9AO7xf3b2+/wY/Zp1CEqlAqFnJ2lUAE2RIMJc3UNWk1o7fdRKRVoNSHkGc3898MvuPGGIW4zh7xJcPh6AQqLS3jx1dd5d+1nHstV1qJQINVYMFbX0EmjcRrSBj4bN2/jhef+wmGDGSHKvsbRTv1lWrX3UKGQF0AfLWHoqL58859/1Z5zKCubqwfdjq5fouxxLS4j7d5ba3NpC4tLePNfH/DpVzsJCwur/b3C4nMU5hcTaTHTzT7Pz/jqB+ie4PIiEE0SVBoxntjd4Li3aaSHBcGe+bNqyUIWzn9UXo1haOSlZLOhse8Ped5m5ZNv9jNq9F3Mmve802a1geG6r0MOR0wcN5oVqxbTJywUMeOMfRMjFShUaFRKNGoVGnWI/ZgCsdKImJnO9IdubiBIURR59/2PIVquQihWGhk1qDe/nXhLbQhkz/4DvPr6+xSXig021o0M13HH+GvY8dWHpH/0FukfvcXhozsY1jXR5dkJOh2crWDFmnUNjnuboKVsBlvSd/L7mU+RZ7Bc0AHkQLJYMRpMxBklBl43gJeff9LvquY1F7Oj5u7sxXLlOm0yRNodL9U2KKsCTtBzxAT+/beFLvdj1959jL5hPELKQLAvlp4/ZSLPP/Mk1HPsrFwnh0moty19N32023u8a+8+Hp//ckNnlD21Usw4iNHovR2+nQmKspmIosj9afPY+skWdP16yUMxharxxcT2drHCAAUGUq7uyry06UwcP67RhO22pXU8xoeystl/8BcyDv0C9nzXwf36MXTQlfY6Og0/VxRFJs1IY8+BEwixUYjlVUwa1Jv58+fVCm3X3n3c8djzYN/f05ENJFmsrP6/OW7uq/wZE6fMbLjEDPnZiBkH+XH/Dhche4uWfwoBjiAIbPzXW6z7cDnGX0/KiQaNCZK61f9CZBhCrzgyC84zdfLjRAtJTEl73EdWoNR1BbMksSV9J3PnL2Tj5m0Nzmou/fv2Zurk21m6aAFLFy3gT4/PYeK40fUKWzX83G93/8iez7fIO17bbOjCtFw1dHADweSfzUfMOIY+OorIcF3tbmEatYrEyxLdvOjkzxh743XEx0S5liHRJrD/oPzS8AWCorworPJ2dgX7mHB9f8SjJR7r/rhD0GoQUi9H128An2zew6RxU4gWBCZOmcn6DZvs888L/T1HWpm7n+aTk5vH+g2bmDhlJtGCwKRxU1j52nIyfz5oP+Pi//aFkf+uyWjkd9OfQ9dPdtCIpWX8ZlBfHp52b+2ZoiiSk5OLrl83wsLCCAsLQxsSgrG6htFX9UWIjKg915nLe3UnosZNhlGkutaS+wJBUV4U8m0TBIGP//kWGza8QWKIDfF8RdPFaS89IsRGIaT2QUgZwdbdh5g6+Sn69BqILuZqpqQ9zoo169iSvpOc3Dwna+pIK3P34xmzJJGTm8eW9J2sWLOOiVNmortsCH169WTq5GfYui8LIWUEQmof6NqbM2eL2mTLQLMk8d5/NkBZBRq1CqnGQnx8R0aOargKpKyikoKiEtlhBBgMBkzV1Qi6UBJiOhDZiCgT9XrC1aEN83QVCtCoOJ3nZj9TLxGcU7Yg6z78lD+/uJQ8jZye1+QNeBzYQwSO2kFieRUUGMAkAnlAJELqIIb06szlXZNJ6BiHNjyc+LgYAEK1cizVbJKtgdFYTXlVJb8czuZcSTHpv55GzMgEzgPxoI2FhDCEyLC6z66HY0i4/Lm0Vt5rxUph8Xm6dk5GSBkhz/PKK5g29lreWLKQUI269qWQk5vHX954ky9/yKjdYLeiykhkuI5xV6c0untYTm4e0x+Zy+4zBfWcPQrE8iqGXtGlgQfYmwRF2QosXbaK99d+Qg4KjCYJnVbTPHG6wxHTM0ly8neVJHsyTRYwmZyGl3aLqdWCVgU6JYSGQIi60T1H3OHYmWzuE4+5mau1DGZJ4vk/L2blmk0IyTFyRYGEDm63dnd4Xjf/mNkgPmkwGC4oykNZ2cx4+Akyyg0N0wErDIwZ3JeN/3qrwfne4hJ7ShB3zJ0zi90/fMGff38/w7omEqXTIhpMtRvBXhT2d6eg1cgOo4RohOQYhF5xCClJ8j4dKV3r9uxISZLbkmMQ9NEIUeHNFiQAHaPZcyi7VSvHFxQVs/K1lxGSZYuP1crgAX1cBOkgJqJOjBeDy94rVRKXd01ueMyLBEXZSoRqNMyeMZWvP/83S+bNZMLVvRmWnCBnp1QYLk2gzcVmcxmaNhVdqIatew5y/IT7XbEvFbMk8dgzf4JOA+WhpEliWJdE0h6c6nwqANGREQj6urWWDsLCwjhy6gwVFZXOTbWIntoqaujqQ3uPBEXZBtw1aQIf//Mtli99kbn3TWLSIFmgIJeyFA2mOoE616JtS+yfLVmssmWvkEtX9oiNpLyqslXCNt/u3svWT95DiJXXSupC1IwYfY3HmKFWpyM+LoYKp20PdKoa9h49TX5xcYPjMvK9/TnzF3KqncJXCgWYzjDy2qENj3uR4JzSC5gliZ8OZPLD7r38mplFUanIoYLzFJ4rk0+wD69qHUUOoTbvUTVOrVNJhWg0ykubrFaokRBiormqcwIdYwR69OjC1dcOveQ9JN1hliS6XDcB0WBCCNchGssZltyFd1cv9bAETv7/Lek7mTp3IfoY+ftUVBkxVVcjZpxh4eLZPDztXqfvKoty4pRH2Zpx1GUYL2ZmYzSecvN53iEoSi/jEOjxEzlkHPqF84XFFJWKFJVXkllwHiqNoFRCiKqBWKmf2H0BHFbYWF1Tt7aw2iL/O0JHSkIHOkZFkNQpgYjICLp268LAlCu46sqUVnPuYPdWP/LAEwipfeSKAsCff38/s2e4H7o6yMnN45Uly1j7329qCzJrVEoki5WesVFuHUQbN2/jpYWvc9QoNUiIF0+fZ/4ffmdP4QuKMogbCotLOHYyh/yz+eTk5FJQVEJlRSWSwVgr1iKDRLnRhNEkgdFeoxGgxulRalQQokIXpiU2TEmMLoKwKB3dYzqgCdMRERlBQsc4unRJIvGyRHp26+JmC7nW6aiiKJKQMARdv85y0rnBxKQBvVj95mtNsMhW1m/4nKlPvoJOkJPWHYgFZYwZMYCH7r+b7p2TiYyMIPvESY87XounCji8/wuPHltvEBSlHyCKImfyCxArKjlXWorZZKSwRK4UZ6qqAnDZuiC0XrhAGx5OVHgE0TFRaHVaEvV6khMTmtD5W48XXl/Gq8+tlndetlvzd19+monjRjfpReCIV777yXaEuHrXoVAhnisDbQhjrpTnpVt3H3IRL4CYeZolK569oGVua4KiDGLnwkJoKXJy8+jTayBCysDa1R3XpfZi1ZKX7Gc07bvU7mPyfYaLMKWaankkAei0GqJ0chK6SZK93sZf80l7cjJLFy2o+z0fISjKIG3O3Q/MZNN3h4hPiCVeH0th8Tk+e3upR4+rZ6zs2vszb6xazabvDkFoiMtC5gYoFPJua9oQFs6czGOPzmjVOfPFEhRlkDbFsVaSXlcg6ELRR0cx7uqUS7JYh7Ky+WrrDj5L38WedHuSg6AFtV1w5mo4WwGdBKbdej2jbrrOzfIu3yEoyiBthlmSmPH40+SdKsCmkxcYJ8XrW6RkiiiKHMo+xvETOZw5dZIzZ4s4X15KhFpLh3h9rUNryOBBl/xZrU2zRSlJElZrG2WiBAkoHOGf+giRERcxbPWMWZIoE8sxmc212T2RkREkdNT7rGV0ptmirK6uxmK5wKLeIEHc4smB4+l4c2iJv+EbNPsqlMpm/0qQIHY89R1Px5tDS/wN36DZV6Ly4XLvQYIEAs0WJUFhBgnSqlyUKNVqp/VoQYIEaTEuSpQKhSIozCBBWgG1Wn1xosTxy0GnT5AgLYZSqbw0UQKEhNTtmhQkSJBLw6GnSxKlQqEgNNT9xqpBggRpOqGhoSjsi9mbnTzgDpvNRnV1dTDTJ0iQZqJUKgkJCakVJC0lSgc1NTXtYv/KIEFaArVa7dZh2qKixG41a2pqgql4QYJ4QKVSoVarG1jH+rS4KOtjsViwWq3YbLbg0DZIu0WpVKJQKFAqlU1KvGlVUQYJEqT5XJL3NUiQIC3P/wMShLLjgGf5CwAAAABJRU5ErkJggg=="""

APP_TITLE = "Concilia (Multi-Bandeiras)"
DEFAULT_DB = "db/concilia.sqlite"

# Atalho de pesquisa: transferências/recebimentos de repasse Alelo/Naip no extrato bancário (memo)
# Use o token abaixo no campo "bank_keyword" (Etapa 3) ou "Termo (memo)" (Banco - Pesquisa) para buscar todos.
# (token usado no campo "Filtro Banco (token)" para buscar repasses no banco)
BANK_TRANSF_TOKEN_ALELO = "TRANSF_ALELO_NAIP"
BANK_TRANSF_MEMO_TERMS_ALELO = [
    "RECEBIMENTO FORNECEDOR ALELO INSTITUICAO DE PAGAMENTO S",
    "SISPAG ALELO SA",
    "RECEBIMENTO FORNECEDOR NAIP INSTITUICAO DE PAGAMENTO S",
    "NAIP INSTITUICAO DE PAGAMENTO",
    "ALELO INSTITUICAO DE PAGAMENTO",
]

# FarmaciasApp (Zoop/OTB): termos típicos do extrato
BANK_TRANSF_TOKEN_FARM = "TRANSF_ZOOP_OTB"
BANK_TRANSF_MEMO_TERMS_FARM = [
    "ZOOP",
    "OTB",
    "SISPAG ZOOP",
    "INSTITUICAO DE PAGAMENTO",
    "RECEBIMENTO FORNECEDOR ZOOP",
    "RECEBIMENTO FORNECEDOR OTB",
]

# Ticket: termos do extrato variam; base vazia (usuario cadastra)
BANK_TRANSF_TOKEN_TICKET = "TRANSF_TICKET"
BANK_TRANSF_MEMO_TERMS_TICKET: list[str] = []


# Mantém compatibilidade com versões antigas
BANK_TRANSF_TOKEN = BANK_TRANSF_TOKEN_ALELO
BANK_TRANSF_MEMO_TERMS = BANK_TRANSF_MEMO_TERMS_ALELO



# Palavras-chave adicionais (definidas pelo usuário em tempo de execução)
BANK_USER_MEMO_TERMS: list[str] = []

def bank_all_memo_terms(provider: str | None = None) -> list[str]:
    """Retorna termos fixos + termos do usuário (normalizados, únicos).

    - Para ALELO/NAIP: usa BANK_TRANSF_MEMO_TERMS_ALELO
    - Para FARMACIASAPP: usa BANK_TRANSF_MEMO_TERMS_FARM (Zoop/OTB)
    """
    prov = (provider or "").strip().upper()
    base = (BANK_TRANSF_MEMO_TERMS_FARM if prov == "FARMACIASAPP" else (BANK_TRANSF_MEMO_TERMS_TICKET if prov == "TICKET" else BANK_TRANSF_MEMO_TERMS_ALELO))

    terms: list[str] = []
    for t in (base + (BANK_USER_MEMO_TERMS or [])):
        if not t:
            continue
        tt = str(t).strip().upper()
        if tt and tt not in terms:
            terms.append(tt)
    return terms



def bank_memo_match(token: str | None, memo: str | None, provider: str | None = None) -> bool:
    """Match de memo bancário com padrão único.
    - token vazio => aceita tudo
    - token == BANK_TRANSF_TOKEN => aceita qualquer termo conhecido (fixo + usuário)
    - caso contrário => substring simples (token dentro do memo)
    """
    tk = (token or "").strip().upper()
    m = (memo or "").strip().upper()
    if not tk:
        return True
    prov = (provider or "").strip().upper()
    token_std = BANK_TRANSF_TOKEN_FARM if prov == "FARMACIASAPP" else (BANK_TRANSF_TOKEN_TICKET if prov == "TICKET" else BANK_TRANSF_TOKEN_ALELO)
    if tk == token_std:
        # termos conhecidos (fixos + usuário)
        for tt in bank_all_memo_terms(prov):
            if tt and tt in m:
                return True
        # fallback seguro apenas para Alelo/Naip
        if prov not in ("TICKET","FARMACIASAPP"):
            return ("ALELO" in m) or ("NAIP" in m)
        return False
    return tk in m



# --- Normalização de valor bancário (2 casas) para evitar duplicidades por float/format ---
def _bank_amt2_str(x) -> str:
    try:
        d = Decimal(str(x)).quantize(Decimal("0.01"))
    except Exception:
        return str(x)
    # manter ponto como separador decimal (persistência interna)
    return format(d, "f")

def _bank_sender_token(memo: str) -> str:
    m = (memo or "").upper()
    if "NAIP" in m:
        return "NAIP"
    if "ALELO" in m:
        return "ALELO"
    return "OUTRO"

def _bank_is_transfer_memo(memo: str) -> bool:
    m = (memo or "").upper()
    for t in bank_all_memo_terms():
        if t in m:
            return True
    # fallback: se tiver ALELO/NAIP no memo, tratar como transferência (mais seguro p/ duplicidade)
    return ("ALELO" in m) or ("NAIP" in m)

def _bank_row_exists_transfer(conn: sqlite3.Connection, provider: str, dt_iso: str, amount_s: str, memo: str) -> bool:
    token = _bank_sender_token(memo)
    like = f"%{token}%" if token in ("ALELO", "NAIP") else None
    cur = conn.cursor()
    if like:
        cur.execute(
            "SELECT 1 FROM bank_tx WHERE provider=? AND is_deleted=0 AND dt=? AND amount=? AND UPPER(COALESCE(memo,'')) LIKE ? LIMIT 1",
            (provider, dt_iso, amount_s, like),
        )
    else:
        cur.execute(
            "SELECT 1 FROM bank_tx WHERE provider=? AND is_deleted=0 AND dt=? AND amount=? LIMIT 1",
            (provider, dt_iso, amount_s),
        )
    return cur.fetchone() is not None

def br_money(x: Decimal | float | int | None) -> str:
    if x is None:
        return "R$ -"
    try:
        d = Decimal(str(x))
    except Exception:
        return "R$ -"
    q = d.quantize(Decimal("0.01"))
    s = f"{q:.2f}".replace(".", ",")
    parts = s.split(",")
    parts[0] = "{:,}".format(int(parts[0])).replace(",", ".")
    return f"R$ {parts[0]},{parts[1]}"


def center_window(win, width=None, height=None):
    """Centraliza uma janela/Toplevel na tela."""
    try:
        win.update_idletasks()
        w = width or win.winfo_width() or win.winfo_reqwidth()
        h = height or win.winfo_height() or win.winfo_reqheight()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = max(0, int((sw - w) / 2))
        y = max(0, int((sh - h) / 2))
        win.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        pass

fmt_money = br_money  # compat: antigo nome usado em alguns pontos

import os

# ---- Log em arquivo (logs/app) ----
def _app_log_paths() -> tuple[str, str]:
    base = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base, "logs", "app")
    os.makedirs(log_dir, exist_ok=True)
    return os.path.join(log_dir, "import.log"), os.path.join(log_dir, "error.log")

def _append_line(path: str, line: str):
    try:
        with open(path, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass

def normalize_col(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _log_columns_hint(df: pd.DataFrame, limit: int = 60) -> str:
    """Return a compact, normalized list of columns to help diagnose mapping issues."""
    try:
        cols = [normalize_col(str(c)) for c in df.columns]
    except Exception:
        cols = []
    if not cols:
        return "<sem_colunas>"
    if len(cols) > limit:
        return ", ".join(cols[:limit]) + " ..."
    return ", ".join(cols)

def parse_date(val) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        dtv = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(dtv):
            return None
        return dtv.date()
    except Exception:
        return None


def parse_any_date(val):
    """Interpreta datas vindas como texto, date/datetime ou números (epoch ms)."""
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    try:
        n = float(s.replace(",", "."))
        if n > 1e10:
            return datetime.fromtimestamp(n/1000.0).date()
    except Exception:
        pass
    return parse_date(s)

def fmt_br_date(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def fmt_date(val) -> str:
    """Compatibilidade: versões antigas chamavam fmt_date.
    Aceita date/datetime/str e retorna dd/mm/aaaa (ou '' se inválido).
    """
    try:
        d = parse_any_date(val)
        return fmt_br_date(d) if d else ""
    except Exception:
        return ""

def iso_from_ddmmyyyy(s: str) -> str:
    """Converte 'dd/mm/aaaa' -> 'aaaa-mm-dd'. Retorna '' se inválido."""
    try:
        s = (s or '').strip()
        if not s:
            return ''
        d, m, y = s.split('/')
        return f"{y.zfill(4)}-{m.zfill(2)}-{d.zfill(2)}"
    except Exception:
        return ''

def _json_default(o):
    """Default serializer para json.dumps (Decimal/date/datetime)."""
    try:
        if isinstance(o, Decimal):
            return str(o)
        if isinstance(o, (date, datetime)):
            return o.isoformat()
    except Exception:
        pass
    return str(o)

def _ascii_sanitize(s: str) -> str:
    """Sanitiza MEMO (PDF/OFX/CSV): remove quebras, normaliza espaços."""
    try:
        s = (s or '')
        s = s.replace('\n', ' ').replace('\r', ' ')
        s = ' '.join(s.split())
        return s.strip()
    except Exception:
        return (s or '').strip()

def parse_br_date_str(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    return parse_date(s)

def parse_decimal(val) -> Decimal | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None

def month_start(mm_yyyy: str) -> date | None:
    m = re.match(r"^(\d{1,2})/(\d{4})$", (mm_yyyy or "").strip())
    if not m:
        return None
    mm = int(m.group(1)); yy = int(m.group(2))
    if not (1 <= mm <= 12):
        return None
    return date(yy, mm, 1)

def month_range(any_day: date) -> tuple[date, date]:
    start = date(any_day.year, any_day.month, 1)
    if any_day.month == 12:
        end = date(any_day.year + 1, 1, 1)
    else:
        end = date(any_day.year, any_day.month + 1, 1)
    return start, end

def iter_months(start: date, end: date) -> list[date]:
    ms = date(start.year, start.month, 1)
    me = date(end.year, end.month, 1)
    out = []
    cur = ms
    while cur <= me:
        out.append(cur)
        cur = date(cur.year + (cur.month // 12), ((cur.month % 12) + 1), 1) if cur.month != 12 else date(cur.year + 1, 1, 1)
    return out

SCHEMA = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS imports (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  kind TEXT NOT NULL,
  source_path TEXT,
  imported_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS erp_tx (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  dt DATE NOT NULL,
  bruto NUMERIC,
  liquido NUMERIC,
  autorizacao TEXT,
  raw_json TEXT,
  is_deleted INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS sales_tx (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  dt DATE NOT NULL,
  bruto NUMERIC,
  liquido NUMERIC,
  autorizacao TEXT,
  raw_json TEXT,
  is_deleted INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS receb_tx (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  dt DATE NOT NULL,
  pay_dt DATE,
  bruto NUMERIC,
  liquido NUMERIC,
  autorizacao TEXT,
  raw_json TEXT,
  is_deleted INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS bank_tx (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  dt DATE NOT NULL,
  amount NUMERIC NOT NULL,
  memo TEXT,
  bank_id TEXT,
  raw_json TEXT,
  is_deleted INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS closed_periods (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  month TEXT NOT NULL,
  closed_at TEXT NOT NULL,
  UNIQUE(provider, month)
);

CREATE TABLE IF NOT EXISTS carryover_balances (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  month TEXT NOT NULL,          -- mm/yyyy do mês que GEROU o saldo
  amount NUMERIC NOT NULL,      -- saldo a carregar para o próximo mês (pode ser + ou -)
  confirmed_at TEXT NOT NULL,
  UNIQUE(provider, month)
);

CREATE TABLE IF NOT EXISTS ui_kv (
  key TEXT PRIMARY KEY,
  value TEXT,
  updated_at TEXT
);

CREATE TABLE IF NOT EXISTS fee_rules_tx (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  provider TEXT NOT NULL,
  label TEXT NOT NULL,
  match_text TEXT NOT NULL,
  mdr_percent NUMERIC NOT NULL DEFAULT 0,
  fee_fixed NUMERIC NOT NULL DEFAULT 0,
  transfer_fee NUMERIC NOT NULL DEFAULT 0,
  is_active INTEGER NOT NULL DEFAULT 1
);

"""


# =========================
# Taxas/Tarifas (regras dinâmicas por bandeira) - helper simples
# =========================
def _get_active_fee_rule(conn: sqlite3.Connection, provider: str):
    """Retorna (mdr_percent, fee_fixed, transfer_fee) da regra ativa mais recente para o provider.
    Se não existir, retorna (0,0,0).
    """
    try:
        row = conn.execute(
            "SELECT mdr_percent, fee_fixed, transfer_fee FROM fee_rules_tx WHERE provider=? AND is_active=1 ORDER BY id DESC LIMIT 1",
            (provider.strip().upper(),),
        ).fetchone()
        if not row:
            return Decimal("0"), Decimal("0"), Decimal("0")
        mdr = parse_decimal(row[0]) or Decimal("0")
        fixed = parse_decimal(row[1]) or Decimal("0")
        transf = parse_decimal(row[2]) or Decimal("0")
        return mdr, fixed, transf
    except Exception:
        return Decimal("0"), Decimal("0"), Decimal("0")
# =========================
# Calendário: dias úteis (BR) para conciliação (finais de semana + feriados nacionais)
# =========================
def _easter_sunday(year: int) -> date:
    """Computa o domingo de Páscoa (algoritmo de Meeus/Jones/Butcher)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def br_national_holidays(year: int) -> set[date]:
    """Feriados nacionais (fixos + móveis comuns)."""
    hol = set()
    # Fixos
    hol.add(date(year, 1, 1))   # Confraternização Universal
    hol.add(date(year, 4, 21))  # Tiradentes
    hol.add(date(year, 5, 1))   # Dia do Trabalho
    hol.add(date(year, 9, 7))   # Independência
    hol.add(date(year, 10, 12)) # Nossa Senhora Aparecida
    hol.add(date(year, 11, 2))  # Finados
    hol.add(date(year, 11, 15)) # Proclamação da República
    hol.add(date(year, 12, 25)) # Natal
    # Consciência Negra (mantemos para evitar falso negativo em 20/11)
    hol.add(date(year, 11, 20))

    # Móveis (Páscoa base)
    easter = _easter_sunday(year)
    hol.add(easter - timedelta(days=2))   # Sexta-feira Santa
    hol.add(easter - timedelta(days=48))  # Carnaval (segunda)
    hol.add(easter - timedelta(days=47))  # Carnaval (terça)
    hol.add(easter + timedelta(days=60))  # Corpus Christi
    return hol

def is_business_day(d: date) -> bool:
    if d.weekday() >= 5:
        return False
    if d in br_national_holidays(d.year):
        return False
    return True

def next_business_day(d: date) -> date:
    dd = d
    while not is_business_day(dd):
        dd = dd + timedelta(days=1)
    return dd

def add_business_days(d: date, n: int) -> date:
    """Soma (ou subtrai) n dias úteis."""
    step = 1 if n >= 0 else -1
    remaining = abs(int(n))
    dd = d
    while remaining > 0:
        dd = dd + timedelta(days=step)
        if is_business_day(dd):
            remaining -= 1
    return dd

def connect(db_path: str) -> sqlite3.Connection:
    # Permite usar caminhos com subpastas (ex.: db/concilia.sqlite)
    try:
        d = os.path.dirname(db_path)
        if d:
            os.makedirs(d, exist_ok=True)
    except Exception:
        pass
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA)
    _ensure_schema_migrations(conn)
    conn.commit()

def _ensure_schema_migrations(conn: sqlite3.Connection) -> None:
    """Pequenas migrações sem quebrar DBs antigos."""
    try:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(receb_tx)").fetchall()}
        if "pay_dt" not in cols:
            conn.execute("ALTER TABLE receb_tx ADD COLUMN pay_dt DATE")
            conn.commit()
    except Exception:
        # Se der erro, não aborta o app (ex.: tabela ainda não existe em DB vazio)
        pass

    # bank_tx: versões antigas podem não ter o campo bank_name
    try:
        cols_b = {r[1] for r in conn.execute("PRAGMA table_info(bank_tx)").fetchall()}
        if "bank_name" not in cols_b:
            conn.execute("ALTER TABLE bank_tx ADD COLUMN bank_name TEXT")
            conn.commit()
    except Exception:
        pass


    # fee_rules_tx: versões antigas podem não ter o campo transfer_fee
    try:
        cols_f = {r[1] for r in conn.execute("PRAGMA table_info(fee_rules_tx)").fetchall()}
        if "transfer_fee" not in cols_f:
            conn.execute("ALTER TABLE fee_rules_tx ADD COLUMN transfer_fee NUMERIC NOT NULL DEFAULT 0")
            conn.commit()
    except Exception:
        pass

@dataclass
class ImportResult:
    inserted: int
    dropped_dupe: int
    warnings: list[str]
    errors: list[str]

def _detect_columns(df: pd.DataFrame, candidates: dict[str, list[str]]) -> dict[str, str | None]:
    """Detect columns by normalized name.

    Strategy:
      1) exact match against normalized names
      2) prefix/substring fallback (helps when headers have '(R$)', 'R$', etc.)
    """
    norm_map = {c: normalize_col(str(c)) for c in df.columns}
    inv = {norm: orig for orig, norm in norm_map.items()}
    norms = list(inv.keys())

    out: dict[str, str | None] = {}
    for key, opts in candidates.items():
        found: str | None = None

        # 1) exact
        for o in opts:
            if o in inv:
                found = inv[o]
                break

        # 2) prefix / substring fallback
        if found is None:
            for o in opts:
                # prefix match first
                pref = next((n for n in norms if n.startswith(o + "_") or n == o), None)
                if pref and pref in inv:
                    found = inv[pref]
                    break
                # substring match (last resort)
                sub = next((n for n in norms if o in n), None)
                if sub and sub in inv:
                    found = inv[sub]
                    break

        out[key] = found
    return out

def _read_excel_best_sheet(path: str, prefer_sheets: list[str] | None) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    names = xls.sheet_names
    ordered = []
    if prefer_sheets:
        for s in prefer_sheets:
            if s in names:
                ordered.append(s)
    for s in names:
        if s not in ordered:
            ordered.append(s)

    def score(df: pd.DataFrame) -> int:
        if df is None or df.empty:
            return -1
        cols = [normalize_col(str(c)) for c in df.columns]
        sc = 0
        if any(c in cols for c in ("data_da_venda","data_venda","data","emissao","data_da_transacao","data_transacao")):
            sc += 4
        if any(c in cols for c in ("valor_liquido","vl_liq","liquido","tot_c_desconto","tot_cdesconto")):
            sc += 3
        if any(c in cols for c in ("valor_bruto","vl_bruto","bruto","total","valor")):
            sc += 1
        if any(c in cols for c in ("numero_da_autorizacao","n_da_autorizacao","autorizacao","nsu","numero_autorizacao")):
            sc += 2
        return sc

    best_df = None
    best_sc = -999
    for sh in ordered:
        try:
            df = pd.read_excel(xls, sheet_name=sh, dtype=object)
        except Exception:
            continue
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if df.shape[0] == 0:
            continue
        sc = score(df)
        if sc > best_sc:
            best_sc = sc
            best_df = df
    return best_df if best_df is not None else pd.DataFrame()

def _to_records(df: pd.DataFrame, col_dt: str, col_bruto: str | None, col_liq: str | None, col_auth: str | None):
    recs = []
    for _, r in df.iterrows():
        d = parse_date(r.get(col_dt))
        if not d:
            continue
        bruto = parse_decimal(r.get(col_bruto)) if col_bruto else None
        liq = parse_decimal(r.get(col_liq)) if col_liq else None
        auth = None
        if col_auth:
            v = r.get(col_auth)
            if v is not None and str(v).strip():
                auth = str(v).strip()
        recs.append((d, bruto, liq, auth, r.to_json(force_ascii=False)))
    return recs


def _to_records_receb(df: pd.DataFrame, col_sale_dt: str, col_pay_dt: str | None,
                      col_bruto: str | None, col_liq: str | None, col_auth: str | None):
    """Recebimentos: preserva dt (data da venda) e captura pay_dt (data de pagamento), quando houver."""
    recs = []
    for _, r in df.iterrows():
        sale_d = parse_date(r.get(col_sale_dt))
        if not sale_d:
            continue
        pay_d = parse_date(r.get(col_pay_dt)) if col_pay_dt else None
        bruto = parse_decimal(r.get(col_bruto)) if col_bruto else None
        liq = parse_decimal(r.get(col_liq)) if col_liq else None
        auth = None
        if col_auth:
            v = r.get(col_auth)
            if v is not None and str(v).strip():
                auth = str(v).strip()
        recs.append((sale_d, pay_d, bruto, liq, auth, r.to_json(force_ascii=False)))
    return recs

def _drop_dupes_keep_first(recs: list[tuple], auth_index: int = 3) -> tuple[list[tuple], int]:
    seen = set()
    kept = []
    dropped = 0
    for rec in recs:
        a = rec[auth_index]
        if a:
            if a in seen:
                dropped += 1
                continue
            seen.add(a)
        kept.append(rec)
    return kept, dropped

def _row_exists(conn: sqlite3.Connection, table: str, provider: str, dt_iso: str, bruto_s: str | None, liq_s: str | None, auth: str | None, amount_s: str | None = None) -> bool:
    if table in ("erp_tx","sales_tx","receb_tx"):
        if auth:
            row = conn.execute(f"SELECT 1 FROM {table} WHERE provider=? AND is_deleted=0 AND autorizacao=? LIMIT 1", (provider, auth)).fetchone()
            return row is not None
        row = conn.execute(
            f"""SELECT 1 FROM {table}
                WHERE provider=? AND is_deleted=0 AND dt=?
                  AND COALESCE(bruto,'')=COALESCE(?, '')
                  AND COALESCE(liquido,'')=COALESCE(?, '')
                  AND (autorizacao IS NULL OR autorizacao='' )
                LIMIT 1""",
            (provider, dt_iso, bruto_s, liq_s)
        ).fetchone()
        return row is not None
    if table == "bank_tx":
        row = conn.execute(
            """SELECT 1 FROM bank_tx
               WHERE provider=? AND is_deleted=0 AND dt=? AND COALESCE(amount,'')=COALESCE(?, '')
               LIMIT 1""",
            (provider, dt_iso, amount_s)
        ).fetchone()
        return row is not None
    return False

def import_erp_alelo(conn: sqlite3.Connection, path: str, provider: str = "ALELO") -> ImportResult:
    warnings, errors = [], []
    inserted = dropped_total = 0
    try:
        df = _read_excel_best_sheet(path, prefer_sheets=["Planilha1","ERP","Dados"])
    except Exception as e:
        return ImportResult(0,0,[],[f"Falha ao abrir Excel ERP: {e}"])
    if df.empty:
        return ImportResult(0,0,["Arquivo vazio ou sem abas legíveis."],[])
    df.columns = [str(c).strip() for c in df.columns]
    cols = _detect_columns(df, {
        
        "dt": ["data_do_faturamento","data_faturamento","data_da_venda","data_venda","data_emissao","emissao","data","dt"],
        "pay_dt": ["data_de_pagamento","data_do_pagamento","data_pagamento","pagamento","data_pagto","data_de_credito_debito","data_credito_debito","data_de_credito_debito_"],
        "bruto": ["valor_bruto","vl_bruto","bruto","valor_da_transacao","valor_transacao","valor","total"],
        "liquido": ["valor_liquido","vl_liq","vl_liquido","liquido","total_liquido","tot_c_desconto","totc_desconto","total_com_desconto","total_c_desconto"],
        "autorizacao": ["numero_da_autorizacao","n_da_autorizacao","autorizacao","nsu","n_reembolso","numero_do_reembolso","numero_reembolso"]
    
    })
    if cols["dt"] is None:
        return ImportResult(0,0,[f"Não consegui identificar a coluna de data no ERP. Colunas: {_log_columns_hint(df)}"],[])
    recs = _to_records_receb(df, cols["dt"], cols["pay_dt"], cols["bruto"], cols["liquido"], cols["autorizacao"])
    recs, dropped_mem = _drop_dupes_keep_first(recs, auth_index=4)
    cur = conn.cursor()
    for d, pay_d, bruto, liq, auth, raw in recs:
        dt_iso = d.isoformat()
        bruto_s = str(bruto) if bruto is not None else None
        liq_s = str(liq) if liq is not None else None
        if _row_exists(conn, "erp_tx", provider, dt_iso, bruto_s, liq_s, auth):
            dropped_total += 1
            continue
        cur.execute("INSERT INTO erp_tx(provider, dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?)",
                (provider, dt_iso, bruto_s, liq_s, auth, raw))
        inserted += 1
    dropped_total += dropped_mem
    cur.execute("INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
                (provider, "ERP", path, datetime.now().isoformat(sep=" ", timespec="seconds")))
    conn.commit()
    return ImportResult(inserted, dropped_total, warnings, errors)

def import_sales_alelo(conn: sqlite3.Connection, path: str, provider: str = "ALELO") -> ImportResult:
    warnings, errors = [], []
    inserted = dropped_total = 0
    try:
        df = _read_excel_best_sheet(path, prefer_sheets=["Extrato","Vendas"])
    except Exception as e:
        return ImportResult(0,0,[],[f"Falha ao abrir Excel Vendas: {e}"])
    if df.empty:
        return ImportResult(0,0,["Arquivo vazio ou sem abas legíveis."],[])
    df.columns = [str(c).strip() for c in df.columns]
    cols = _detect_columns(df, {
        "dt": ["data_da_venda","data_venda","data_da_transacao","data_transacao","data","dt"],
        "bruto": ["valor_bruto","vl_bruto","bruto","valor"],
        "liquido": ["valor_liquido","vl_liq","vl_liquido","liquido"],
        "autorizacao": ["numero_da_autorizacao","n_da_autorizacao","autorizacao","nsu","numero_autorizacao"]
    })
    if cols["dt"] is None:
        return ImportResult(0,0,["Não consegui identificar a coluna de data no relatório de Vendas (Portal)."],[])
    recs = _to_records(df, cols["dt"], cols["bruto"], cols["liquido"], cols["autorizacao"])
    recs, dropped_mem = _drop_dupes_keep_first(recs)
    cur = conn.cursor()
    for d, bruto, liq, auth, raw in recs:
        dt_iso = d.isoformat()
        bruto_s = str(bruto) if bruto is not None else None
        liq_s = str(liq) if liq is not None else None
        if _row_exists(conn, "sales_tx", provider, dt_iso, bruto_s, liq_s, auth):
            dropped_total += 1
            continue
        cur.execute("INSERT INTO sales_tx(provider, dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?)",
                    (provider, dt_iso, bruto_s, liq_s, auth, raw))
        inserted += 1
    dropped_total += dropped_mem
    cur.execute("INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
                (provider, "SALES", path, datetime.now().isoformat(sep=" ", timespec="seconds")))
    conn.commit()
    return ImportResult(inserted, dropped_total, warnings, errors)

def import_receb_alelo(conn: sqlite3.Connection, path: str, provider: str = "ALELO") -> ImportResult:
    warnings, errors = [], []
    inserted = dropped_total = 0
    try:
        df = _read_excel_best_sheet(path, prefer_sheets=["Recebimentos","Recebimento","Receb"])
    except Exception as e:
        return ImportResult(0,0,[],[f"Falha ao abrir Excel Recebimentos: {e}"])
    if df.empty:
        return ImportResult(0,0,["Arquivo vazio ou sem abas legíveis."],[])
    df.columns = [str(c).strip() for c in df.columns]
    cols = _detect_columns(df, {
        
        "dt": ["data_da_venda","data_venda","data_da_transacao","data_transacao","data","dt"],
        "pay_dt": ["data_de_pagamento","data_do_pagamento","data_pagamento","pagamento","data_pagto","data_de_credito_debito","data_credito_debito","data_de_credito_debito_"],
        "bruto": ["valor_bruto","vl_bruto","bruto","valor_da_transacao","valor_transacao","valor","total"],
        "liquido": ["valor_liquido","vl_liq","vl_liquido","liquido","total_liquido"],
        "autorizacao": ["numero_da_autorizacao","n_da_autorizacao","autorizacao","nsu","n_reembolso","numero_do_reembolso","numero_reembolso"]
    
    })
    if cols["dt"] is None:
        return ImportResult(0,0,[f"Não consegui identificar a coluna de data no relatório de Recebimentos (Portal). Colunas: {_log_columns_hint(df)}"],[])
    # Recebimentos: dt = Data da Venda (para Etapa 2), pay_dt = Data de Pagamento (para Etapa 3)
    recs = _to_records_receb(df, cols["dt"], cols.get("pay_dt"), cols["bruto"], cols["liquido"], cols["autorizacao"])
    recs, dropped_mem = _drop_dupes_keep_first(recs, auth_index=4)
    cur = conn.cursor()
    for sale_d, pay_d, bruto, liq, auth, raw in recs:
        dt_iso = sale_d.isoformat()
        pay_dt_iso = pay_d.isoformat() if pay_d else None
        bruto_s = str(bruto) if bruto is not None else None
        liq_s = str(liq) if liq is not None else None
        if _row_exists(conn, "receb_tx", provider, dt_iso, bruto_s, liq_s, auth):
            dropped_total += 1
            continue
        cur.execute("INSERT INTO receb_tx(provider, dt, pay_dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?,?)",
                    (provider, dt_iso, pay_dt_iso, bruto_s, liq_s, auth, raw))
        inserted += 1
    dropped_total += dropped_mem
    cur.execute("INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
                (provider, "RECEB", path, datetime.now().isoformat(sep=" ", timespec="seconds")))
    conn.commit()
    return ImportResult(inserted, dropped_total, warnings, errors)


# =========================
# Providers: TICKET (Portal)
# =========================

TICKET_VENDAS_REQUIRED_COLS = {"Data da Transação", "Nº Transação", "Vl Transação", "Nº Reembolso"}
TICKET_REEMB_REQUIRED_COLS = {
    "Número do reembolso",
    "Data de corte",
    "Data de crédito/débito",
    "Data da transação",
    "Descrição do lançamento",
    "Valor da transação",
}

def _ticket_safe_float_brl(x) -> float:
    try:
        if pd.isna(x):
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return 0.0
        s = s.replace("R$", "").replace("\u00a0", " ").strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0.0

def _ticket_auth_key(reembolso_id: str, sale_dt: date | None) -> str:
    rid = (reembolso_id or "").strip()
    d = sale_dt.isoformat() if sale_dt else "NA"
    return f"{rid}|{d}"

def import_sales_ticket(conn: sqlite3.Connection, path: str, provider: str = "TICKET") -> ImportResult:
    """Importa 'Vendas' do portal Ticket (Relatório de Vendas).

    ✅ Anti-duplicidade (mesma planilha / mesma importação):
      - remove duplicatas por (dt, Nº Reembolso, Nº Transação) antes de agrupar.
      - isso evita dobrar valor quando o portal repete a mesma transação na exportação.

    Estratégia (compatível com o Concilia):
      - agrupa por (Data da Transação, Nº Reembolso) => 1 linha por reembolso/dia em sales_tx
      - autorizacao = f"<reembolso>|<data_transacao>"
      - bruto = soma do Vl Transação no grupo
      - liquido = bruto (Ticket vendas não traz líquido)
    """
    warnings, errors = [], []
    inserted = dropped_total = 0
    try:
        df = pd.read_excel(path, skiprows=8, dtype=object)
    except Exception as e:
        return ImportResult(0,0,[],[f"Falha ao abrir Excel Vendas (Ticket): {e}"])
    if df is None or df.empty:
        return ImportResult(0,0,["Arquivo vazio ou sem abas legíveis."],[])

    cols = set([str(c).strip() for c in df.columns])
    missing = TICKET_VENDAS_REQUIRED_COLS - cols
    if missing:
        return ImportResult(0,0,[],[f"Ticket Vendas: colunas obrigatórias ausentes: {sorted(list(missing))}"])

    df = df[list(TICKET_VENDAS_REQUIRED_COLS)].copy()
    df["Data da Transação"] = pd.to_datetime(df["Data da Transação"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["Data da Transação"]).copy()
    df["dt"] = df["Data da Transação"].dt.date

    df["Nº Reembolso"] = df["Nº Reembolso"].astype("string").fillna("").str.strip()
    df["Nº Transação"] = df["Nº Transação"].astype("string").fillna("").str.strip()
    df["Vl Transação"] = df["Vl Transação"].apply(_ticket_safe_float_brl)

    # filtra linhas inválidas
    df = df[(df["Nº Reembolso"] != "") & (df["Nº Transação"] != "") & (df["Vl Transação"] > 0)].copy()

    # ✅ remove duplicatas dentro da planilha (mesmo reembolso + transação no mesmo dia)
    before = len(df)
    df = df.drop_duplicates(subset=["dt","Nº Reembolso","Nº Transação"], keep="first").copy()
    dropped_plan = before - len(df)
    if dropped_plan:
        warnings.append(f"Ticket Vendas: {dropped_plan} duplicados removidos (mesmo reembolso+transação).")

    # agrupa por reembolso/dia (preserva lista de NSU / Nº Transação para auditoria)
    g = df.groupby(["dt","Nº Reembolso"], as_index=False).agg({
        "Vl Transação":"sum",
        "Nº Transação": lambda s: list(pd.unique(s.astype("string").fillna("").str.strip()))
    })
    cur = conn.cursor()
    for _, r in g.iterrows():
        d: date = r["dt"]
        rid = str(r["Nº Reembolso"])
        amt = Decimal(str(r["Vl Transação"])).quantize(Decimal("0.01"))
        auth = _ticket_auth_key(rid, d)
        dt_iso = d.isoformat()
        bruto_s = str(amt)
        liq_s = str(amt)
        raw = json.dumps({"reembolso_id": rid, "nsu_list": (r.get("Nº Transação") or []), "qtd_transacoes": int(len((r.get("Nº Transação") or []))), "fonte": "Ticket Vendas"}, ensure_ascii=False)
        if _row_exists(conn, "sales_tx", provider, dt_iso, bruto_s, liq_s, auth):
            dropped_total += 1
            continue
        cur.execute("INSERT INTO sales_tx(provider, dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?)",
                    (provider, dt_iso, bruto_s, liq_s, auth, raw))
        inserted += 1

    cur.execute("INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
                (provider, "SALES", path, datetime.now().isoformat(sep=" ", timespec="seconds")))
    conn.commit()
    return ImportResult(inserted, dropped_total, warnings, errors)

def import_receb_ticket(conn: sqlite3.Connection, path: str, provider: str = "TICKET") -> ImportResult:
    """Importa 'Reembolsos/Recebimentos' do portal Ticket.

    ✅ ROTA NOVA (Ticket):
      - BRUTO (Etapa 2): soma das linhas COMPRA por (reembolso_id + data_da_transacao) -> 1 registro por dia.
      - LÍQUIDO (Etapa 2): calculado a partir do cadastro em Taxas/Tarifas (fee_rules_tx):
            liq = bruto - (bruto * mdr_percent/100) - (qtd_trans * fee_fixed)
        • mdr_percent = % Adm (ex.: 4,5)
        • fee_fixed   = tarifa por transação (ex.: 0,52)
      - PAY_DT (Etapa 3): vem da coluna "Data de crédito/débito" (1 por reembolso) e é gravada em receb_tx.
      - A taxa de transferência (transfer_fee, ex.: 8,30) NÃO entra na Etapa 2.
        Ela será subtraída na Etapa 3 por lote (1x por reembolso).

    ✅ Anti-duplicidade:
      - autorizacao = f"<reembolso>|<data_transacao>" (um registro por reembolso+dia)
    """
    warnings, errors = [], []
    inserted = dropped_total = 0

    prov = (provider or "").strip().upper() or "TICKET"

    # pega taxas do cadastro (fallback seguro)
    rules = fee_rules_tx_list(conn, prov)
    rule = fee_rule_tx_match(rules, "VOUCHER") if rules else None
    if rule is None and rules:
        # fallback: primeira regra ativa
        rule = next((r for r in rules if r.get("is_active")), None)

    mdr_percent = Decimal(str(rule.get("mdr_percent", 4.5))) if rule else Decimal("4.5")
    fee_fixed = Decimal(str(rule.get("fee_fixed", 0.52))) if rule else Decimal("0.52")

    try:
        df = pd.read_excel(path, skiprows=13, dtype=object)
    except Exception as e:
        return ImportResult(0, 0, [], [f"Falha ao abrir Excel Recebimentos (Ticket): {e}"])
    if df is None or df.empty:
        return ImportResult(0, 0, ["Arquivo vazio ou sem abas legíveis."], [])

    cols = set([str(c).strip() for c in df.columns])
    missing = TICKET_REEMB_REQUIRED_COLS - cols
    if missing:
        return ImportResult(0, 0, [], [f"Ticket Reembolso: colunas obrigatórias ausentes: {sorted(list(missing))}"])

    df = df[list(TICKET_REEMB_REQUIRED_COLS)].copy()
    df["Número do reembolso"] = df["Número do reembolso"].astype("string").fillna("").str.strip()
    df.loc[df["Número do reembolso"] == "", "Número do reembolso"] = pd.NA
    df["Número do reembolso"] = df["Número do reembolso"].ffill()
    df = df.dropna(subset=["Número do reembolso"]).copy()
    df["reembolso_id"] = df["Número do reembolso"].astype("string").str.strip()

    # datas
    for c in ["Data de corte", "Data de crédito/débito", "Data da transação"]:
        df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

    df["Descrição do lançamento"] = df["Descrição do lançamento"].astype("string").fillna("").str.strip()
    df["Valor da transação"] = df["Valor da transação"].apply(_ticket_safe_float_brl)

    cur = conn.cursor()

    for rid, g in df.groupby("reembolso_id", dropna=False):
        rid = str(rid).strip()
        if not rid:
            continue

        # data de pagamento (1 por lote)
        dcred = g["Data de crédito/débito"].dropna()
        pay_dt = dcred.iloc[0].date() if len(dcred) else None

        # linhas COMPRA
        compras = g[g["Descrição do lançamento"].str.fullmatch(r"COMPRA", case=False, na=False)].copy()
        compras = compras.dropna(subset=["Data da transação"]).copy()
        if compras.empty:
            compras = g[g["Descrição do lançamento"].str.contains(r"\bCOMPRA\b", case=False, na=False)].copy()
            compras = compras.dropna(subset=["Data da transação"]).copy()

        if compras.empty:
            warnings.append(f"Ticket Reembolso {rid}: não encontrei linhas 'COMPRA' para gerar Etapa 2.")
            continue

        compras["sale_dt"] = compras["Data da transação"].dt.date
        compras["bruto"] = compras["Valor da transação"].astype(float)

        # 1 registro por reembolso+dia: soma bruto + conta qtd transações do dia
        by_day = (
            compras.groupby("sale_dt", as_index=False)
            .agg(bruto_sum=("bruto", "sum"), qtd=("bruto", "count"))
        )

        for _, rr in by_day.iterrows():
            sale_dt: date = rr["sale_dt"]
            bruto_d = Decimal(str(rr["bruto_sum"] or 0)).quantize(Decimal("0.01"))
            qtd_d = int(rr["qtd"] or 0)

            # calcula taxas pelo cadastro
            taxa_adm = (bruto_d * (mdr_percent / Decimal("100"))).quantize(Decimal("0.01"))
            tarifa_tot = (fee_fixed * Decimal(str(qtd_d))).quantize(Decimal("0.01"))
            taxa_total = (taxa_adm + tarifa_tot).quantize(Decimal("0.01"))
            liq_d = (bruto_d - taxa_total).quantize(Decimal("0.01"))

            auth = _ticket_auth_key(rid, sale_dt)
            dt_iso = sale_dt.isoformat()
            pay_iso = pay_dt.isoformat() if pay_dt else None

            bruto_s = str(bruto_d)
            liq_s = str(liq_d)

            raw = json.dumps({
                "reembolso_id": rid,
                "pay_dt": pay_iso,
                "qtd_transacoes_dia": qtd_d,
                "mdr_percent": float(mdr_percent),
                "fee_fixed": float(fee_fixed),
                "taxa_adm": float(taxa_adm),
                "tarifa_total": float(tarifa_tot),
                "taxa_total": float(taxa_total),
                "fonte": "Ticket Reembolso (rota_nova)"
            }, ensure_ascii=False)

            if _row_exists(conn, "receb_tx", prov, dt_iso, bruto_s, liq_s, auth):
                dropped_total += 1
                continue

            cur.execute(
                "INSERT INTO receb_tx(provider, dt, pay_dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?,?)",
                (prov, dt_iso, pay_iso, bruto_s, liq_s, auth, raw),
            )
            inserted += 1

    cur.execute(
        "INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
        (prov, "RECEB", path, datetime.now().isoformat(sep=" ", timespec="seconds")),
    )
    conn.commit()
    return ImportResult(inserted, dropped_total, warnings, errors)


# ==========================================================
# FARMACIASAPP - imports (ERP / Recebimentos / Vendas)
# ==========================================================
def import_erp_farmaciasapp(conn: sqlite3.Connection, path: str, provider: str = "FARMACIASAPP") -> ImportResult:
    """Importa ERP para FarmaciasAPP.

    Layout esperado (colunas):
      - Emissão (data da venda)
      - Operadora (filtra FARMACIAS_APP)
      - Total (valor pago na loja)
    """
    prov = (provider or "FARMACIASAPP").strip().upper()
    warnings, errors = [], []
    inserted = dropped = 0

    try:
        df = _read_excel_best_sheet(path, prefer_sheets=["Planilha1","ERP","Dados","Sheet1"]) 
    except Exception as e:
        return ImportResult(0, 0, [], [f"Falha ao ler ERP FarmaciasAPP: {e}"])

    # normaliza colunas
    cols = {normalize_text(c): c for c in df.columns}
    c_dt = cols.get("EMISSAO") or cols.get("EMISSAO")
    c_op = cols.get("OPERADORA")
    c_total = cols.get("TOTAL")

    if not c_dt or not c_op or not c_total:
        return ImportResult(0, 0, [], [f"ERP FarmaciasAPP: não encontrei colunas obrigatórias (Emissão/Operadora/Total). Colunas: {list(df.columns)[:20]}"])

    # filtra operadora
    op_txt = df[c_op].astype(str).str.upper()
    df = df[op_txt.str.contains("FARMACIAS_APP", na=False)].copy()
    if df.empty:
        warnings.append("ERP FarmaciasAPP: nenhum registro com Operadora=FARMACIAS_APP.")
        return ImportResult(0, 0, warnings, errors)

    cur = conn.cursor()
    for _, r in df.iterrows():
        dt = parse_date(r.get(c_dt))
        if not dt:
            continue
        bruto = parse_decimal(r.get(c_total)) or Decimal("0")
        if bruto == 0:
            continue
        # No ERP FarmaciasAPP não há NSU -> autorizacao vazio (match por janela + valor)
        auth = ""
        dt_iso = dt.isoformat()
        bruto_s = str(bruto)
        liq_s = str(bruto)

        raw = json.dumps({
            "fonte": "ERP FarmaciasAPP",
            "emissao": dt_iso,
            "operadora": str(r.get(c_op) or "").strip(),
            "valor_pago_loja": float(bruto),
        }, ensure_ascii=False)

        if _row_exists(conn, "erp_tx", prov, dt_iso, bruto_s, liq_s, auth):
            dropped += 1
            continue

        cur.execute(
            "INSERT INTO erp_tx(provider, dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?)",
            (prov, dt_iso, bruto_s, liq_s, auth, raw),
        )
        inserted += 1

    cur.execute(
        "INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
        (prov, "ERP", path, datetime.now().isoformat(sep=" ", timespec="seconds")),
    )
    conn.commit()
    return ImportResult(inserted, dropped, warnings, errors)


def _farmapp_detect_receb_layout(df: pd.DataFrame) -> str:
    """Detecta layout do recebimento FarmaciasAPP (OTB vs ZOOP).
    Nota: normalize_text() no app retorna normalmente string normalizada em MAIÚSCULO.
    """
    cols_n = {normalize_text(c) for c in getattr(df, "columns", [])}
    # OTB típico: "Numero Pedido", "Data do Repasse", etc.
    if ("NUMERO PEDIDO" in cols_n) or ("NÚMERO PEDIDO" in cols_n) or ("NUMERO_PEDIDO" in cols_n):
        return "OTB"
    # ZOOP típico: "Pedido" + "Valor Total" + "Valor de Repasse"
    if ("PEDIDO" in cols_n) and ("VALOR TOTAL" in cols_n):
        return "ZOOP"
    # fallback: se não detectou, assume OTB (é o mais comum) mas somente se houver colunas
    return "OTB"

def _farmapp_locate_otb_sheet_and_header(path: str, max_rows: int = 12):
    """Tenta localizar em qual aba e em qual linha está o cabeçalho do OTB.
    Retorna (sheet_name, header_row_index) onde header_row_index é 0-based para pandas.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return None, None

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, None

    # chaves que identificam o header OTB
    must = {"NUMERO PEDIDO", "DATA PEDIDO", "VALOR DE REPASSE"}
    for sname in wb.sheetnames:
        try:
            ws = wb[sname]
            for r in range(1, max_rows + 1):
                row_vals = []
                for c in range(1, 18):  # até coluna R
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    row_vals.append(normalize_text(str(v)))
                if not row_vals:
                    continue
                if must.issubset(set(row_vals)):
                    # header achado na linha r
                    return sname, r - 1
        except Exception:
            continue
    return None, None

def import_receb_farmaciasapp(conn: sqlite3.Connection, path: str, provider: str = "FARMACIASAPP") -> ImportResult:
    """Importa Recebimentos do portal FarmaciasAPP.

    Suporta dois layouts:
      - OTB: colunas (Numero Pedido, Data Pedido, Forma de Pagamento, Valor Total, Frete, Subsidio de Desconto,
                     Taxa de Servico, Valor de Repasse, Data do Repasse, Status)
      - ZOOP: colunas (Data Pedido, Pedido, Valor Total, Frete, Subsidio de Desconto, Taxa de Servico, Valor de Repasse,
                      Data, Status)  (podem variar; ignoramos extras)
    """
    prov = (provider or "FARMACIASAPP").strip().upper()
    warnings, errors = [], []
    inserted = dropped = 0

    df = None
    # 1) Primeiro tenta OTB de forma determinística (aba/linha do header), porque o layout é padronizado
    sheet_otb, header_otb = _farmapp_locate_otb_sheet_and_header(path)
    if sheet_otb is not None and header_otb is not None:
        try:
            df = pd.read_excel(path, sheet_name=sheet_otb, header=header_otb)
            warnings.append(f"Receb FarmaciasAPP: OTB detectado em aba='{sheet_otb}' header_linha={header_otb+1}.")
        except Exception as e:
            warnings.append(f"Receb FarmaciasAPP: falha lendo OTB aba='{sheet_otb}' header_linha={header_otb+1}: {e}")

    # 2) Se não conseguiu, usa o helper genérico
    if df is None:
        try:
            df = _read_excel_best_sheet(path, prefer_sheets=["Sheet1","OTB","ZOOP","Recebimentos","Recebimento","Extrato"])
        except Exception as e:
            return ImportResult(0, 0, [], [f"Falha ao ler Recebimentos FarmaciasAPP: {e}"])

    # fallback: alguns exports vêm com header deslocado/linha mesclada; tenta leitura direta da 1ª aba
    if df is None or len(getattr(df, "columns", [])) == 0:
        for h in (0, 1, 2):
            try:
                df_try = pd.read_excel(path, sheet_name=0, header=h)
                if df_try is not None and len(getattr(df_try, "columns", [])) > 0:
                    df = df_try
                    break
            except Exception:
                continue

    # alguns arquivos podem vir com "Unnamed" nas colunas quando o header está uma linha abaixo
    if df is not None and df.columns.tolist() and str(df.columns[0]).startswith("Unnamed"):
        try:
            df_alt = pd.read_excel(path, sheet_name=0, header=1)
            if df_alt is not None and len(getattr(df_alt, "columns", [])) > 0:
                df = df_alt
        except Exception:
            pass

    if df is None or len(getattr(df, "columns", [])) == 0:
        return ImportResult(0, 0, warnings, [f"Receb FarmaciasAPP: não consegui ler colunas do arquivo. Abas/header podem estar fora do padrão."])

    layout = _farmapp_detect_receb_layout(df)

    # map colunas por normalize_text
    cmap = {normalize_text(c): c for c in df.columns}

    def col(*names):
        for n in names:
            k = normalize_text(n)
            if k in cmap:
                return cmap[k]
        return None

    if layout == "OTB":
        c_pedido = col("Numero Pedido", "Número Pedido")
        c_dt_ped = col("Data Pedido")
        c_forma = col("Forma de Pagamento", "Forma Pagamento")
        c_total = col("Valor Total")
        c_frete = col("Frete")
        c_sub = col("Subsidio de Desconto", "Subsídio de Desconto", "Subsidio", "Subsídio")
        c_taxa = col("Taxa de Servico", "Taxa de Serviço", "Taxa Servico")
        c_rep = col("Valor de Repasse")
        c_pay = col("Data do Repasse", "Data de Repasse", "Data Repasse")
        c_status = col("Status")
    else:
        # ZOOP
        c_pedido = col("Pedido", "Numero Pedido", "Número Pedido")
        c_dt_ped = col("Data Pedido")
        c_forma = col("Forma de Pagamento", "Forma Pagamento")
        c_total = col("Valor Total")
        c_frete = col("Frete")
        c_sub = col("Subsidio de Desconto", "Subsídio de Desconto", "Subsidio", "Subsídio")
        c_taxa = col("Taxa de Servico", "Taxa de Serviço", "Taxa Servico")
        c_rep = col("Valor de Repasse", "Valor Repasse")
        c_pay = col("Data", "Data do Repasse", "Data de Repasse")
        c_status = col("Status")

    if not c_pedido or not c_dt_ped or not c_total or not c_rep:
        # fallback robusto: detectar linha de header varrendo as primeiras linhas
        def _scan_header_row(path_: str, max_rows: int = 40):
            try:
                raw0 = pd.read_excel(path_, sheet_name=0, header=None, nrows=max_rows, engine="openpyxl")
            except Exception:
                return None
            markers = {
                "NUMERO PEDIDO","NÚMERO PEDIDO","PEDIDO","DATA PEDIDO","VALOR TOTAL","VALOR DE REPASSE",
                "DATA DO REPASSE","DATA","STATUS","SUBSIDIO DE DESCONTO","TAXA DE SERVICO"
            }
            best_i, best_hits = None, 0
            norm_markers = {normalize_text(x) for x in markers}
            for i in range(min(len(raw0), max_rows)):
                row = raw0.iloc[i].astype(str).tolist()
                norms = {normalize_text(x) for x in row if x and str(x).strip() and str(x).strip().lower() != "nan"}
                hits = len(norm_markers & norms)
                if hits > best_hits:
                    best_hits, best_i = hits, i
            if best_i is None or best_hits < 2:
                return None
            try:
                return pd.read_excel(path_, sheet_name=0, header=int(best_i), engine="openpyxl")
            except Exception:
                return None

        df2 = _scan_header_row(path)
        if df2 is not None and len(getattr(df2, "columns", [])) > 0:
            df = df2
            layout = _farmapp_detect_receb_layout(df)
            cmap = {normalize_text(c): c for c in df.columns}

            def col(*names):
                for n in names:
                    k = normalize_text(n)
                    if k in cmap:
                        return cmap[k]
                return None

            if layout == "OTB":
                c_pedido = col("Numero Pedido", "Número Pedido")
                c_dt_ped = col("Data Pedido")
                c_forma = col("Forma de Pagamento", "Forma Pagamento")
                c_total = col("Valor Total")
                c_frete = col("Frete")
                c_sub = col("Subsidio de Desconto", "Subsídio de Desconto", "Subsidio", "Subsídio")
                c_taxa = col("Taxa de Servico", "Taxa de Serviço", "Taxa Servico")
                c_rep = col("Valor de Repasse")
                c_pay = col("Data do Repasse", "Data de Repasse", "Data Repasse")
                c_status = col("Status")
            else:
                c_pedido = col("Pedido", "Numero Pedido", "Número Pedido")
                c_dt_ped = col("Data Pedido")
                c_forma = col("Forma de Pagamento", "Forma Pagamento")
                c_total = col("Valor Total")
                c_frete = col("Frete")
                c_sub = col("Subsidio de Desconto", "Subsídio de Desconto", "Subsidio", "Subsídio")
                c_taxa = col("Taxa de Servico", "Taxa de Serviço", "Taxa Servico")
                c_rep = col("Valor de Repasse", "Valor Repasse")
                c_pay = col("Data", "Data do Repasse", "Data de Repasse")
                c_status = col("Status")

        if not c_pedido or not c_dt_ped or not c_total or not c_rep:
            return ImportResult(0, 0, [], [f"Receb FarmaciasAPP: colunas obrigatórias ausentes. Layout={layout}. Colunas: {list(getattr(df,'columns',[]))[:30]}"])

    cur = conn.cursor()
    for _, r in df.iterrows():
        pedido = str(r.get(c_pedido) or "").strip()
        if not pedido or pedido.lower().startswith("nan"):
            continue

        # status: se cancelado, ignora
        st = str(r.get(c_status) or "").strip().upper() if c_status else ""
        if "CANCEL" in st:
            continue

        dt = parse_date(r.get(c_dt_ped))
        if not dt:
            continue

        bruto_total = parse_decimal(r.get(c_total)) or Decimal("0")
        repasse = parse_decimal(r.get(c_rep)) or Decimal("0")

        if bruto_total == 0:
            continue

        # campos auxiliares
        frete = parse_decimal(r.get(c_frete)) if c_frete else None
        subsidio = parse_decimal(r.get(c_sub)) if c_sub else None
        taxa_serv = parse_decimal(r.get(c_taxa)) if c_taxa else None
        forma = str(r.get(c_forma) or "").strip().upper() if c_forma else ""

        # valor pago na loja (cliente) = total - subsidio (se existir)
        pago_loja = bruto_total
        if subsidio not in (None, Decimal("0")):
            try:
                pago_loja = (bruto_total - (subsidio or Decimal("0"))).quantize(Decimal("0.01"))
            except Exception:
                pago_loja = bruto_total

        pay_dt = parse_date(r.get(c_pay)) if c_pay else None
        dt_iso = dt.isoformat()
        pay_iso = pay_dt.isoformat() if pay_dt else None

        bruto_s = str(bruto_total)
        liq_s = str(repasse)

        raw = json.dumps({
            "fonte": f"Receb FarmaciasAPP ({layout})",
            "pedido": pedido,
            "data_pedido": dt_iso,
            "pay_dt": pay_iso,
            "forma_pagamento": forma,
            "valor_total_compra": float(bruto_total),
            "valor_pago_loja": float(pago_loja),
            "subsidio_desconto": float(subsidio) if subsidio is not None else None,
            "frete": float(frete) if frete is not None else None,
            "taxa_servico": float(taxa_serv) if taxa_serv is not None else None,
        }, ensure_ascii=False)

        auth = pedido  # usamos pedido como chave principal para match (quando disponível)
        if _row_exists(conn, "receb_tx", prov, dt_iso, bruto_s, liq_s, auth):
            dropped += 1
            continue

        cur.execute(
            "INSERT INTO receb_tx(provider, dt, pay_dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?,?)",
            (prov, dt_iso, pay_iso, bruto_s, liq_s, auth, raw),
        )
        inserted += 1

    cur.execute(
        "INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
        (prov, "RECEB", path, datetime.now().isoformat(sep=" ", timespec="seconds")),
    )
    conn.commit()
    return ImportResult(inserted, dropped, warnings, errors)


def import_sales_farmaciasapp(conn: sqlite3.Connection, path: str, provider: str = "FARMACIASAPP") -> ImportResult:
    """Importa Vendas do portal FarmaciasAPP.

    Observação: alguns arquivos podem vir com XML inválido (números com vírgula em célula numérica).
    Se isso acontecer, reexporte como CSV ou Excel corrigido.
    """
    prov = (provider or "FARMACIASAPP").strip().upper()
    warnings, errors = [], []
    inserted = dropped = 0
    try:
        df = _read_excel_best_sheet(path, prefer_sheets=["Sheet1","Vendas","Extrato"]) 
    except Exception as e:
        return ImportResult(0, 0, [], [f"Falha ao ler Vendas FarmaciasAPP: {e}. Sugestão: reexportar o relatório (CSV ou Excel)"])

    cmap = {normalize_text(c): c for c in df.columns}
    c_ped = cmap.get("PEDIDO")
    c_status = cmap.get("STATUS")
    c_total = cmap.get("TOTAL")
    c_dt = cmap.get("DATAHORA") or cmap.get("DATA HORA") or cmap.get("DATAHORA DA VENDA")
    if not c_ped or not c_status or not c_total or not c_dt:
        return ImportResult(0, 0, [], [f"Vendas FarmaciasAPP: colunas obrigatórias ausentes (Pedido/Status/Total/DataHora). Colunas: {list(df.columns)[:20]}"])

    cur = conn.cursor()
    for _, r in df.iterrows():
        st = str(r.get(c_status) or "").strip().upper()
        if "CANCEL" in st:
            continue
        pedido = str(r.get(c_ped) or "").strip()
        if not pedido:
            continue
        dt = parse_date(r.get(c_dt))
        if not dt:
            continue
        bruto = parse_decimal(r.get(c_total)) or Decimal("0")
        if bruto == 0:
            continue
        dt_iso = dt.isoformat()
        bruto_s = str(bruto)
        liq_s = str(bruto)
        auth = pedido
        raw = json.dumps({
            "fonte":"Vendas Portal FarmaciasAPP",
            "pedido": pedido,
            "status": st
        }, ensure_ascii=False)

        if _row_exists(conn, "sales_tx", prov, dt_iso, bruto_s, liq_s, auth):
            dropped += 1
            continue
        cur.execute(
            "INSERT INTO sales_tx(provider, dt, bruto, liquido, autorizacao, raw_json) VALUES (?,?,?,?,?,?)",
            (prov, dt_iso, bruto_s, liq_s, auth, raw),
        )
        inserted += 1

    cur.execute(
        "INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
        (prov, "VENDAS", path, datetime.now().isoformat(sep=" ", timespec="seconds")),
    )
    conn.commit()
    return ImportResult(inserted, dropped, warnings, errors)

def import_sales_by_provider(conn: sqlite3.Connection, path: str, provider: str) -> ImportResult:
    prov = (provider or "").strip().upper() or "ALELO"
    if prov == "TICKET":
        return import_sales_ticket(conn, path, provider="TICKET")
    if provider == "FARMACIASAPP":
        return import_sales_farmaciasapp(conn, path, provider="FARMACIASAPP")
    return import_sales_alelo(conn, path, provider=prov)

def import_receb_by_provider(conn: sqlite3.Connection, path: str, provider: str) -> ImportResult:
    prov = (provider or "").strip().upper() or "ALELO"
    if prov == "TICKET":
        return import_receb_ticket(conn, path, provider="TICKET")
    if provider == "FARMACIASAPP":
        return import_receb_farmaciasapp(conn, path, provider="FARMACIASAPP")
    return import_receb_alelo(conn, path, provider=prov)

def import_erp_by_provider(conn: sqlite3.Connection, path: str, provider: str) -> ImportResult:
    prov = (provider or "").strip().upper() or "ALELO"
    if provider == "FARMACIASAPP":
        return import_erp_farmaciasapp(conn, path, provider="FARMACIASAPP")
    # Ticket usa o mesmo ERP do Alelo (estrutura padrão)
    return import_erp_alelo(conn, path, provider=prov)


def import_bank_ofx(conn: sqlite3.Connection, path: str, provider: str = "BANCO") -> ImportResult:
    """
    Import OFX robusto (Itaú/Bradesco).
    Observação: alguns OFX vêm com ENCODING/CHARSET inconsistentes (ex.: USASCII com acentos),
    o que faz o parser falhar com erro de ASCII. Aqui nós:
      - decodificamos em cp1252/latin1
      - forçamos header para UTF-8
      - sanitizamos para ASCII (remove acentos do memo, mas preserva datas/valores)
    """
    warnings, errors = [], []
    inserted = dropped = 0
    if OfxParser is None:
        return ImportResult(0, 0, [], ["Dependência 'ofxparse' não encontrada. Rode: pip install ofxparse"])

    try:
        raw = Path(path).read_bytes()
    except Exception as e:
        return ImportResult(0, 0, [], [f"Falha ao ler OFX: {e}"])

    def _detect_encoding(raw_bytes: bytes) -> str | None:
        head_txt = raw_bytes[:4096].decode("latin1", errors="ignore")
        m1 = re.search(r"(?im)^\s*charset\s*:\s*([^\r\n]+)", head_txt)
        if m1:
            return m1.group(1).strip()
        m2 = re.search(r"(?im)^\s*encoding\s*:\s*([^\r\n]+)", head_txt)
        if m2:
            return m2.group(1).strip()
        return None

    def _norm_to_python_codec(enc: str | None) -> str:
        if not enc:
            return "cp1252"
        e = enc.strip().lower()
        if e in ("1252", "windows-1252", "cp1252", "ansi"):
            return "cp1252"
        if e in ("latin1", "iso-8859-1", "iso8859-1"):
            return "latin1"
        if e in ("utf-8", "utf8", "unicode"):
            return "utf-8"
        if e in ("usascii", "ascii"):
            return "cp1252"
        return "cp1252"

    def _force_utf8_header(ofx_text: str) -> str:
        lines = ofx_text.splitlines()
        out = []
        in_header = True
        for line in lines:
            if in_header and line.strip().startswith("<"):
                in_header = False
            if in_header:
                if re.match(r"(?i)^\s*encoding\s*:", line):
                    out.append("ENCODING:UTF-8"); continue
                if re.match(r"(?i)^\s*charset\s*:", line):
                    out.append("CHARSET:UTF-8"); continue
            out.append(line)
        return "\n".join(out)

    def _ascii_sanitize(text: str) -> str:
        t = unicodedata.normalize("NFKD", text)
        t = "".join(ch for ch in t if not unicodedata.combining(ch))
        return t.encode("ascii", errors="ignore").decode("ascii", errors="ignore")


    def _extract_tag(block: str, tag: str) -> str:
        # OFX SGML: <TAG>value may include line breaks until next tag
        m = re.search(rf"(?is)<{tag}>(.*?)(?=\n<|</|$)", block)
        if not m:
            return ""
        v = m.group(1)
        # collapse whitespace/newlines
        v = re.sub(r"[\r\n\t]+", " ", v)
        v = re.sub(r"\s{2,}", " ", v).strip()
        return v

    def _build_raw_tx_map(ofx_text: str) -> dict:
        tx_map: dict[str, str] = {}
        # split by STMTTRN blocks
        parts = re.split(r"(?is)<STMTTRN>", ofx_text)
        for part in parts[1:]:
            block = part
            # truncate at end tag if present
            endm = re.search(r"(?is)</STMTTRN>", block)
            if endm:
                block = block[:endm.start()]
            fitid = _extract_tag(block, "FITID")  # FITID should exist
            fitid = fitid.strip()
            if not fitid:
                continue
            memo = _extract_tag(block, "MEMO")
            name = _extract_tag(block, "NAME")
            payee = _extract_tag(block, "PAYEE")
            combined = " | ".join([x for x in [memo, name, payee] if x])
            combined = _ascii_sanitize(combined)
            if combined:
                tx_map[fitid] = combined
        return tx_map
    def _try_parse(raw_bytes: bytes):
        import io
        with io.BytesIO(raw_bytes) as bio:
            return OfxParser.parse(bio)

    try:
        guessed = _norm_to_python_codec(_detect_encoding(raw))
        txt = raw.decode(guessed, errors="ignore")
        txt = _force_utf8_header(txt)
        raw_txt = txt
        raw_tx_map = _build_raw_tx_map(raw_txt)
        txt = _ascii_sanitize(txt)
        ofx = _try_parse(txt.encode("ascii", errors="ignore"))
    except Exception as e:
        errors.append(f"Falha ao importar OFX: {e}")
        return ImportResult(0, 0, [], errors)

    try:
        cur = conn.cursor()
        for acct in getattr(ofx, "accounts", []):
            stmt = getattr(acct, "statement", None)
            if stmt is None:
                continue
            for tx in getattr(stmt, "transactions", []):
                d = tx.date.date() if isinstance(tx.date, datetime) else tx.date
                dt_iso = d.isoformat()
                amt = Decimal(str(tx.amount))
                amount_s = _bank_amt2_str(amt)
                # Compose memo with parser fields (some OFX provide counterparty in NAME/PAYEE)
                parts = []
                for fld in ("memo", "name", "payee"):
                    val = getattr(tx, fld, None)
                    if val is None and fld == "memo":
                        val = getattr(tx, "memo", None)
                    s = str(val).strip() if val is not None else ""
                    if s:
                        parts.append(s)
                memo = " | ".join(parts).strip(" |")
                # Caixa (e alguns OFX): recover dropped multiline description from raw OFX by FITID
                key = (getattr(tx, "id", None) or getattr(tx, "fitid", None) or "").strip()
                if key:
                    extra = raw_tx_map.get(key)
                    if extra:
                        try:
                            if extra.upper() not in (memo or "").upper():
                                memo = (memo + " | " + extra).strip(" |")
                        except Exception:
                            pass
                memo = _ascii_sanitize(memo)
                bank_id = (getattr(tx, "id", None) or "").strip()
                if (_bank_row_exists_transfer(conn, provider, dt_iso, amount_s, memo) if _bank_is_transfer_memo(memo) else _row_exists(conn, "bank_tx", provider, dt_iso, None, None, None, amount_s=amount_s)):
                    dropped += 1
                    continue
                cur.execute(
                    "INSERT INTO bank_tx(provider, dt, amount, memo, bank_id, raw_json) VALUES (?,?,?,?,?,?)",
                    (provider, dt_iso, amount_s, memo, bank_id, None),
                )
                inserted += 1

        cur.execute(
            "INSERT INTO imports(provider, kind, source_path, imported_at) VALUES (?,?,?,?)",
            (provider, "BANK", path, datetime.now().isoformat(sep=" ", timespec="seconds")),
        )
        conn.commit()
    except Exception as e:
        errors.append(f"Falha ao salvar transações OFX no banco: {e}")

    return ImportResult(inserted, dropped, warnings, errors)


def _pdf_extract_text(path: str) -> str:
    """Extract text from PDF without OCR (works when PDF has selectable text)."""
    try:
        import PyPDF2  # type: ignore
    except Exception as e:
        raise RuntimeError("PyPDF2 não disponível para leitura de PDF.") from e
    reader = PyPDF2.PdfReader(path)
    out = []
    for p in reader.pages:
        try:
            out.append(p.extract_text() or "")
        except Exception:
            out.append("")
    return "\n".join(out)

def _parse_money_ptbr(s: str) -> Optional[Decimal]:
    if not s:
        return None
    s = s.strip().replace("\xa0", " ")
    m = re.search(r'R\$\s*([0-9\.\,]+)', s)
    if not m:
        return None
    return parse_decimal(m.group(1))

def _parse_caixa_pdf_transactions(text: str) -> list[dict]:
    """
    Parser robusto para Extrato PDF Caixa (texto selecionável).

    Já vimos dois layouts comuns no PDF extraído:
      A) (data sozinha)
         "02/01/2025"
         "02/01 15:34000341CRED PAG0108R2 DIF TITULA"
         "TICKET SERVICOS ... R$ 13,99 R$ 3.710,19 D"

      B) (data + histórico na mesma linha)
         "02/01/2025 CRED PAG0108R2 DIF TITULA"
         "000341 R$ 13,99 R$ 3.710,19 D"
         "02/01 15:34 TICKET SERVICOS ..."

    Regras:
      - Valor do lançamento = primeiro valor monetário encontrado no bloco da transação.
      - Sinal do valor:
          • negativo se o próprio valor vier com "-" OU se palavras-chave de débito aparecerem.
          • NÃO usar o "C/D" do saldo (isso é do SALDO da conta, não do lançamento).
      - "SALDO DIA ..." não pode fazer a transação anterior sumir.
      - bank_id no PDF nem sempre é único (ex.: "000341" repete). Geramos um ID estável por (doc+data+índice).
    """
    if not text:
        return []

    raw_lines = [re.sub(r"\s+", " ", (ln or "").strip()) for ln in text.splitlines()]
    lines = [ln for ln in raw_lines if ln]

    debit_kw = ("DEBITO", "PAGTO", "PAGAMENTO", "TARIFA", "IOF", "JUROS", "COBRANCA", "DEB ")
    txs: list[dict] = []

    i = 0
    while i < len(lines):
        ln = lines[i]
        up = ln.upper()

        # pula cabeçalhos
        if "EXTRATO" in up or up.startswith("DATA ") or "CNPJ" in up or "AGÊNCIA" in up:
            i += 1
            continue
        if "SALDO DIA" in up:
            i += 1
            continue

        mdate = re.match(r"^(\d{2}/\d{2}/\d{4})(?:\s+(.*))?$", ln)
        if not mdate:
            i += 1
            continue

        dt = mdate.group(1)
        hist_part = (mdate.group(2) or "").strip().strip(" -")

        # precisamos capturar: doc, hist, payee (opcional) e valor
        doc = ""
        hist = ""
        payee = ""
        money_block = ""

        # --- Caso A: data sozinha -> próxima linha tem efetiva+doc+hist; linha seguinte tem payee+valores
        if not hist_part and i + 2 < len(lines):
            ln2 = lines[i + 1]
            ln3 = lines[i + 2]
            # ignora quando a "linha 2" for SALDO DIA por algum motivo
            if "SALDO DIA" in ln2.upper():
                i += 1
                continue
            # ex: "02/01 15:34000341CRED PAG0108R2 DIF TITULA"
            m2 = re.match(r"^(\d{2}/\d{2})(\d{2}:\d{2})(\d{3,20})(.*)$", ln2.replace(" ", ""))
            if m2:
                doc = (m2.group(3) or "").strip()
                hist = (m2.group(4) or "").strip()
                # às vezes os valores já vêm na própria ln2 (ex.: IOF/JUROS); outras vezes vêm na ln3 (favorecido + valores)
                if "R$" in ln2:
                    money_block = ln2
                    # payee pode estar na ln3
                    if i + 2 < len(lines):
                        if "SALDO DIA" not in ln3.upper() and "EXTRATO" not in ln3.upper():
                            payee = ln3.strip()
                else:
                    money_block = ln3
                    # payee = texto antes do primeiro "R$" (se houver)
                    if "R$" in ln3:
                        payee = ln3.split("R$", 1)[0].strip()
                    else:
                        payee = ln3.strip()

        # --- Caso B: data+hist -> linha seguinte tem doc+valores; linha seguinte pode ter efetiva+favorecido
        if hist_part and not doc and i + 1 < len(lines):
            ln2 = lines[i + 1]
            if "SALDO DIA" in ln2.upper():
                i += 1
                continue
            mdoc = re.match(r"^(\d{1,20})\s+(.*)$", ln2)
            if mdoc:
                doc = (mdoc.group(1) or "").strip()
                hist = hist_part or (mdoc.group(2) or "").strip()
                money_block = ln2
                # payee pode vir na linha seguinte (dd/mm hh:mm ...)
                if i + 2 < len(lines):
                    ln3 = lines[i + 2]
                    mp = re.match(r"^(\d{2}/\d{2})\s+(\d{2}:\d{2})\s+(.*)$", ln3)
                    if mp:
                        payee = (mp.group(3) or "").strip()

        # se não pegou doc/hist, tenta um fallback leve: procurar doc+hist na linha seguinte sem efetiva
        if not doc and i + 1 < len(lines):
            ln2 = lines[i + 1]
            mdoc2 = re.match(r"^(\d{3,20})(.*)$", ln2.replace(" ", ""))
            if mdoc2 and "R$" not in ln2:
                # pode ser doc+hist na linha 2 e valores na linha 3
                doc = (mdoc2.group(1) or "").strip()
                hist = (mdoc2.group(2) or "").strip()
                if i + 2 < len(lines):
                    money_block = lines[i + 2]
                    if "R$" in money_block:
                        payee = money_block.split("R$", 1)[0].strip()

        # precisa de algum bloco com dinheiro
        if not money_block or "R$" not in money_block or not doc:
            i += 1
            continue

        if "SALDO DIA" in money_block.upper():
            i += 1
            continue

        monies = re.findall(r"R\$\s*([\-]?\d[\d\.]*,\d{2})", money_block)
        if not monies:
            i += 1
            continue

        amt_txt = monies[0].strip()
        amt = parse_decimal(amt_txt.lstrip('R$').strip())

        # sinal: só pelo "-" do próprio valor ou por palavras-chave de débito
        ref_text = f"{hist_part} {hist} {money_block}"
        is_debit = amt_txt.startswith("-") or any(k in ref_text.upper() for k in debit_kw)
        if is_debit and amt > 0:
            amt = -amt

        # memo final
        hist_clean = (hist_part or hist or "").strip()
        memo = hist_clean
        if payee:
            memo = f"{memo} | {payee}".strip(" |")

        memo = _ascii_sanitize(memo)

        bank_id = f"{doc}_{dt.replace('/','')}_{i}"

        txs.append({
            "dt": dt,
            "doc": doc,
            "amount": amt.quantize(Decimal("0.01")),
            "memo": memo,
            "raw": {"dt": dt, "doc": doc, "line1": ln, "money_block": money_block, "payee": payee}
        })

        # avança para evitar duplicar parsing no mesmo bloco
        i += 2
        continue

    return txs



def import_bank_pdf_caixa(conn: sqlite3.Connection, path: str, provider: str = "BANCO") -> ImportResult:
    """
    Importa Extrato PDF da Caixa (texto selecionável).
    Estratégia:
      - parseia transações do PDF
      - tenta ENRIQUECER linhas já importadas (ex.: via OFX) por (dt, amount, bank_id/doc, memo contém hist)
      - se não encontrar, insere como novo lançamento
    """
    warnings, errors = [], []
    inserted = 0
    dropped = 0
    updated = 0
    try:
        raw_text = _pdf_extract_text(path)
        txs = _parse_caixa_pdf_transactions(raw_text)
    except Exception as e:
        return ImportResult(inserted=0, dropped_dupe=0, warnings=warnings, errors=[str(e)])

    cur = conn.cursor()
    for t in txs:
        dt_iso = iso_from_ddmmyyyy(t["dt"])
        amount_s = str(t["amount"])
        memo = _ascii_sanitize(t["memo"])
        bank_id = (t.get("doc") or "").strip()

        # 1) tenta atualizar linha existente com memo genérico (mesmo dt/amount/doc)
        try:
            if bank_id:
                cur.execute(
                    "UPDATE bank_tx SET memo = CASE WHEN memo IS NULL OR memo = '' THEN ? "
                    "WHEN instr(upper(memo), upper(?)) = 0 THEN memo || ' | ' || ? ELSE memo END "
                    "WHERE provider=? AND dt=? AND amount=? AND bank_id=? AND is_deleted=0",
                    (memo, memo, memo, provider, dt_iso, amount_s, bank_id)
                )
                if cur.rowcount and cur.rowcount > 0:
                    updated += cur.rowcount
                    continue
        except Exception:
            pass

        # 2) se não achou por doc, tenta por dt+amount e hist contido
        try:
            hist_key = memo.split("|", 1)[0].strip()
            if hist_key:
                cur.execute(
                    "UPDATE bank_tx SET memo = CASE WHEN instr(upper(memo), upper(?)) = 0 THEN memo || ' | ' || ? ELSE memo END "
                    "WHERE provider=? AND dt=? AND amount=? AND memo LIKE ? AND is_deleted=0",
                    (memo, memo, provider, dt_iso, amount_s, f"%{hist_key}%")
                )
                if cur.rowcount and cur.rowcount > 0:
                    updated += cur.rowcount
                    continue
        except Exception:
            pass

        # 3) inserir novo
        try:
            if (_row_exists(conn, "bank_tx", provider, dt_iso, None, None, None, amount_s=amount_s)):
                dropped += 1
                continue
            cur.execute(
                "INSERT INTO bank_tx(provider, dt, amount, memo, bank_id, raw_json) VALUES (?,?,?,?,?,?)",
                (provider, dt_iso, amount_s, memo, bank_id, json.dumps(t, ensure_ascii=False, default=_json_default))
            )
            inserted += 1
        except Exception as e:
            errors.append(str(e))

    conn.commit()
    if updated:
        warnings.append(f"PDF Caixa: {updated} lançamento(s) enriquecido(s) com favorecido/linha 2.")
    return ImportResult(inserted=inserted, dropped_dupe=dropped, warnings=warnings, errors=errors)


def import_bank_csv_bradesco(conn: sqlite3.Connection, path: str, provider: str = "BANCO") -> ImportResult:
    """
    Importador CSV bancário (robusto para Bradesco/Itaú/outros).

    Estratégia:
    1) Tenta o layout clássico Bradesco (linha com Data;...;Crédito;Débito).
    2) Se não reconhecer, cai para um modo genérico (sep autodetect, colunas variadas).
    """
    warnings, errors = [], []
    inserted = dropped = 0

    # -------- 1) Tentativa layout "clássico Bradesco" --------
    header_idx = None
    chosen_enc = None
    for enc in ("utf-8-sig", "latin1", "cp1252"):
        try:
            with open(path, "r", encoding=enc) as f:
                lines = f.read().splitlines()
            chosen_enc = enc
            for i, line in enumerate(lines[:200]):
                l = line.strip().lower()
                if l.startswith("data;") and ("credito" in l or "crédito" in l) and ("debito" in l or "débito" in l):
                    header_idx = i
                    break
            break
        except Exception:
            continue

    def _load_csv_generic() -> pd.DataFrame:
        last_err = None
        for enc in (chosen_enc, "utf-8-sig", "latin1", "cp1252"):
            if not enc:
                continue
            try:
                return pd.read_csv(path, dtype=str, encoding=enc, sep=None, engine="python")
            except Exception as e:
                last_err = e
                continue
        raise last_err if last_err else Exception("Falha ao ler CSV (encoding).")

    try:
        if header_idx is not None:
            df = pd.read_csv(path, dtype=str, encoding=chosen_enc, sep=";", skiprows=header_idx, engine="python")
        else:
            df = _load_csv_generic()
            warnings.append("CSV não bateu no layout clássico; usei modo genérico (auto-separador).")

        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        return ImportResult(0, 0, [], [f"Falha ao ler CSV do banco: {e}"])

    if df.empty:
        return ImportResult(0, 0, warnings, [])

    # -------- 2) Mapeamento de colunas (genérico) --------
    norm_cols = {c: normalize_col(c) for c in df.columns}

    def pick(*cands):
        for c in df.columns:
            nc = norm_cols.get(c, "")
            if nc in cands:
                return c
        return None

    col_data = pick(
        "data", "dt", "data_lancamento", "data_mov", "data_movimento", "data_transacao", "data_da_transacao", "data_do_lancamento"
    )

    col_memo = pick(
        "historico", "histórico", "descricao", "descrição", "lancamento", "lançamento", "descricao_do_lancamento", "descricao_lancamento", "memo"
    )

    col_doc = pick("documento", "doc", "dcto", "numero_documento", "n_documento", "num_documento", "id")

    col_cred = None
    col_deb = None
    for c in df.columns:
        nc = norm_cols.get(c, "")
        if col_cred is None and "credito" in nc:
            col_cred = c
        if col_deb is None and "debito" in nc:
            col_deb = c

    col_val = pick("valor", "vlr", "valor_rs", "valor_r", "amount", "valor_do_lancamento")

    if not col_data:
        return ImportResult(0, 0, warnings, [f"CSV do banco não reconhecido (não achei coluna de data). Colunas: {list(df.columns)}"])

    if not (col_val or col_cred or col_deb):
        return ImportResult(0, 0, warnings, [f"CSV do banco não reconhecido (não achei coluna de valor). Colunas: {list(df.columns)}"])

    cur = conn.cursor()

    def parse_amt_csv(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none"):
            return None
        s = s.replace("R$", "").replace("r$", "").strip()
        s = s.replace(".", "").replace(",", ".")
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()
        try:
            d = Decimal(s)
            return -d if neg else d
        except Exception:
            return None

    for _, r in df.iterrows():
        d = parse_date(r.get(col_data))
        if not d:
            continue
        dt_iso = d.isoformat()

        memo = str(r.get(col_memo) or "").strip() if col_memo else ""
        bank_id = str(r.get(col_doc) or "").strip() if col_doc else ""

        if col_cred or col_deb:
            cred = parse_amt_csv(r.get(col_cred)) if col_cred else Decimal("0")
            deb = parse_amt_csv(r.get(col_deb)) if col_deb else Decimal("0")
            cred = cred or Decimal("0")
            deb = deb or Decimal("0")
            amt = cred - deb
        else:
            amt = parse_amt_csv(r.get(col_val))
            if amt is None:
                continue

        amount_s = _bank_amt2_str(amt)

        if (_bank_row_exists_transfer(conn, provider, dt_iso, amount_s, memo) if _bank_is_transfer_memo(memo) else _row_exists(conn, "bank_tx", provider, dt_iso, None, None, None, amount_s=amount_s)):
            dropped += 1
            continue

        cur.execute(
            "INSERT INTO bank_tx(provider, dt, amount, memo, bank_id, raw_json) VALUES (?,?,?,?,?,?)",
            (provider, dt_iso, amount_s, memo, bank_id, None),
        )
        inserted += 1

    conn.commit()
    return ImportResult(inserted, dropped, warnings, errors)

def _daily_series(conn: sqlite3.Connection, table: str, provider: str, start: date, end_inclusive: date, value_col: str) -> pd.DataFrame:
    q = f"""SELECT dt, COALESCE({value_col},0) AS v FROM {table}
             WHERE provider=? AND is_deleted=0 AND dt>=? AND dt<=?"""
    rows = conn.execute(q, (provider, start.isoformat(), end_inclusive.isoformat())).fetchall()
    sums: dict[date, Decimal] = {}
    for r in rows:
        d = parse_date(r["dt"])
        v = parse_decimal(r["v"]) or Decimal("0")
        if d:
            sums[d] = sums.get(d, Decimal("0")) + v
    all_days = pd.date_range(start=start, end=end_inclusive, freq="D")
    data = [{"dt": ts.date(), "total": sums.get(ts.date(), Decimal("0"))} for ts in all_days]
    return pd.DataFrame(data)

def _get_period(app_month: date, from_d: date | None, to_d: date | None) -> tuple[date, date]:
    if from_d and to_d:
        if to_d < from_d:
            return from_d, from_d
        return from_d, to_d
    start, end_excl = month_range(app_month)
    return start, (end_excl - timedelta(days=1))

def run_step1_daily(conn: sqlite3.Connection, provider: str, month: date, from_d: date | None, to_d: date | None) -> pd.DataFrame:
    """Etapa 1 (diária): ERP x Vendas (Portal).

    Regra geral: mantém colunas bruto e liq.
    Exceção Ticket: relatórios de vendas não trazem líquido -> concilia apenas BRUTO (liquido fica em branco).
    Observação global: NÃO exibe dias onde ERP=0 e Portal=0 (coluna 1 e 2 sem movimento).
    """
    start, end = _get_period(month, from_d, to_d)

    erp_b = _daily_series(conn, "erp_tx", provider, start, end, "bruto")
    sal_b = _daily_series(conn, "sales_tx", provider, start, end, "bruto")

    # Exceção FarmaciasAPP: Etapa 1 = ERP x Recebimentos (Portal), comparando VALOR PAGO NA LOJA.
    # No recebimento, esse valor é armazenado em raw_json["valor_pago_loja"].
    if provider.upper() == "FARMACIASAPP":
        rec_pago = _daily_series_receb_pago_loja(conn, provider, start, end)
        df = pd.DataFrame({"dt": erp_b["dt"]})
        df["erp_bruto"] = erp_b["total"]
        df["portal_bruto"] = rec_pago["total"]
        df["div_bruto"] = df["portal_bruto"] - df["erp_bruto"]
        df["erp_liq"] = pd.NA
        df["portal_liq"] = pd.NA
        df["div_liq"] = pd.NA
        a = pd.to_numeric(df["erp_bruto"], errors="coerce").fillna(0)
        b = pd.to_numeric(df["portal_bruto"], errors="coerce").fillna(0)
        df = df[(a != 0) | (b != 0)].copy()
        return df


    # mantém eixo de datas alinhado pelo ERP (como já era) e junta Portal pelo dt
    df = pd.DataFrame({"dt": erp_b["dt"]})
    df["erp_bruto"] = erp_b["total"]
    df["portal_bruto"] = sal_b["total"]
    df["div_bruto"] = df["portal_bruto"] - df["erp_bruto"]

    if provider.upper() == "TICKET":
        # Ticket: relatório de vendas não tem líquido -> deixa vazio (NaN)
        df["erp_liq"] = pd.NA
        df["portal_liq"] = pd.NA
        df["div_liq"] = pd.NA
    else:
        erp_l = _daily_series(conn, "erp_tx", provider, start, end, "liquido")
        sal_l = _daily_series(conn, "sales_tx", provider, start, end, "liquido")
        df["erp_liq"] = erp_l["total"]
        df["portal_liq"] = sal_l["total"]
        df["div_liq"] = df["portal_liq"] - df["erp_liq"]

    # ✅ regra global: se coluna 1 e 2 são zero, não exibir linha
    a = pd.to_numeric(df["erp_bruto"], errors="coerce").fillna(0)
    b = pd.to_numeric(df["portal_bruto"], errors="coerce").fillna(0)
    df = df[(a != 0) | (b != 0)].copy()

    return df

def run_step2_daily(conn: sqlite3.Connection, provider: str, month: date, from_d: date | None, to_d: date | None) -> pd.DataFrame:
    """Etapa 2 (diária): Vendas (Portal) x Recebimentos (Portal), usando a DATA DA VENDA para o período.

    Ticket (novo padrão):
      Col1 = vendas_portal_bruto
      Col2 = receb_portal_bruto
      Col3 = dif_bruto (receb - vendas)  -> verifica se todas as vendas do relatório de vendas aparecem no de recebimentos
      Col4 = erp_liq
      Col5 = receb_portal_liq
      Col6 = perc_cobrado (taxa) = (1 - receb_liq / erp_liq) * 100

    Outras bandeiras (mantém comportamento atual):
      vendas_liq x receb_liq por dia.

    Observação global: NÃO exibe dias onde col1=0 e col2=0.
    """
    start, end = _get_period(month, from_d, to_d)

    prov = provider.upper()
    if prov == "TICKET":
        sal_b = _daily_series(conn, "sales_tx", provider, start, end, "bruto")
        rec_b = _daily_series(conn, "receb_tx", provider, start, end, "bruto")
        erp_l = _daily_series(conn, "erp_tx", provider, start, end, "liquido")

        # --- RECEB_LIQ (contratual) recalculado em tempo real conforme Taxas/Tarifas ---
        rules = fee_rules_tx_list(conn, prov)
        rule = fee_rule_tx_match(rules, "VOUCHER") if rules else None
        if rule is None and rules:
            rule = next((r for r in rules if isinstance(r, dict) and r.get("is_active")), None)
        try:
            mdr_pct = Decimal(str((rule or {}).get("mdr_percent", 0) or 0).replace(",", "."))
        except Exception:
            mdr_pct = Decimal("0")
        try:
            fee_fixed = Decimal(str((rule or {}).get("fee_fixed", 0) or 0).replace(",", "."))
        except Exception:
            fee_fixed = Decimal("0")

        cols_r = {r[1] for r in conn.execute("PRAGMA table_info(receb_tx)").fetchall()}
        has_raw = "raw_json" in cols_r
        q = f"SELECT dt, bruto{', raw_json' if has_raw else ''} FROM receb_tx WHERE provider=? AND is_deleted=0 AND dt>=? AND dt<=?"
        rows = conn.execute(q, (prov, start.isoformat(), end.isoformat())).fetchall()
        rec_liq_by_day = {}
        for rr in rows:
            rr = dict(rr)
            d = parse_date(rr.get("dt"))
            if not d:
                continue
            bruto = parse_decimal(rr.get("bruto")) or Decimal("0")
            if bruto <= 0:
                continue
            qty = 1
            if has_raw:
                try:
                    j = json.loads(rr.get("raw_json") or "{}")
                    qty = int(j.get("qtd_transacoes") or j.get("qtd_transacoes_do_dia") or 1)
                except Exception:
                    qty = 1
            taxa = (bruto * (mdr_pct / Decimal("100"))).quantize(Decimal("0.01"))
            tarifa = (fee_fixed * Decimal(str(max(1, qty)))).quantize(Decimal("0.01"))
            liq = (bruto - taxa - tarifa).quantize(Decimal("0.01"))
            rec_liq_by_day[d] = rec_liq_by_day.get(d, Decimal("0")) + liq

        # monta série diária no mesmo formato de _daily_series()
        rec_l = pd.DataFrame({"dt": sal_b["dt"]})
        rec_l["total"] = rec_l["dt"].apply(lambda x: float(rec_liq_by_day.get(parse_date(x) or x, Decimal("0"))))

        df = pd.DataFrame({"dt": sal_b["dt"]})
        df["vendas_portal_bruto"] = sal_b["total"]
        df["receb_portal_bruto"] = rec_b["total"]
        df["dif_bruto"] = df["receb_portal_bruto"] - df["vendas_portal_bruto"]

        df["erp_liq"] = erp_l["total"]
        df["receb_portal_liq"] = rec_l["total"]

        # taxa (% cobrado): se ERP=0, deixa vazio
        erp = pd.to_numeric(df["erp_liq"], errors="coerce")
        rec = pd.to_numeric(df["receb_portal_liq"], errors="coerce")
        perc = (1 - (rec / erp)) * 100
        perc = perc.replace([np.inf, -np.inf], np.nan)
        df["perc_cobrado"] = perc.where(erp.notna() & (erp != 0), pd.NA)

        a = pd.to_numeric(df["vendas_portal_bruto"], errors="coerce").fillna(0)
        b = pd.to_numeric(df["receb_portal_bruto"], errors="coerce").fillna(0)
        df = df[(a != 0) | (b != 0)].copy()
        return df

    # --- padrão antigo (Alelo etc.) ---
    sal_l = _daily_series(conn, "sales_tx", provider, start, end, "liquido")
    rec_l = _daily_series(conn, "receb_tx", provider, start, end, "liquido")
    df = pd.DataFrame({"dt": sal_l["dt"]})
    df["vendas_liq"] = sal_l["total"]
    df["receb_liq"] = rec_l["total"]
    df["div_liq"] = df["receb_liq"] - df["vendas_liq"]

    a = pd.to_numeric(df["vendas_liq"], errors="coerce").fillna(0)
    b = pd.to_numeric(df["receb_liq"], errors="coerce").fillna(0)
    df = df[(a != 0) | (b != 0)].copy()

    return df


# ==========================================================
# NOVO PADRÃO (v1.4 - captura/recebíveis/banco por evento)
# ==========================================================

def _table_cols(conn: sqlite3.Connection, table: str) -> set:
    """Retorna o conjunto de colunas existentes em uma tabela (cacheado)."""
    cache = getattr(conn, "_col_cache", None)
    if cache is None:
        cache = {}
        try:
            setattr(conn, "_col_cache", cache)
        except Exception:
            # conexão pode não aceitar attrs; usa cache global fallback
            pass
    # fallback global se necessário
    global _GLOBAL_COL_CACHE
    try:
        _GLOBAL_COL_CACHE
    except NameError:
        _GLOBAL_COL_CACHE = {}

    key = (id(conn), table)
    if key in _GLOBAL_COL_CACHE:
        return _GLOBAL_COL_CACHE[key]

    if table in cache:
        cols = cache[table]
    else:
        rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
        cols = {r[1] if isinstance(r, tuple) else r["name"] for r in rows}
        cache[table] = cols
    _GLOBAL_COL_CACHE[key] = cols
    return cols



def _daily_series_receb_pago_loja(conn: sqlite3.Connection, provider: str, start: date, end_inclusive: date) -> pd.DataFrame:
    """Série diária usando o campo valor_pago_loja armazenado em raw_json de receb_tx (FarmaciasAPP)."""
    rows = conn.execute(
        "SELECT dt, raw_json FROM receb_tx WHERE provider=? AND is_deleted=0 AND dt>=? AND dt<=?",
        (provider, start.isoformat(), end_inclusive.isoformat())
    ).fetchall()
    sums: dict[date, Decimal] = {}
    for r in rows:
        d = parse_date(r["dt"])
        if not d:
            continue
        try:
            j = json.loads(r["raw_json"] or "{}")
        except Exception:
            j = {}
        v = None
        if isinstance(j, dict):
            v = j.get("valor_pago_loja")
        v = parse_decimal(v) or Decimal("0")
        if d:
            sums[d] = sums.get(d, Decimal("0")) + v
    all_days = pd.date_range(start=start, end=end_inclusive, freq="D").date
    out = []
    for d in all_days:
        out.append({"dt": d.isoformat(), "total": float(sums.get(d, Decimal("0")))})
    return pd.DataFrame(out)


def _q_tx(conn: sqlite3.Connection, table: str, provider: str, start: date, end: date, date_expr: str = "dt"):
    """Consulta registros de uma tabela (compatível com schemas antigos).

    Alguns bancos antigos não têm colunas como pay_dt/raw_json. Esta função
    seleciona apenas o que existir e preenche faltantes com NULL/''.

    date_expr:
      - "dt" (padrão)
      - "COALESCE(pay_dt, dt)" (recebíveis por data de pagamento) -> cai para "dt" se pay_dt não existir.
    """
    cols = _table_cols(conn, table)

    # Se a expressão usa pay_dt mas a coluna não existe, regride para dt
    if "pay_dt" not in cols and "pay_dt" in date_expr:
        date_expr = "dt"

    sel_pay_dt = "pay_dt" if "pay_dt" in cols else "NULL"
    sel_aut = "autorizacao" if "autorizacao" in cols else "''"
    sel_raw = "raw_json" if "raw_json" in cols else "NULL"
    sel_bruto = "bruto" if "bruto" in cols else "0"
    sel_liq = "liquido" if "liquido" in cols else "0"

    q = f"""SELECT id,
                         dt,
                         {sel_pay_dt} AS pay_dt,
                         {sel_bruto} AS bruto,
                         {sel_liq} AS liquido,
                         {sel_aut} AS autorizacao,
                         {sel_raw} AS raw_json
                  FROM {table}
                  WHERE provider=? AND is_deleted=0
                    AND {date_expr}>=? AND {date_expr}<=?
                  ORDER BY {date_expr}, id"""
    rows = conn.execute(q, (provider, start.isoformat(), end.isoformat())).fetchall()

    out = []
    for r in rows:
        # sqlite Row suporta [] por nome
        d = parse_date(r["dt"]) if "dt" in r.keys() else None
        pay_d = parse_date(r["pay_dt"]) if ("pay_dt" in r.keys() and r["pay_dt"] is not None) else None
        out.append({
            "id": int(r["id"]),
            "dt": d,
            "pay_dt": pay_d,
            "bruto": (parse_decimal(r["bruto"]) or Decimal("0")),
            "liquido": (parse_decimal(r["liquido"]) or Decimal("0")),
            "autorizacao": (r["autorizacao"] or "").strip(),
            "raw_json": r["raw_json"],
        })
    return out


def _cents(d: Decimal) -> int:
    try:
        return int((d.quantize(Decimal("0.01")) * 100))
    except Exception:
        return int(Decimal(str(d)).quantize(Decimal("0.01")) * 100)

def _find_subset_sum(items, target_cents: int, max_items: int = 6):
    """Backtracking limitado para achar combinação que fecha exatamente target."""
    if target_cents == 0:
        return []
    # ordena por valor desc para reduzir busca
    items = sorted(items, key=lambda x: x["_c"], reverse=True)
    best = None

    def bt(i, cur_sum, chosen):
        nonlocal best
        if best is not None:
            return
        if cur_sum == target_cents:
            best = chosen[:]
            return
        if cur_sum > target_cents:
            return
        if len(chosen) >= max_items:
            return
        if i >= len(items):
            return
        # bound simples
        rem_max = cur_sum + sum(it["_c"] for it in items[i:i+(max_items-len(chosen))])
        if rem_max < target_cents:
            return
        # escolhe
        bt(i+1, cur_sum + items[i]["_c"], chosen + [items[i]])
        # não escolhe
        bt(i+1, cur_sum, chosen)

    bt(0, 0, [])
    return best

def run_step1_capture(conn: sqlite3.Connection, provider: str, month: date, from_d: date | None, to_d: date | None,
                      window_days: int = 2) -> pd.DataFrame:
    """Etapa 1 (evento): ERP x Vendas (captura) - usa BRUTO.

    - ERP pode vir sem NSU e/ou agrupando vendas.
    - Portal (Vendas) traz NSU (Alelo) ou Nº Transação (Ticket) via raw_json.
    - Matching:
        1) 1↔1 por valor (centavos) dentro de ±window_days
        2) n↔1 (soma de vendas) para fechar o valor do ERP dentro da mesma janela
    """
    start, end = _get_period(month, from_d, to_d)
    prov = provider.upper()

    erp = _q_tx(conn, "erp_tx", provider, start, end, date_expr="dt")
    # FarmaciasAPP: Etapa 1 usa REL. VENDAS (sales_tx) como fonte de verdade do "Total pago na loja"
    if prov == "FARMACIASAPP":
        sales = _q_tx(conn, "sales_tx", "FARMACIASAPP", start, end, date_expr="dt")
        # normaliza: ignora cancelados e garante bruto = valor pago na loja
        cleaned = []
        for s in sales:
            try:
                rj = json.loads(s.get("raw_json") or "{}")
            except Exception:
                rj = {}
            status = ""
            if isinstance(rj, dict):
                status = str(rj.get("status") or rj.get("Status") or "").strip().upper()
            if status in ("CANCELADO", "CANCELED", "CANCELLED"):
                continue
            # para FarmaciasAPP, autorizacao guarda o "Pedido"
            s["autorizacao"] = (s.get("autorizacao") or "").strip()
            s["bruto"] = parse_decimal(s.get("bruto")) or Decimal("0")
            s["liquido"] = s["bruto"]
            cleaned.append(s)
        sales = cleaned
    else:
        sales = _q_tx(conn, "sales_tx", provider, start, end, date_expr="dt")

    # prepara lista de NSU para Ticket (quando agrupado por reembolso)
    for s in sales:
        s["nsu_list"] = []
        if s.get("autorizacao"):
            # Alelo: autorizacao já é o NSU
            s["nsu_list"] = [s["autorizacao"]]
        try:
            rj = json.loads(s.get("raw_json") or "{}")
            if isinstance(rj, dict) and rj.get("nsu_list"):
                s["nsu_list"] = [x for x in rj.get("nsu_list") if str(x).strip()]
        except Exception:
            pass
        s["_c"] = _cents(s["bruto"])
        s["used"] = False

    rows = []
    # index sales por janela rápida
    sales_sorted = sorted(sales, key=lambda x: (x["dt"], x["id"]))
    for e in erp:
        e_amt_c = _cents(e["bruto"])
        e_dt = e["dt"]
        # candidatos por janela
        cand = [s for s in sales_sorted if (not s["used"]) and s["dt"] and e_dt and abs((s["dt"] - e_dt).days) <= window_days]
        # 1) 1-1
        one = [s for s in cand if s["_c"] == e_amt_c]
        chosen = None
        if one:
            one = sorted(one, key=lambda s: (abs((s["dt"] - e_dt).days), s["id"]))
            chosen = [one[0]]
        else:
            # 2) subset sum (n vendas)
            subset = _find_subset_sum(cand, e_amt_c, max_items=6)
            if subset:
                chosen = subset

        if chosen:
            for s in chosen:
                s["used"] = True
            portal_sum = sum([s["bruto"] for s in chosen], Decimal("0"))
            nsus = []
            for s in chosen:
                nsus += (s.get("nsu_list") or [])
            nsus = [n for n in nsus if str(n).strip()]
            # fallback: se não achou nsu_list, guarda IDs
            if not nsus:
                nsus = [f"ID:{s['id']}" for s in chosen]
            delta_days = min([abs((s["dt"] - e_dt).days) for s in chosen]) if chosen else 0
            diff = portal_sum - e["bruto"]
            ok = (diff == 0)
            rows.append({
                "data": e_dt,
                "erp_id": e["id"],
                "erp_bruto": e["bruto"],
                "vendas_ref": ", ".join(nsus[:5]) + ("..." if len(nsus) > 5 else ""),
                "vendas_bruto": portal_sum,
                "delta_dias": delta_days,
                "status": "✅" if ok else "❌",
                "diferenca": diff,
            })
        else:
            # ERP sem captura
            rows.append({
                "data": e_dt,
                "erp_id": e["id"],
                "erp_bruto": e["bruto"],
                "vendas_ref": "",
                "vendas_bruto": Decimal("0"),
                "delta_dias": "",
                "status": "❌",
                "diferenca": -e["bruto"],
            })

    # vendas sobrando (capturado mas não está no ERP)
    for s in sales_sorted:
        if s.get("used"):
            continue
        nsus = s.get("nsu_list") or ([s.get("autorizacao")] if s.get("autorizacao") else [])
        if not nsus:
            nsus = [f"ID:{s['id']}"]
        rows.append({
            "data": s["dt"],
            "erp_id": f"CLONE:{s['id']}",
            "erp_bruto": s["bruto"],
            "vendas_ref": ", ".join(nsus[:5]) + ("..." if len(nsus) > 5 else ""),
            "vendas_bruto": s["bruto"],
            "delta_dias": 0,
            "status": "🟦✅ (ERP ausente - clonado)",
            "diferenca": Decimal("0"),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.sort_values(by=["data","erp_id"], kind="mergesort").reset_index(drop=True)
    return df

def run_step2_recebiveis(
    conn: sqlite3.Connection,
    provider: str,
    month: date,
    from_d: date | None,
    to_d: date | None,
    window_days: int = 0
) -> pd.DataFrame:
    """Etapa 2 (evento): Vendas x Recebíveis.

    Mostra a taxa como 'OK' quando:
      - Vendas (BRUTO) == Recebível (BRUTO) e
      - Recebível (LÍQ) <= Recebível (BRUTO)

    Matching principal:
      - por autorizacao (NSU / chave Ticket reembolso|data)
      - fallback por valor bruto no mesmo dia (ou janela, se window_days>0)
    """

    start, end = _get_period(month, from_d, to_d)
    prov = (provider or "").strip().upper()

    # ============================================================
    # FARMACIASAPP — Etapa 2 por PEDIDO (layout semelhante ao relatório)
    # ============================================================

    if prov == "FARMACIASAPP":
        # Fonte de verdade do VALOR_PAGO_LOJA / FORMA_PAGAMENTO / DATA_PEDIDO: REL. VENDAS (sales_tx)
        sales_rows = conn.execute(
            """
            SELECT dt, autorizacao, bruto, raw_json
            FROM sales_tx
            WHERE provider=? AND is_deleted=0 AND dt BETWEEN ? AND ?
            ORDER BY dt, id
            """,
            ("FARMACIASAPP", start.isoformat(), end.isoformat()),
        ).fetchall()

        sales_by_pedido: dict[str, dict] = {}
        for dt_s, pedido_s, bruto_s, raw_s in sales_rows:
            pedido_s = (pedido_s or "").strip()
            if not pedido_s:
                continue
            # status (ignora cancelado)
            status = ""
            forma = ""
            try:
                rj = json.loads(raw_s or "{}")
            except Exception:
                rj = {}
            if isinstance(rj, dict):
                status = str(rj.get("status") or rj.get("Status") or "").strip().upper()
                forma = str(rj.get("forma_pagamento") or rj.get("Forma de Pagamento") or rj.get("forma pgto") or "").strip().upper()
            if status in ("CANCELADO", "CANCELED", "CANCELLED"):
                continue

            sales_by_pedido[pedido_s] = {
                "dt": dt_s,
                "valor_pago_loja": parse_decimal(bruto_s) or Decimal("0"),
                "forma_pagamento": forma,
            }

        # Regras: separar Taxa ADM (fixa) e Taxa Modalidade (PIX/CREDIT)
        rules = fee_rules_tx_list(conn, "FARMACIASAPP")
        active = [r for r in rules if isinstance(r, dict) and int(r.get("is_active") or 0) == 1]

        adm_pct = Decimal("0")
        for r in active:
            mt = normalize_text(r.get("match_text") or "")
            lb = normalize_text(r.get("label") or "")
            if "ADM" in mt or "ADM" in lb:
                adm_pct = parse_decimal(r.get("mdr_percent")) or Decimal("0")
                break

        def _modal_pct(forma: str) -> Decimal:
            t = (forma or "").strip().upper()
            for r in active:
                mt = normalize_text(r.get("match_text") or "")
                lb = normalize_text(r.get("label") or "")
                if "ADM" in mt or "ADM" in lb:
                    continue
                if not mt:
                    continue
                if mt in normalize_text(t):
                    return parse_decimal(r.get("mdr_percent")) or Decimal("0")
            # fallback: tenta pelo label
            for r in active:
                lb = normalize_text(r.get("label") or "")
                if "ADM" in lb:
                    continue
                if not lb:
                    continue
                if lb in normalize_text(t):
                    return parse_decimal(r.get("mdr_percent")) or Decimal("0")
            return Decimal("0")

        # Recebíveis (relatório da operadora): total_compra (bruto), repasse_aplicado (liquido), subsidio etc. via raw_json
        rows_db = conn.execute(
            """
            SELECT id, dt, bruto, liquido, autorizacao, raw_json
            FROM receb_tx
            WHERE provider=? AND is_deleted=0 AND dt BETWEEN ? AND ?
            ORDER BY dt, id
            """,
            ("FARMACIASAPP", start.isoformat(), end.isoformat()),
        ).fetchall()

        out = []
        for rr in rows_db:
            rr = dict(rr)
            pedido = (rr.get("autorizacao") or "").strip()
            total_compra = parse_decimal(rr.get("bruto")) or Decimal("0")
            repasse_apl = parse_decimal(rr.get("liquido")) or Decimal("0")
            taxa_apl = (total_compra - repasse_apl).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            # subsidio vem do raw_json do recebível
            subsidio = None
            forma_rx = ""
            try:
                j = json.loads(rr.get("raw_json") or "{}")
            except Exception:
                j = {}
            if isinstance(j, dict):
                subsidio = j.get("subsidio_desconto")
                forma_rx = str(j.get("forma_pagamento") or "").strip().upper()

            # fonte verdade: rel vendas (se existir)
            s_info = sales_by_pedido.get(pedido) or {}
            dt_pedido = s_info.get("dt") or (rr.get("dt") or "")
            forma = (s_info.get("forma_pagamento") or forma_rx or "").strip().upper()
            valor_pago_loja = Decimal("0")
            try:
                vpl = s_info.get("valor_pago_loja", None)
                valor_pago_loja = vpl if isinstance(vpl, Decimal) else (parse_decimal(vpl) or Decimal("0"))
            except Exception:
                valor_pago_loja = Decimal("0")

            subsidio_d = (parse_decimal(subsidio) or Decimal('0'))

            # % cobrado aplicado (em %): taxa_apl / total_compra
            perc_apl = pd.NA
            if total_compra and total_compra != 0:
                try:
                    perc_apl = (taxa_apl / total_compra) * Decimal("100")
                except Exception:
                    perc_apl = pd.NA

            modal_pct = _modal_pct(forma)

            # total da compra (corrigido) = pago na loja + subsídio (não confiar no total do relatório)
            total_compra_calc = (valor_pago_loja + subsidio_d).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            # taxa esperada:
            # - TX ADM % sobre o total da compra (corrigido)
            # - TX MODAL % apenas sobre o valor pago na loja
            taxa_adm = (total_compra_calc * (adm_pct / Decimal("100"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            taxa_modal = (valor_pago_loja * (modal_pct / Decimal("100"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            taxa_esp = (taxa_adm + taxa_modal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            repasse_esp = (total_compra_calc - taxa_esp).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            # % cobrado esperado (apenas referência)
            perc_esp = (adm_pct + modal_pct).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            diff = (repasse_apl - repasse_esp).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            out.append({
                "pedido": pedido,
                "data_pedido": dt_pedido,
                "forma_pagamento": forma,
                "valor_pago_loja": valor_pago_loja,
                "valor_subsidio": subsidio_d,
                "valor_total_compra": total_compra_calc,
                "valor_repasse_aplicado": repasse_apl,
                "valor_taxa_aplicada": taxa_apl,
                "perc_cobrado_aplicado": perc_apl,
                "tx_adm_pct": adm_pct,
                "tx_modal_pct": modal_pct,
                "valor_repasse_esperado": repasse_esp,
                "valor_taxa_esperada": taxa_esp,
                "perc_cobrado_esperado": perc_esp,
                "diferenca": diff,
            })

        df = pd.DataFrame(out)
        return df


    # Fluxo padrão — Vendas x Recebíveis por evento/autorização
    # ============================================================
    sales = _q_tx(conn, "sales_tx", prov, start, end, date_expr="dt")
    rec = _q_tx(conn, "receb_tx", prov, start, end, date_expr="dt")

    sales_by_auth = {}
    for s in sales:
        a = (s.get("autorizacao") or "").strip()
        if a:
            sales_by_auth.setdefault(a, []).append(s)

    used_sales = set()
    used_rec = set()

    rows = []
    # 1) por autorizacao
    for r in rec:
        a = (r.get("autorizacao") or "").strip()
        chosen_sales = []
        if a and a in sales_by_auth:
            for s in sales_by_auth[a]:
                if s["id"] in used_sales:
                    continue
                chosen_sales.append(s)
            if chosen_sales:
                for s in chosen_sales:
                    used_sales.add(s["id"])
                used_rec.add(r["id"])

        # 2) fallback por soma de vendas (mesmo dia/janela)
        if not chosen_sales:
            cand = [
                s for s in sales
                if s["id"] not in used_sales and s["dt"] and r["dt"]
                and abs((s["dt"] - r["dt"]).days) <= window_days
            ]
            for s in cand:
                s["_c"] = _cents(s["bruto"])
            target = _cents(r["bruto"]) if r["bruto"] else _cents(r["liquido"])
            subset = _find_subset_sum(cand, target, max_items=8)
            if subset:
                chosen_sales = subset
                for s in chosen_sales:
                    used_sales.add(s["id"])
                used_rec.add(r["id"])

        vendas_bruto = sum([s["bruto"] for s in chosen_sales], Decimal("0")) if chosen_sales else Decimal("0")
        receb_bruto = r["bruto"]

        # Ticket: RECEB_LIQ é recalculado dinamicamente a partir das regras cadastradas
        if prov == "TICKET":
            mdr_p, fee_fixed, _txf = _get_active_fee_rule(conn, prov)
            qtd_ops = max(1, len(chosen_sales)) if chosen_sales else 1
            mdr_val = (vendas_bruto * (mdr_p / Decimal("100"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            fee_val = (fee_fixed * Decimal(qtd_ops)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            taxa_total = (mdr_val + fee_val).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            receb_liq = (vendas_bruto - taxa_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            taxa = taxa_total
        else:
            receb_liq = r["liquido"]
            taxa = (vendas_bruto - receb_liq) if receb_liq is not None else Decimal("0")

        dif_bruto = (receb_bruto - vendas_bruto)
        ok = (dif_bruto == 0) and (receb_liq <= receb_bruto if receb_bruto is not None else True)

        refs = []
        for s in chosen_sales:
            refs.append(s.get("autorizacao") or f"ID:{s['id']}")
        rep_aplic = r["liquido"]
        taxa_aplic = (receb_bruto - rep_aplic) if (receb_bruto is not None and rep_aplic is not None) else Decimal("0")
        rep_esp = receb_liq
        diff_rep = (rep_aplic - rep_esp) if (rep_aplic is not None and rep_esp is not None) else Decimal("0")
        rows.append({
            "data": r["dt"],
            "vendas_ref": ", ".join(refs[:6]) + ("..." if len(refs) > 6 else ""),
            "vendas_bruto": vendas_bruto,
            "receb_id": r["id"],
            "receb_bruto": receb_bruto,
            "repasse_aplicado": rep_aplic,
            "taxa_aplicada": taxa_aplic,
            "repasse_esperado": rep_esp,
            "diferenca": diff_rep,
            "status": "✅" if ok else "❌",
        })

    # vendas sem recebível
    for s in sales:
        if s["id"] in used_sales:
            continue
        nsus = [n for n in (s.get("nsu_list") or []) if str(n).strip()]
        nsu_txt = ", ".join([str(n) for n in nsus[:5]]) + ("..." if len(nsus) > 5 else "") if nsus else (s.get("autorizacao") or f"ID:{s['id']}")
        rows.append({
            "data": s["dt"],
            "vendas_ref": nsu_txt,
            "vendas_bruto": s["bruto"],
            "receb_id": "",
            "receb_bruto": Decimal("0"),
            "repasse_aplicado": Decimal("0"),
            "taxa_aplicada": Decimal("0"),
            "repasse_esperado": Decimal("0"),
            "diferenca": Decimal("0"),
            "status": "❌ (sem recebível)",
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.sort_values(by=["data","receb_id"], kind="mergesort").reset_index(drop=True)
    return df



def run_step3_banco(
    conn: sqlite3.Connection,
    provider: str,
    month: date,
    window_days: int = 2,
    bank_term: str | None = None,
    spillover_days: int = 15,
    bank_start_override: date | None = None,
    bank_end_override: date | None = None,
    expected_end: date | None = None,
) -> tuple[pd.DataFrame, dict[int, list[int]]]:
    """Etapa 3 (evento): Banco x Recebíveis (lógico, sem misturar lotes).

    Objetivo (melhor prática):
      - A fonte "Esperado" para Etapa 3 é o **Recebimentos (Portal) por data de pagamento** (pay_dt).
      - A fonte "Banco" é o extrato (créditos) filtrado por memo/termos.
      - A alocação deve ser **conservadora** para não puxar recebimentos de outros períodos/lotes:
          1) Primeiro aloca apenas recebíveis dentro da janela (bank_dt ± window_days)
          2) Só então (SE o depósito ainda tiver saldo) expande para frente até spillover_days
             (evita "puxar" valores futuros à toa)

    Retorna:
      - df_bank: 1 linha por lançamento bancário (com quanto foi alocado e saldo banco remanescente)
      - composition: dict[bank_tx_id] -> list[receb_tx_id]
    """
    prov = (provider or "").strip().upper() or "ALELO"
    start, end_excl = month_range(month)
    end_date = end_excl - timedelta(days=1)

    bank_start = (bank_start_override or start)

    spillover_days = int(max(0, spillover_days))
    window_days = int(max(0, window_days))
    bank_end = (bank_end_override or end_date)

    # FARMACIASAPP: crédito por lote pode liquidar alguns dias após a "data de transferência".
    if provider == "FARMACIASAPP":
        window_days = max(window_days, 7)
        spillover_days = max(spillover_days, 20)


    def _memo(s: str | None) -> str:
        return (s or "").strip().upper()

    # =========================
    # 1) Seleção de BANCO
    # =========================
    where = "WHERE provider='BANCO' AND is_deleted=0 AND dt>=? AND dt<=?"
    params: list = [bank_start.isoformat(), bank_end.isoformat()]

    # termos que indicam "não é repasse"
    NEG_MEMO = ["PIX", "TARIFA", "JUROS", "IOF", "SAQUE", "TED ENVIADA", "DOC ENVIADO"]

    if bank_term:
        t = bank_term.strip().upper()
        prov_up = str(provider or "").strip().upper()
        token_std = BANK_TRANSF_TOKEN_FARM if prov_up == "FARMACIASAPP" else (BANK_TRANSF_TOKEN_TICKET if prov_up == "TICKET" else BANK_TRANSF_TOKEN_ALELO)
        if (t == token_std or t.startswith(token_std) or token_std.startswith(t)):
            # Token padrão: usa termos conhecidos + termos do usuário
            terms = bank_all_memo_terms(provider)
            if terms:
                ors = " OR ".join(["UPPER(COALESCE(memo,'')) LIKE ?"] * len(terms))
                where += f" AND ({ors})"
                params += [f"%{x.upper()}%" for x in terms]
            else:
                where += " AND (UPPER(COALESCE(memo,'')) LIKE ? OR UPPER(COALESCE(memo,'')) LIKE ?)"
                params += ["%ALELO%", "%NAIP%"]
            for neg in NEG_MEMO:
                where += " AND UPPER(COALESCE(memo,'')) NOT LIKE ?"
                params.append(f"%{neg}%")
        else:
            where += " AND UPPER(COALESCE(memo,'')) LIKE ?"
            params.append(f"%{t}%")

    elif provider == "FARMACIASAPP":
        # FARMACIASAPP: depósitos por lote (Zoop/OTB). Se o usuário não informar termo, filtramos por memos típicos.
        terms = ["ZOOP", "OTB TECH"]
        ors = " OR ".join(["UPPER(COALESCE(memo,'')) LIKE ?"] * len(terms))
        where += f" AND ({ors})"
        params += [f"%{t}%" for t in terms]
        for neg in NEG_MEMO:
            where += " AND UPPER(COALESCE(memo,'')) NOT LIKE ?"
            params.append(f"%{neg}%")

    bank_rows = conn.execute(
        f"SELECT id, dt, amount, memo, COALESCE(bank_name,'') AS bank_name "
        f"FROM bank_tx {where} ORDER BY dt, id",
        tuple(params),
    ).fetchall()

    bank_state = []
    for r in bank_rows:
        r = dict(r)
        dtv = parse_date(r.get("dt"))
        if not dtv:
            continue
        amt = parse_decimal(r.get("amount")) or Decimal("0")
        if amt <= Decimal("0"):
            continue  # Etapa 3 trabalha com créditos
        bank_state.append({
            "id": int(r["id"]),
            "dt": dtv,
            "amount": amt,
            "memo": _memo(r.get("memo")),
            "bank_name": (r.get("bank_name") or "").strip(),
            "remaining": amt,
            "allocated": Decimal("0"),
        })

    # =========================
    # 2) Seleção de RECEBÍVEIS (por pay_dt)
    # =========================
    cols_r = {r[1] for r in conn.execute("PRAGMA table_info(receb_tx)").fetchall()}
    dt_target_expr = "COALESCE(pay_dt, dt)" if "pay_dt" in cols_r else "dt"

    def _pick_expr(cols: set[str], candidates: list[str], default: str) -> str:
        for c in candidates:
            if c in cols:
                return c
        return default

    amt_expr = _pick_expr(cols_r, ["amount", "liquido", "valor_liquido", "valor_liq", "net", "net_amount"], "0")
    nsu_expr = _pick_expr(cols_r, ["nsu", "autorizacao", "autorizacao_nsu"], "NULL")
    auth_expr = _pick_expr(cols_r, ["auth", "autorizacao", "autorizacao_nsu"], "NULL")
    raw_expr = "raw_json" if "raw_json" in cols_r else "NULL"

    exp_end = (expected_end or bank_end)

    r_where = f"WHERE provider=? AND is_deleted=0 AND {dt_target_expr}>=? AND {dt_target_expr}<=?"
    r_params = [prov, start.isoformat(), exp_end.isoformat()]

    receb_rows = conn.execute(
        f"SELECT id, dt AS dt, pay_dt AS pay_dt, {dt_target_expr} AS dt_target, "
        f"{amt_expr} AS amount, {nsu_expr} AS nsu, {auth_expr} AS auth, {raw_expr} AS raw_json "
        f"FROM receb_tx {r_where} ORDER BY dt_target, id",
        tuple(r_params),
    ).fetchall()

    receb = []
    if provider == "FARMACIASAPP":
        # Agrupa por lote: TransferId (Zoop) quando existir; senão por pay_dt (OTB).
        batches: dict[str, dict] = {}
        for r in receb_rows:
            r = dict(r)
            amt = parse_decimal(r.get("amount")) or Decimal("0")
            if amt <= Decimal("0"):
                continue
            dt_target = parse_date(r.get("dt_target")) or parse_date(r.get("pay_dt")) or parse_date(r.get("dt"))
            if not dt_target:
                continue
            # tenta extrair TransferId e data de transferência do raw_json
            transfer_id = ""
            transfer_dt = None
            try:
                j = json.loads(r.get("raw_json") or "{}")
                if isinstance(j, dict):
                    transfer_id = str(j.get("transfer_id") or j.get("TransferId") or "").strip()
                    transfer_dt = parse_date(j.get("transfer_date") or j.get("DataTransferencia") or j.get("data_transferencia"))
            except Exception:
                pass
            key = transfer_id or f"PAYDT:{dt_target.isoformat()}"
            base_dt = transfer_dt or dt_target
            bdt = next_business_day(base_dt)
            if key not in batches:
                batches[key] = {"key": key, "dt": bdt, "amount": Decimal("0"), "ids": []}
            batches[key]["amount"] += amt
            batches[key]["ids"].append(int(r["id"]))
            # mantém a menor data (segurança)
            if bdt < batches[key]["dt"]:
                batches[key]["dt"] = bdt
        for b in sorted(batches.values(), key=lambda x: (x["dt"], x["key"])):
            receb.append({"id": -1, "dt": b["dt"], "amount": b["amount"], "ref": b["key"], "ids": b["ids"]})
    else:
        for r in receb_rows:
            r = dict(r)
            amt = parse_decimal(r.get("amount")) or Decimal("0")
            if amt <= Decimal("0"):
                continue
            dt_target = parse_date(r.get("dt_target")) or parse_date(r.get("pay_dt")) or parse_date(r.get("dt"))
            if not dt_target:
                continue
            receb.append({
                "id": int(r["id"]),
                "dt": next_business_day(dt_target),
                "amount": amt,
                "ref": (str(r.get("nsu") or r.get("auth") or "").strip() or ""),
            })

    # 3) Alocação conservadora (depósito -> recebíveis)
    # =========================
    composition: dict[int, list[int]] = {b["id"]: [] for b in bank_state}
    remaining_receb = [{"id": r["id"], "dt": r["dt"], "amount": r["amount"], "ids": r.get("ids", [r["id"]])} for r in receb]

    # índice auxiliar para acelerar (lista já vem ordenada por dt)
    def _eligible_receb_indices(bank_dt: date, lo: int, hi: int) -> list[int]:
        idx = []
        for i in range(lo, hi):
            rr = remaining_receb[i]
            if rr["amount"] <= Decimal("0"):
                continue
            idx.append(i)
        return idx

    # Para buscar rápido, calculamos janelas por varredura simples (volumes típicos são OK).
    for b in bank_state:
        if b["remaining"] <= Decimal("0"):
            continue
        bd = b["dt"]

        # FARMACIASAPP: tenta casar 1 depósito (banco) com 1 lote (TransferId/pay_dt) por valor exato (tolerância).
        if provider == "FARMACIASAPP":
            tol = Decimal("0.05")
            best_rr = None
            best_idx = None
            for idx_rr, rr in enumerate(remaining_receb):
                if rr["amount"] <= Decimal("0"):
                    continue
                # janela: aceita alguns dias de liquidação após a data de transferência/pagamento
                if rr["dt"] < (bd - timedelta(days=window_days)) or rr["dt"] > (bd + timedelta(days=spillover_days)):
                    continue
                if abs(rr["amount"] - b["remaining"]) <= tol:
                    best_rr = rr; best_idx = idx_rr; break
            if best_rr is not None:
                alloc = best_rr["amount"]
                best_rr["amount"] = Decimal("0")
                b["remaining"] -= alloc
                b["allocated"] += alloc
                composition[b["id"]].extend(best_rr.get("ids") or [])
                # depósito fechado: segue para o próximo
                continue


        # 3.1) Primeira passada: apenas dt dentro de [bd-window, bd+window]
        w_lo = bd - timedelta(days=window_days)
        w_hi = bd + timedelta(days=window_days)

        for rr in remaining_receb:
            if b["remaining"] <= Decimal("0"):
                break
            if rr["amount"] <= Decimal("0"):
                continue
            if rr["dt"] < w_lo or rr["dt"] > w_hi:
                continue
            alloc = rr["amount"] if b["remaining"] >= rr["amount"] else b["remaining"]
            if alloc <= Decimal("0"):
                continue
            rr["amount"] -= alloc
            b["remaining"] -= alloc
            b["allocated"] += alloc
            composition[b["id"]].extend(rr.get("ids") or [rr["id"]])

        # 3.2) Segunda passada (spillover): só se ainda faltar fechar o depósito
        if b["remaining"] > Decimal("0") and spillover_days > 0:
            s_hi = bd + timedelta(days=spillover_days)
            for rr in remaining_receb:
                if b["remaining"] <= Decimal("0"):
                    break
                if rr["amount"] <= Decimal("0"):
                    continue
                # spillover apenas para FRENTE (evita puxar de outros lotes antigos)
                if rr["dt"] <= w_hi or rr["dt"] > s_hi:
                    continue
                alloc = rr["amount"] if b["remaining"] >= rr["amount"] else b["remaining"]
                if alloc <= Decimal("0"):
                    continue
                rr["amount"] -= alloc
                b["remaining"] -= alloc
                b["allocated"] += alloc
                composition[b["id"]].extend(rr.get("ids") or [rr["id"]])

    # =========================
    # 4) Saída (df_bank)
    # =========================
    open_receb_total = sum((rr["amount"] for rr in remaining_receb if rr["amount"] > 0), Decimal("0"))
    carryover_bank_total = sum((b["remaining"] for b in bank_state if b["remaining"] > 0), Decimal("0"))

    out_rows = []
    for b in bank_state:
        amt = b["amount"]
        al = b["allocated"]
        rem = b["remaining"]
        if al == Decimal("0") and amt > 0:
            status = "❌ (nenhum recebível)"
        elif rem > Decimal("0"):
            status = "🟨 (parcial - saldo banco)"
        else:
            status = "✅"
        out_rows.append({
            "data": b["dt"].strftime("%d/%m/%Y"),
            "bank_id": b["id"],
            "banco": b["bank_name"],
            "memo": b["memo"],
            "valor_banco": float(amt),
            "alocado": float(al),
            "saldo_banco": float(rem),
            "qtd_receb": len(composition.get(b["id"], [])),
            "status": status,
        })

    df = pd.DataFrame(out_rows)
    if not df.empty:
        df = df.sort_values(by=["data", "bank_id"], kind="mergesort").reset_index(drop=True)

    # atributos úteis
    df.attrs["open_receb_total"] = float(open_receb_total)
    df.attrs["carryover_bank_total"] = float(carryover_bank_total)
    df.attrs["spillover_days"] = spillover_days
    df.attrs["window_days"] = window_days
    df.attrs["bank_end"] = bank_end
    df.attrs["month_end"] = end_date
    return df, composition



def run_step3_daily_view(
    conn: sqlite3.Connection,
    provider: str,
    month: date,
    window_days: int = 3,
    bank_term: str | None = None,
    spillover_days: int = 3,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[int, list[int]]]:
    """Etapa 3 (visão diária) — **por mês de pagamento (pay_dt)**.

    - Filtro mm/aaaa define o mês do **pay_dt** (recebimento agendado/pago).
    - Ajuste automático de finais de semana/feriados nacionais:
        dt_efetiva = next_business_day(pay_dt)
      (é essa data efetiva que manda na busca/alocação no banco).
    - Janela (dias) aplica tolerância de data na alocação (bank_dt ± window_days).
    - Spillover só amplia a busca NO BANCO se faltar fechar o esperado (não estica recebíveis).

    Retorna:
      df_day (diário), df_bank (detalhe banco), comp (bank_id -> receb_ids)
    """
    prov = (provider or '').strip().upper() or 'ALELO'
    # Ticket: taxa de transferência (subtrair 1x por reembolso na Etapa 3)
    transfer_fee = Decimal('0')
    if prov == 'TICKET':
        rules = fee_rules_tx_list(conn, prov)
        rule = fee_rule_tx_match(rules, 'VOUCHER') if rules else None
        if rule is None and rules:
            rule = next((r for r in rules if r.get('is_active')), None)
        if rule:
            try:
                transfer_fee = Decimal(str(rule.get('transfer_fee', 0) or 0))
            except Exception:
                transfer_fee = Decimal('0')
    start, end_excl = month_range(month)
    end_date = end_excl - timedelta(days=1)

    window_days = int(max(0, window_days))
    spillover_days = int(max(0, spillover_days))

    cols_r = {r[1] for r in conn.execute("PRAGMA table_info(receb_tx)").fetchall()}
    dt_target_expr = "COALESCE(pay_dt, dt)" if "pay_dt" in cols_r else "dt"

    def _pick_expr(cols: set[str], candidates: list[str], default: str) -> str:
        for c in candidates:
            if c in cols:
                return c
        return default

    amt_expr = _pick_expr(cols_r, ["liquido", "amount", "valor_liquido", "valor_liq", "net", "net_amount"], "0")

    # Para TICKET: "Esperado" deve ser recalculado a partir do BRUTO + regras atuais (Taxas/Tarifas),
    # e NÃO do valor líquido armazenado (que pode ter sido importado com outra regra).
    if prov == 'TICKET':
        rules = fee_rules_tx_list(conn, prov)
        rule = fee_rule_tx_match(rules, 'VOUCHER') if rules else None
        if rule is None and rules:
            rule = next((r for r in rules if isinstance(r, dict) and r.get('is_active')), None)
        try:
            mdr_pct = Decimal(str((rule or {}).get('mdr_percent', 0) or 0).replace(',', '.'))
        except Exception:
            mdr_pct = Decimal('0')
        try:
            fee_fixed = Decimal(str((rule or {}).get('fee_fixed', 0) or 0).replace(',', '.'))
        except Exception:
            fee_fixed = Decimal('0')

        has_raw = "raw_json" in cols_r
        exp_rows = conn.execute(
            f"SELECT id, dt AS dt_base, pay_dt AS pay_dt, {dt_target_expr} AS dt_target, bruto AS bruto, autorizacao AS autorizacao"
            + (", raw_json AS raw_json" if has_raw else "")
            + f" FROM receb_tx WHERE provider=? AND is_deleted=0 AND {dt_target_expr}>=? AND {dt_target_expr}<=? ORDER BY {dt_target_expr}, id",
            (prov, start.isoformat(), end_date.isoformat()),
        ).fetchall()
    else:
        exp_rows = conn.execute(
            f"SELECT id, dt AS dt_base, pay_dt AS pay_dt, {dt_target_expr} AS dt_target, {amt_expr} AS amount, autorizacao AS autorizacao "
            f"FROM receb_tx WHERE provider=? AND is_deleted=0 AND {dt_target_expr}>=? AND {dt_target_expr}<=? "
            f"ORDER BY {dt_target_expr}, id",
            (prov, start.isoformat(), end_date.isoformat()),
        ).fetchall()

    exp_by_day: dict[date, Decimal] = {}
    exp_last_dt = start
    # Ticket: controlar reembolsos por dia efetivo para subtrair a taxa de transferência 1x por lote
    reemb_by_day: dict[date, set[str]] = {}

    for r in exp_rows:
        r = dict(r)
        dt_target = parse_date(r.get("dt_target")) or parse_date(r.get("pay_dt")) or parse_date(r.get("dt_base"))
        if not dt_target:
            continue
        dt_eff = next_business_day(dt_target)
        if prov == 'TICKET':
            auth = str(r.get('autorizacao') or '')
            rid = auth.split('|', 1)[0].strip() if auth else ''
            if rid:
                reemb_by_day.setdefault(dt_eff, set()).add(rid)
        if prov == 'TICKET':
            bruto = parse_decimal(r.get('bruto')) or Decimal('0')
            if bruto <= 0:
                v = Decimal('0')
            else:
                qty = 1
                try:
                    j = json.loads(r.get('raw_json') or '{}') if 'raw_json' in r else {}
                    qty = int(j.get('qtd_transacoes') or j.get('qtd_transacoes_do_dia') or 1)
                except Exception:
                    qty = 1
                taxa = (bruto * (mdr_pct / Decimal('100'))).quantize(Decimal('0.01'))
                tarifa = (fee_fixed * Decimal(str(max(1, qty)))).quantize(Decimal('0.01'))
                v = (bruto - taxa - tarifa).quantize(Decimal('0.01'))
        else:
            v = parse_decimal(r.get("amount")) or Decimal('0')
        if v <= 0:
            continue
        exp_by_day[dt_eff] = exp_by_day.get(dt_eff, Decimal('0')) + v
        if dt_eff > exp_last_dt:
            exp_last_dt = dt_eff

    # Ticket: subtrai taxa de transferência 1x por reembolso (lote) na data efetiva de pagamento
    if prov == 'TICKET' and transfer_fee > 0:
        for d_eff, reembs in reemb_by_day.items():
            if not reembs:
                continue
            exp_by_day[d_eff] = (exp_by_day.get(d_eff, Decimal('0')) - (transfer_fee * Decimal(str(len(reembs))))).quantize(Decimal('0.01'))
            if exp_by_day[d_eff] < Decimal('0'):
                exp_by_day[d_eff] = Decimal('0')

    if not exp_by_day:
        exp_last_dt = end_date

    # 1ª passada: banco só até o último esperado
    bank_start = start - timedelta(days=max(0, window_days))
    df_bank, comp = run_step3_banco(
        conn, prov, month,
        window_days=window_days,
        bank_term=bank_term,
        spillover_days=0,
        bank_start_override=bank_start,
        bank_end_override=end_date,
        expected_end=exp_last_dt,
    )

    tot_exp = sum(exp_by_day.values(), Decimal('0'))
    tot_bnk = Decimal('0')
    if df_bank is not None and not df_bank.empty:
        tot_bnk = sum([parse_decimal(x) or Decimal('0') for x in df_bank['valor_banco'].tolist()], Decimal('0'))

    used_spillover = False
    if tot_bnk < tot_exp and spillover_days > 0:
        used_spillover = True
        bank_end2 = max(end_date, exp_last_dt) + timedelta(days=spillover_days)
        df_bank, comp = run_step3_banco(
            conn, prov, month,
            window_days=window_days,
            bank_term=bank_term,
            spillover_days=0,
            bank_start_override=bank_start,
            bank_end_override=bank_end2,
            expected_end=exp_last_dt,
        )

    bank_by_day: dict[date, Decimal] = {}
    alloc_by_day: dict[date, Decimal] = {}
    rem_by_day: dict[date, Decimal] = {}
    bank_end = exp_last_dt

    if df_bank is not None and not df_bank.empty:
        for _, rr in df_bank.iterrows():
            d = parse_br_date_str(str(rr.get("data") or ""))
            if not d:
                continue
            vb = parse_decimal(rr.get("valor_banco")) or Decimal("0")
            al = parse_decimal(rr.get("alocado")) or Decimal("0")
            sb = parse_decimal(rr.get("saldo_banco")) or Decimal("0")
            bank_by_day[d] = bank_by_day.get(d, Decimal("0")) + vb
            alloc_by_day[d] = alloc_by_day.get(d, Decimal("0")) + al
            rem_by_day[d] = rem_by_day.get(d, Decimal("0")) + sb
            if d > bank_end:
                bank_end = d

    end_view = max(exp_last_dt, bank_end)
    all_days = pd.date_range(start=start, end=end_view, freq="D")

    saldo_inicial = get_confirmed_carryover(conn, prov, month)
    saldo = saldo_inicial
    saldo_mes = saldo_inicial

    rows_out = []
    for ts in all_days:
        d = ts.date()
        exp = exp_by_day.get(d, Decimal("0"))
        bnk = bank_by_day.get(d, Decimal("0"))
        al = alloc_by_day.get(d, Decimal("0"))
        sb = rem_by_day.get(d, Decimal("0"))

        if exp == 0 and bnk == 0 and al == 0 and sb == 0:
            continue

        dif = bnk - exp
        saldo += dif
        if d <= end_date:
            saldo_mes = saldo

        if exp > 0 and bnk > 0 and exp == bnk and sb == 0:
            st = "🟢 OK"
        elif exp > 0 and bnk == 0:
            st = "🔴 Atraso"
        elif exp == 0 and bnk > 0:
            st = "🟡 Crédito sem esperado"
        elif exp > 0 and bnk > 0 and exp != bnk:
            st = "🟡 Diferença no dia"
        else:
            st = "⚪ Info"

        if sb > 0:
            st = f"{st} | 🟨 saldo banco em aberto"

        rows_out.append({
            "data": d,
            "esperado": exp,
            "banco": bnk,
            "diferenca": dif,
            "saldo_acum": saldo,
            "alocado": al,
            "saldo_banco": sb,
            "status": st,
        })

    df_day = pd.DataFrame(rows_out)
    if not df_day.empty:
        df_day = df_day.sort_values(by=["data"], kind="mergesort").reset_index(drop=True)

    df_day.attrs["start"] = start
    df_day.attrs["month_end"] = end_date
    df_day.attrs["expected_end"] = exp_last_dt
    df_day.attrs["end_view"] = end_view
    df_day.attrs["used_spillover"] = bool(used_spillover)
    df_day.attrs["saldo_inicial"] = float(saldo_inicial)
    df_day.attrs["saldo_mes"] = float(saldo_mes)
    return df_day, (df_bank if df_bank is not None else pd.DataFrame()), (comp or {})


def run_step3_monthly(conn: sqlite3.Connection, provider: str, month: date, bank_keyword: str | None = None) -> dict:
    start, end_excl = month_range(month)
    end = end_excl - timedelta(days=1)
    def sum_tbl(tbl: str, col: str, prov: str, extra_where: str="", params=(), date_col: str = "dt"):
        q = f"SELECT COALESCE(SUM(COALESCE({col},0)),0) AS s FROM {tbl} WHERE provider=? AND is_deleted=0 AND {date_col}>=? AND {date_col}<=? {extra_where}"
        row = conn.execute(q, (prov, start.isoformat(), end.isoformat(), *params)).fetchone()
        return parse_decimal(row["s"]) or Decimal("0")
    vendas_liq = sum_tbl("sales_tx","liquido", provider)
    receb_liq = sum_tbl("receb_tx","liquido", provider, date_col="COALESCE(pay_dt, dt)")
    if bank_keyword:
        bk = bank_keyword.strip().upper()
        if bk == BANK_TRANSF_TOKEN:
            # OR de termos conhecidos de repasse (Alelo/Naip)
            ors = " OR ".join(["UPPER(COALESCE(memo,'')) LIKE ?"] * len(bank_all_memo_terms(prov)))
            params = tuple([f"%{t.upper()}%" for t in bank_all_memo_terms(prov)])
            banco = sum_tbl("bank_tx","amount","BANCO", f"AND ({ors})", params)
        else:
            kw = f"%{bank_keyword.strip().upper()}%"
            banco = sum_tbl("bank_tx","amount","BANCO","AND UPPER(COALESCE(memo,'')) LIKE ?", (kw,))
    else:
        banco = sum_tbl("bank_tx","amount","BANCO")

    return {
        "mes": f"{month.month:02d}/{month.year}",
        "vendas_liq": vendas_liq,
        "receb_liq": receb_liq,
        "banco": banco,
        "div_banco_vs_receb": banco - receb_liq,
        "div_receb_vs_vendas": receb_liq - vendas_liq
    }

def is_month_closed(conn: sqlite3.Connection, provider: str, month_mm_yyyy: str) -> bool:
    row = conn.execute("SELECT 1 FROM closed_periods WHERE provider=? AND month=? LIMIT 1", (provider, month_mm_yyyy)).fetchone()
    return row is not None

def close_month(conn: sqlite3.Connection, provider: str, month_mm_yyyy: str) -> None:
    conn.execute("INSERT OR IGNORE INTO closed_periods(provider, month, closed_at) VALUES (?,?,?)",
                 (provider, month_mm_yyyy, datetime.now().isoformat(sep=" ", timespec="seconds")))
    conn.commit()

def undo_month(conn: sqlite3.Connection, provider: str, month_mm_yyyy: str) -> None:
    conn.execute("DELETE FROM closed_periods WHERE provider=? AND month=?", (provider, month_mm_yyyy))
    conn.commit()


def _month_mm_yyyy(d: date) -> str:
    return f"{d.month:02d}/{d.year}"

def get_confirmed_carryover(conn: sqlite3.Connection, provider: str, month: date) -> Decimal:
    """Retorna o saldo confirmado do mês anterior (que entra como saldo inicial do mês atual)."""
    # mês anterior
    prev = (month.replace(day=1) - timedelta(days=1)).replace(day=1)
    mm = _month_mm_yyyy(prev)
    row = conn.execute(
        "SELECT amount FROM carryover_balances WHERE provider=? AND month=?",
        ((provider or '').strip().upper() or 'ALELO', mm),
    ).fetchone()
    return parse_decimal(row["amount"]) if row else Decimal("0")

def set_confirmed_carryover(conn: sqlite3.Connection, provider: str, month: date, amount: Decimal) -> None:
    """Confirma o saldo do mês (para ser usado no mês seguinte)."""
    mm = _month_mm_yyyy(month)
    conn.execute(
        "INSERT INTO carryover_balances(provider, month, amount, confirmed_at) VALUES (?,?,?,?) "
        "ON CONFLICT(provider, month) DO UPDATE SET amount=excluded.amount, confirmed_at=excluded.confirmed_at",
        ((provider or '').strip().upper() or 'ALELO', mm, str(amount.quantize(Decimal('0.01'))), datetime.now().isoformat(sep=' ', timespec='seconds')),
    )
    conn.commit()


# ---------------- UI state persistence (simple key/value) ----------------
def ui_get(conn: sqlite3.Connection, key: str, default: str | None = None) -> str | None:
    try:
        row = conn.execute("SELECT value FROM ui_kv WHERE key=?", (key,)).fetchone()
        return row[0] if row else default
    except Exception:
        return default

def ui_set(conn: sqlite3.Connection, key: str, value: str) -> None:
    try:
        now = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            "INSERT INTO ui_kv(key, value, updated_at) VALUES (?,?,?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
            (key, value, now),
        )
        conn.commit()
    except Exception:
        pass

def first_imported_date(conn: sqlite3.Connection, provider: str) -> date | None:
    q = """SELECT MIN(dt) AS m FROM (
      SELECT MIN(dt) AS dt FROM erp_tx WHERE provider=? AND is_deleted=0
      UNION ALL
      SELECT MIN(dt) AS dt FROM sales_tx WHERE provider=? AND is_deleted=0
      UNION ALL
      SELECT MIN(dt) AS dt FROM receb_tx WHERE provider=? AND is_deleted=0
    )"""
    row = conn.execute(q, (provider, provider, provider)).fetchone()
    return parse_date(row["m"]) if row else None

def export_tree_to_excel(tree: ttk.Treeview, filepath: str) -> None:
    cols = list(tree["columns"])
    data = []
    for item in tree.get_children():
        data.append(list(tree.item(item, "values")))
    pd.DataFrame(data, columns=[c.upper() for c in cols]).to_excel(filepath, index=False)

class EditRecordPopup(tk.Toplevel):
    def __init__(self, master, conn: sqlite3.Connection, table: str, rec_id: int, refresh_cb):
        super().__init__(master)
        self.transient(master)
        self.grab_set()
        self.focus_set()
        self.title("Editar / Excluir registro")
        self.conn = conn
        self.table = table
        self.rec_id = rec_id
        self.refresh_cb = refresh_cb
        row = conn.execute(f"SELECT * FROM {table} WHERE id=?", (rec_id,)).fetchone()
        if not row:
            messagebox.showerror("Erro", "Registro não encontrado.")
            self.destroy(); return
        frm = ttk.Frame(self, padding=10); frm.pack(fill="both", expand=True)
        ttk.Label(frm, text=f"Tabela: {table} | ID: {rec_id}", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,8))
        self.var_dt = tk.StringVar(value=str(row["dt"]))
        self.var_pay_dt = tk.StringVar(value=str(row["pay_dt"]) if "pay_dt" in row.keys() else "")
        self.var_bruto = tk.StringVar(value=str(row["bruto"]) if "bruto" in row.keys() else "")
        self.var_liq = tk.StringVar(value=str(row["liquido"]) if "liquido" in row.keys() else "")
        self.var_amount = tk.StringVar(value=str(row["amount"]) if "amount" in row.keys() else "")
        self.var_auth = tk.StringVar(value=str(row["autorizacao"]) if "autorizacao" in row.keys() else "")
        self.var_memo = tk.StringVar(value=str(row["memo"]) if "memo" in row.keys() else "")
        r = 1
        ttk.Label(frm, text="Data (aaaa-mm-dd ou dd/mm/aaaa):").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_dt, width=22).grid(row=r, column=1, sticky="w"); r += 1
        if "pay_dt" in row.keys():
            ttk.Label(frm, text="Data de Pagamento (opcional):").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_pay_dt, width=22).grid(row=r, column=1, sticky="w"); r += 1
        if "bruto" in row.keys():
            ttk.Label(frm, text="Bruto:").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_bruto, width=22).grid(row=r, column=1, sticky="w"); r += 1
        if "liquido" in row.keys():
            ttk.Label(frm, text="Líquido:").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_liq, width=22).grid(row=r, column=1, sticky="w"); r += 1
        if "amount" in row.keys():
            ttk.Label(frm, text="Valor (Banco):").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_amount, width=22).grid(row=r, column=1, sticky="w"); r += 1
        if "autorizacao" in row.keys():
            ttk.Label(frm, text="Autorização/NSU:").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_auth, width=32).grid(row=r, column=1, sticky="w"); r += 1
        if "memo" in row.keys():
            ttk.Label(frm, text="Histórico/Memo:").grid(row=r, column=0, sticky="w"); ttk.Entry(frm, textvariable=self.var_memo, width=70).grid(row=r, column=1, sticky="w"); r += 1
        btns = ttk.Frame(frm); btns.grid(row=r, column=0, columnspan=2, sticky="e", pady=(10,0))
        ttk.Button(btns, text="Salvar (Editar)", command=self._save).pack(side="left", padx=5)
        ttk.Button(btns, text="Excluir", command=self._delete).pack(side="left", padx=5)
        ttk.Button(btns, text="Fechar", command=self.destroy).pack(side="left", padx=5)
        self.resizable(False, False)
    def _save(self):
        d = parse_date(self.var_dt.get())
        if not d:
            messagebox.showerror("Erro", "Data inválida."); return
        row = self.conn.execute(f"SELECT * FROM {self.table} WHERE id=?", (self.rec_id,)).fetchone()
        sets = ["dt=?"]; params = [d.isoformat()]
        if "pay_dt" in row.keys():
            pdx = parse_date(self.var_pay_dt.get()) if (self.var_pay_dt.get() or "").strip() else None
            sets.append("pay_dt=?"); params.append(pdx.isoformat() if pdx else None)
        def add_num(field, var):
            v = parse_decimal(var.get()); sets.append(f"{field}=?"); params.append(str(v) if v is not None else None)
        def add_txt(field, var):
            sets.append(f"{field}=?"); params.append((var.get() or "").strip())
        if "bruto" in row.keys(): add_num("bruto", self.var_bruto)
        if "liquido" in row.keys(): add_num("liquido", self.var_liq)
        if "amount" in row.keys(): add_num("amount", self.var_amount)
        if "autorizacao" in row.keys(): add_txt("autorizacao", self.var_auth)
        if "memo" in row.keys(): add_txt("memo", self.var_memo)
        params.append(self.rec_id)
        self.conn.execute(f"UPDATE {self.table} SET {', '.join(sets)} WHERE id=?", tuple(params))
        self.conn.commit(); self.refresh_cb(); self.destroy()
    def _delete(self):
        if not messagebox.askyesno("Confirmar", "Excluir este registro?"):
            return
        self.conn.execute(f"UPDATE {self.table} SET is_deleted=1 WHERE id=?", (self.rec_id,))
        self.conn.commit(); self.refresh_cb(); self.destroy()

class UnderlyingByDayPopup(tk.Toplevel):
    def __init__(self, master, conn: sqlite3.Connection, provider: str, day: date, table: str, refresh_parent_cb):
        super().__init__(master)
        self.transient(master)
        self.grab_set()
        self.focus_set()
        self.title(f"Registros do dia {fmt_br_date(day)} - {table}")
        self.conn = conn; self.provider = provider; self.day = day; self.table = table; self.refresh_parent_cb = refresh_parent_cb
        frm = ttk.Frame(self, padding=10); frm.pack(fill="both", expand=True)
        ttk.Label(frm, text=f"{table} | {fmt_br_date(day)}", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,8))
        cols = ["id","dt","amount","memo","bank_id"] if table == "bank_tx" else ["id","dt","bruto","liquido","autorizacao"]
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=160 if c in ("memo",) else 120, anchor="w" if c in ("memo","autorizacao","bank_id") else "e")
        self.tree.column("id", width=70, anchor="e"); self.tree.column("dt", width=110, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self._dbl_edit)
        btns = ttk.Frame(frm); btns.pack(fill="x", pady=(8,0))
        ttk.Button(btns, text="Editar selecionado", command=self._edit_selected).pack(side="left", padx=(0,6))
        ttk.Button(btns, text="Excluir selecionados", command=self._bulk_delete).pack(side="left")
        ttk.Button(btns, text="Fechar", command=self.destroy).pack(side="right")
        self._reload()
    def _selected_id(self):
        sel = self.tree.selection()
        if not sel: return None
        return int(self.tree.item(sel[0], "values")[0])
    def _reload(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        q = f"SELECT * FROM {self.table} WHERE is_deleted=0 AND dt=?"
        params = [self.day.isoformat()]
        if self.table != "bank_tx":
            q += " AND provider=?"; params.append(self.provider)
        q += " ORDER BY id"
        for r in self.conn.execute(q, tuple(params)).fetchall():
            if self.table == "bank_tx":
                self.tree.insert("", "end", values=(r["id"], r["dt"], r["amount"], (r["memo"] or "")[:120], r["bank_id"] or ""))
            else:
                self.tree.insert("", "end", values=(r["id"], r["dt"], r["bruto"], r["liquido"], r["autorizacao"] or ""))
    def _open_editor(self, rec_id: int):
        def refresh():
            self._reload(); self.refresh_parent_cb()
        EditRecordPopup(self, self.conn, self.table, rec_id, refresh)
    def _dbl_edit(self, _evt):
        rid = self._selected_id()
        if rid is None: return
        self._open_editor(rid)
    def _edit_selected(self):
        rid = self._selected_id()
        if rid is None:
            messagebox.showinfo("Info","Selecione um registro."); return
        self._open_editor(rid)
    def _bulk_delete(self):
        ids = [int(self.tree.item(item, "values")[0]) for item in self.tree.selection()]
        if not ids:
            messagebox.showinfo("Info","Selecione um ou mais registros."); return
        if not messagebox.askyesno("Confirmar", f"Excluir {len(ids)} registro(s)?"):
            return
        for rid in ids:
            self.conn.execute(f"UPDATE {self.table} SET is_deleted=1 WHERE id=?", (rid,))
        self.conn.commit(); self._reload(); self.refresh_parent_cb()

class BulkDeletePopup(tk.Toplevel):
    def __init__(self, master, conn: sqlite3.Connection, provider: str, refresh_cb, log_cb, err_cb):
        super().__init__(master)
        self.transient(master)
        self.grab_set()
        self.focus_set()
        self.title("Excluir em massa por período")
        self.conn = conn; self.provider = provider; self.refresh_cb = refresh_cb; self.log_cb = log_cb; self.err_cb = err_cb
        frm = ttk.Frame(self, padding=10); frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Tabela:", font=("Segoe UI",10,"bold")).grid(row=0, column=0, sticky="w")
        self.table = tk.StringVar(value="sales_tx")
        ttk.Combobox(frm, textvariable=self.table, values=["erp_tx","sales_tx","receb_tx","bank_tx"], width=18, state="readonly").grid(row=0, column=1, sticky="w", padx=6)
        ttk.Label(frm, text="Ou mês (mm/aaaa):").grid(row=1, column=0, sticky="w", pady=(8,0))
        self.mm = tk.StringVar(value=""); ttk.Entry(frm, textvariable=self.mm, width=12).grid(row=1, column=1, sticky="w", padx=6, pady=(8,0))
        ttk.Label(frm, text="Ou período (dd/mm/aaaa):").grid(row=2, column=0, sticky="w", pady=(8,0))
        p = ttk.Frame(frm); p.grid(row=2, column=1, sticky="w", padx=6, pady=(8,0))
        self.d1 = tk.StringVar(value=""); self.d2 = tk.StringVar(value="")
        ttk.Entry(p, textvariable=self.d1, width=12).pack(side="left"); ttk.Label(p, text=" até ").pack(side="left"); ttk.Entry(p, textvariable=self.d2, width=12).pack(side="left")
        ttk.Button(frm, text="Executar exclusão", command=self._run).grid(row=3, column=0, columnspan=2, sticky="e", pady=(12,0))
        self.resizable(False, False)
    def _run(self):
        table = self.table.get()
        mm = (self.mm.get() or "").strip()
        d1 = parse_br_date_str(self.d1.get()); d2 = parse_br_date_str(self.d2.get())
        if mm:
            m = month_start(mm)
            if not m:
                messagebox.showerror("Erro","Mês inválido (mm/aaaa)."); return
            start, end_excl = month_range(m); start_d = start; end_d = end_excl - timedelta(days=1)
        elif d1 and d2:
            start_d, end_d = (d1, d2) if d2 >= d1 else (d1, d1)
        else:
            messagebox.showerror("Erro","Informe mm/aaaa OU período (dd/mm/aaaa a dd/mm/aaaa)."); return
        try:
            if table == "bank_tx":
                self.conn.execute("UPDATE bank_tx SET is_deleted=1 WHERE provider='BANCO' AND is_deleted=0 AND dt>=? AND dt<=?",
                                  (start_d.isoformat(), end_d.isoformat()))
            else:
                self.conn.execute(f"UPDATE {table} SET is_deleted=1 WHERE provider=? AND is_deleted=0 AND dt>=? AND dt<=?",
                                  (self.provider, start_d.isoformat(), end_d.isoformat()))
            self.conn.commit()
            self.log_cb(f"Exclusão em massa OK | {table} | {fmt_br_date(start_d)} a {fmt_br_date(end_d)}")
            self.refresh_cb(); self.destroy()
        except Exception as e:
            self.err_cb(f"Falha na exclusão em massa: {e}")


# ==========================================================
# TAXAS/TARIFAS (Etapa 4) - Helpers e CRUD
# ==========================================================
import json
import unicodedata

def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().upper()

def iso_to_br(dt_iso: str | None) -> str:
    if not dt_iso:
        return ""
    try:
        # aceita YYYY-MM-DD ou YYYY-MM-DD HH:MM:SS
        part = str(dt_iso).split(" ")[0]
        y, m, d = part.split("-")
        return f"{d.zfill(2)}/{m.zfill(2)}/{y}"
    except Exception:
        return str(dt_iso)


def extract_ticket_fields(raw):
    """Return dict with possible Ticket fields from raw_json/text."""
    if raw is None:
        return {}
    if isinstance(raw, dict):
        d = raw
    else:
        try:
            d = json.loads(raw) if isinstance(raw, str) and raw.strip().startswith(("{","[")) else {}
        except Exception:
            d = {}
    # normalize keys to simplify lookup
    out = {}
    for k,v in (d.items() if isinstance(d, dict) else []):
        kk = str(k).strip().lower()
        out[kk] = v
    return out

def ticket_get(d, *keys):
    for k in keys:
        kk = str(k).strip().lower()
        if kk in d and d[kk] not in (None, ""):
            return d[kk]
    return None

def extract_tipo_cartao(raw_json: str | None) -> str:
    if not raw_json:
        return ""
    try:
        obj = json.loads(raw_json)
        if isinstance(obj, dict):
            for k in ("Tipo Cartão","Tipo Cartao","tipo_cartao","tipo_cartao_desc","TIPO CARTAO","tipo"):
                if k in obj and obj[k]:
                    return str(obj[k])
        return ""
    except Exception:
        return str(raw_json)

def extract_reembolso(raw_json: str | None) -> str:
    if not raw_json:
        return ""
    try:
        obj = json.loads(raw_json)
        if isinstance(obj, dict):
            for k in ("Número do reembolso","Numero do reembolso","numero_reembolso","reembolso","NUMERO DO REEMBOLSO"):
                if k in obj and obj[k]:
                    return str(obj[k])
        return ""
    except Exception:
        return ""

def fee_rules_tx_list(conn: sqlite3.Connection, provider: str):
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT id, provider, label, match_text, mdr_percent, fee_fixed, transfer_fee, is_active
        FROM fee_rules_tx
        WHERE provider=?
        ORDER BY id ASC
        """,
        (provider,)
    ).fetchall()
    out = []
    for r in rows:
        out.append({
            "id": r[0], "provider": r[1], "label": r[2], "match_text": r[3],
            "mdr_percent": r[4], "fee_fixed": r[5], "transfer_fee": r[6], "is_active": int(r[7] or 0)
        })
    return out

def fee_rule_tx_upsert(conn: sqlite3.Connection, provider: str, label: str, match_text: str,
                      mdr_percent: Decimal, fee_fixed: Decimal, transfer_fee: Decimal,
                      is_active: int, rule_id: int | None = None):
    cur = conn.cursor()
    if rule_id is None:
        cur.execute(
            """INSERT INTO fee_rules_tx(provider,label,match_text,mdr_percent,fee_fixed,transfer_fee,is_active)
                 VALUES(?,?,?,?,?,?,?)""",
            (provider, label, match_text, float(mdr_percent), float(fee_fixed), float(transfer_fee), int(is_active))
        )
    else:
        cur.execute(
            """UPDATE fee_rules_tx
                 SET label=?, match_text=?, mdr_percent=?, fee_fixed=?, transfer_fee=?, is_active=?
                 WHERE id=? AND provider=?""",
            (label, match_text, float(mdr_percent), float(fee_fixed), float(transfer_fee), int(is_active), int(rule_id), provider)
        )
    conn.commit()

def fee_rule_tx_delete(conn: sqlite3.Connection, rule_id: int):
    conn.execute("DELETE FROM fee_rules_tx WHERE id=?", (int(rule_id),))
    conn.commit()

def fee_rule_tx_match(rules: list[dict], tipo_cartao: str):
    t = normalize_text(tipo_cartao)
    best = None
    best_len = -1
    for r in rules:
        if not r.get("is_active"):
            continue
        mtxt = normalize_text(r.get("match_text",""))
        if not mtxt:
            continue
        if mtxt in t:
            if len(mtxt) > best_len:
                best = r
                best_len = len(mtxt)
    return best

def seed_rules_alelo_if_empty(conn: sqlite3.Connection):
    rules = fee_rules_tx_list(conn, "ALELO")
    if rules:
        return False
    fee_rule_tx_upsert(conn, "ALELO", "Alelo Alimentação", "Alimentação", Decimal("7.0"), Decimal("0.99"), Decimal("0"), 1, None)
    fee_rule_tx_upsert(conn, "ALELO", "Alelo Multibenefícios", "Multibeneficios", Decimal("4.5"), Decimal("0"), Decimal("0"), 1, None)
    return True

def seed_rules_ticket_if_empty(conn: sqlite3.Connection):
    rules = fee_rules_tx_list(conn, "TICKET")
    if rules:
        return False
    # padrão: Voucher (usuário pode editar)
    fee_rule_tx_upsert(conn, "TICKET", "Ticket Voucher", "VOUCHER", Decimal("4.5"), Decimal("0.52"), Decimal("0"), 1, None)
    return True



def seed_rules_farmaciasapp_if_empty(conn: sqlite3.Connection):
    rules = fee_rules_tx_list(conn, "FARMACIASAPP")
    if rules:
        return False
    # Padrão sugerido: Taxa ADM 5% sobre o TOTAL DA COMPRA; Tx Adquirente 1% sobre o VALOR PAGO NA LOJA.
    # Usamos:
    #   mdr_percent   -> taxa_adm_percent
    #   fee_fixed     -> tx_adquirente_percent  (em %)
    #   transfer_fee  -> tarifa_fixa (R$) por transação (opcional; default 0)
    fee_rule_tx_upsert(conn, "FARMACIASAPP", "PIX", "PIX", Decimal("5.0"), Decimal("1.0"), Decimal("0"), 1, None)
    fee_rule_tx_upsert(conn, "FARMACIASAPP", "CREDIT", "CREDIT", Decimal("5.0"), Decimal("1.0"), Decimal("0"), 1, None)
    return True


# ==========================================================
# ETAPA 4 (Taxas/Tarifas) - runner
# ==========================================================
def run_step4(app):
    if not hasattr(app, "conn") or app.conn is None:
        app.conn = connect(app.db_path.get()); init_db(app.conn)

    app_month = app._parse_month()
    if not app_month:
        return
    d1, d2 = app._parse_period()
    start, end = _get_period(app_month, d1, d2)
    prov = (app.provider.get() or "ALELO").strip().upper()

    tol_s = ui_get(app.conn, "fees_tolerance", "0.05")
    try:
        tol = Decimal(str(tol_s).replace(",", "."))
    except Exception:
        tol = Decimal("0.05")

        # ------------------------------------------------------------
    # FARMACIASAPP - auditoria por pedido (taxa esperada x aplicada)
    #   Esperado (R$) = (taxa_adm% * Valor Total da Compra) + (tx_adquirente% * Valor Pago na Loja) + tarifa_fixa
    #   Banco/Repasse esperado = Valor Total - Taxas
    # ------------------------------------------------------------
    if provider == "FARMACIASAPP":
        # regra por forma (PIX/CREDIT), fallback para a primeira ativa
        tol_s = ui_get(app.conn, "fees_tolerance", "0.05")
        try:
            tol = Decimal(str(tol_s).replace(",", "."))
        except Exception:
            tol = Decimal("0.05")

        # período: usa dt (data do pedido) como eixo principal
        rows = app.conn.execute(
            "SELECT id, dt, pay_dt, bruto, liquido, autorizacao, raw_json FROM receb_tx "
            "WHERE provider=? AND is_deleted=0 AND dt BETWEEN ? AND ? ORDER BY dt, id",
            (prov, start.isoformat(), end.isoformat())
        ).fetchall()

        tree = app.s4_tree
        for iid in tree.get_children():
            tree.delete(iid)

        out = []
        tot_exp = Decimal("0")
        tot_app = Decimal("0")
        div_cnt = 0

        # pega regra default
        active_rules = [r for r in rules if isinstance(r, dict) and r.get("is_active")]
        default_rule = active_rules[0] if active_rules else (rules[0] if rules else None)

        for rr in rows:
            rr = dict(rr)
            sale_dt = rr.get("dt") or ""
            pay_dt = rr.get("pay_dt") or ""
            bruto = parse_decimal(rr.get("bruto")) or Decimal("0")   # Valor Total da Compra
            liq = parse_decimal(rr.get("liquido")) or Decimal("0")   # Valor de Repasse (aplicado)

            auth = (rr.get("autorizacao") or "").strip()
            # extrai valor_pago_loja e forma do raw_json
            forma = ""
            pago_loja = None
            try:
                j = json.loads(rr.get("raw_json") or "{}")
            except Exception:
                j = {}
            if isinstance(j, dict):
                forma = str(j.get("forma_pagamento") or "").strip().upper()
                pago_loja = j.get("valor_pago_loja")

            pago_loja_d = parse_decimal(pago_loja) or bruto  # fallback

            # regra por forma
            rule = fee_rule_tx_match(rules, forma) if forma else None
            if rule is None:
                rule = default_rule
            if isinstance(rule, dict):
                taxa_adm_pct = parse_decimal(rule.get("mdr_percent")) or Decimal("0")
                tx_adq_pct = parse_decimal(rule.get("fee_fixed")) or Decimal("0")  # em %
                tarifa_fixa = parse_decimal(rule.get("transfer_fee")) or Decimal("0")  # R$ por transação
            else:
                taxa_adm_pct = Decimal("0"); tx_adq_pct = Decimal("0"); tarifa_fixa = Decimal("0")

            taxa_esp = (bruto * (taxa_adm_pct/Decimal("100"))).quantize(Decimal("0.01"))
            taxa_esp += (pago_loja_d * (tx_adq_pct/Decimal("100"))).quantize(Decimal("0.01"))
            taxa_esp += tarifa_fixa.quantize(Decimal("0.01"))

            esp_banco = (bruto - taxa_esp).quantize(Decimal("0.01"))
            taxa_aplic = (bruto - liq).quantize(Decimal("0.01"))
            dif = (liq - esp_banco).quantize(Decimal("0.01"))

            status = "OK"
            tag = ""
            if abs(dif) > tol:
                status = "DIV"
                tag = "div"
                div_cnt += 1

            out.append({
                "sale_dt": fmt_br_date(parse_date(sale_dt)) if sale_dt else "",
                "pay_dt": fmt_br_date(parse_date(pay_dt)) if pay_dt else "",
                "tipo": forma or "-",
                "nsu": auth or f"ID:{rr.get('id')}",
                "bruto": float(bruto),
                "liquido": float(liq),
                "taxa_aplic": float(taxa_aplic),
                "taxa_esp": float(taxa_esp),
                "transf": float(tarifa_fixa),
                "esp_banco": float(esp_banco),
                "dif": float(dif),
                "status": status,
            })

            tot_exp += taxa_esp
            tot_app += taxa_aplic

        # render
        for rowd in out:
            tree.insert(
                "", "end", values=(
                    rowd["sale_dt"], rowd["pay_dt"], rowd["tipo"], rowd["nsu"],
                    br_money(Decimal(str(rowd["bruto"]))), br_money(Decimal(str(rowd["liquido"]))),
                    br_money(Decimal(str(rowd["taxa_aplic"]))), br_money(Decimal(str(rowd["taxa_esp"]))),
                    br_money(Decimal(str(rowd["transf"]))), br_money(Decimal(str(rowd["esp_banco"]))),
                    br_money(Decimal(str(rowd["dif"]))), rowd["status"]
                ), tags=(tag,) if (tag := ("div" if rowd["status"]=="DIV" else "")) else ()
            )

        diff_tot = (tot_app - tot_exp).quantize(Decimal("0.01"))
        app.s4_tot_var.set(f"Taxa esperada: {br_money(tot_exp)} | Taxa aplicada: {br_money(tot_app)} | Diferença: {br_money(diff_tot)}")
        app.s4_div_var.set(f"Divergências: {div_cnt}")
        app._s4_last_df = pd.DataFrame(out)
        app._log(f"Etapa 4 (FARMACIASAPP) atualizada. Registros: {len(out)}")
        return

    # ------------------------------------------------------------
    # TICKET (Voucher) - auditoria por LOTE (Número do reembolso)
    # Agrupamento: Data de Pagamento (crédito/débito) -> Reembolso -> Linhas COMPRA
    # ------------------------------------------------------------
    if prov == "TICKET":
        # regra única típica: match "VOUCHER"
        rule = fee_rule_tx_match(rules, "VOUCHER") or fee_rule_tx_match(rules, "TICKET") or (rules[0] if rules else None)
        if rule is None:
            messagebox.showwarning("Taxas/Tarifas", "Nenhuma regra cadastrada para TICKET. Cadastre em Taxas/Tarifas.")
            return

        # unpack rule safely (supports 5 or 7 cols)
        rid, rprov, rlabel, rmatch, rmdr, rfix, rtransf, ractive = (None, None, "", "", 0, 0, 0, 1)
        try:
            if len(rule) >= 8:
                rid, rprov, rlabel, rmatch, rmdr, rfix, rtransf, ractive = rule[:8]
            elif len(rule) == 7:
                rid, rprov, rlabel, rmatch, rmdr, rfix, ractive = rule
                rtransf = 0
            elif len(rule) == 5:
                rid, rmatch, rmdr, rfix, ractive = rule
                rlabel = "Ticket Voucher"
                rtransf = 0
        except Exception:
            pass

        mdr = Decimal(str(rmdr).replace(",", ".") or "0")
        fee_fixed = Decimal(str(rfix).replace(",", ".") or "0")  # por compra
        transfer_fee = Decimal(str(rtransf).replace(",", ".") or "0")  # por lote

        # período (usa pay_dt)
        rows = app.conn.execute(
            "SELECT id, pay_dt, dt, bruto, liquido, raw_json FROM receb_tx WHERE provider=? AND is_deleted=0 AND pay_dt BETWEEN ? AND ? ORDER BY pay_dt, dt, id",
            (prov, d1, d2)
        ).fetchall()

        # montar estrutura: pay_dt -> reembolso -> itens
        data = {}  # pay_dt -> reimb -> dict
        for _id, pay_dt, sale_dt, bruto, liquido, raw in rows:
            rawd = extract_ticket_fields(raw)
            descr = ticket_get(rawd, "descricao do lancamento", "descrição do lançamento", "descricao", "descricao_lancamento", "descricao do lançamento")
            descr_n = normalize_text(descr or "")
            reimb = ticket_get(rawd, "numero do reembolso", "número do reembolso", "reembolso", "numero_reembolso") or ""
            reimb = str(reimb).strip()
            if not reimb:
                # fallback: usa id como lote (não ideal, mas evita perder linhas)
                reimb = f"ROW-{_id}"

            # valor da transação: tenta bruto, senão tenta campos do raw
            val = bruto
            if val in (None, ""):
                v2 = ticket_get(rawd, "valor da transacao", "valor da transação", "valor", "valor_transacao")
                val = v2 if v2 not in (None, "") else 0

            try:
                val = Decimal(str(val).replace(".", "").replace(",", "."))
            except Exception:
                val = Decimal("0")

            # decidir se é compra / tpe / liquido
            is_compra = "COMPRA" in descr_n
            is_tpe = ("TPE" in descr_n)
            is_valor_liq = ("VALOR LIQ" in descr_n) or ("VALOR LÍQ" in descr_n)

            pay_key = pay_dt or ""
            sale_key = sale_dt or ""
            data.setdefault(pay_key, {})
            dpay = data[pay_key]
            dpay.setdefault(reimb, {"compras": [], "others": [], "liq_line": None})
            lot = dpay[reimb]
            rec = {"id": _id, "sale_dt": sale_key, "descr": descr or "", "val": val}
            if is_valor_liq:
                lot["liq_line"] = val
            if is_compra:
                lot["compras"].append(rec)
            else:
                # guarda outras linhas para eventualmente calcular líquido aplicado sem TPE
                if not is_tpe:  # ignora TPE e Taxa TPE
                    lot["others"].append(rec)

        # limpar tree
        tree = app.s4_tree
        for iid in tree.get_children():
            tree.delete(iid)

        tot_exp = Decimal("0")
        tot_app = Decimal("0")
        diff_tot = Decimal("0")

        for pay_dt in sorted(data.keys()):
            pay_iid = f"P:{pay_dt}"
            tree.insert("", "end", iid=pay_iid, text=iso_to_br(pay_dt), values=(iso_to_br(pay_dt), "", "", "", "", "", "", "", "", ""))
            day_exp = Decimal("0")
            day_app = Decimal("0")

            for reimb, lot in sorted(data[pay_dt].items()):
                compras = lot["compras"]
                if not compras:
                    continue
                qtd = len(compras)
                subtotal = sum((r["val"] for r in compras), Decimal("0"))

                # esperado (contrato): por compra
                taxa_adm = (subtotal * (mdr/Decimal("100")))
                tarifa = (fee_fixed * Decimal(qtd))
                exp_fee = taxa_adm + tarifa
                exp_liq = subtotal - exp_fee
                exp_bank = exp_liq - transfer_fee

                # aplicado: se existir linha "valor líquido", usa; senão subtotal - sum(others)
                if lot["liq_line"] is not None:
                    app_liq = lot["liq_line"]
                else:
                    descontos = sum((r["val"] for r in lot["others"]), Decimal("0"))
                    app_liq = subtotal - descontos

                diff = app_liq - exp_bank
                status = "OK" if abs(diff) <= tol else "DIV"

                day_exp += exp_bank
                day_app += app_liq

                # linha lote (filho)
                lot_iid = f"L:{pay_dt}:{reimb}"
                tree.insert(pay_iid, "end", iid=lot_iid, text=str(reimb), values=(
                    iso_to_br(pay_dt),
                    str(reimb),
                    f"{qtd}",
                    br_money(subtotal),
                    "",  # bruto não exibido
                    br_money(app_liq),
                    br_money(exp_bank),
                    br_money(diff),
                    status,
                    "TICKET"
                ), tags=("div",) if status=="DIV" else ())

                # detalhes (neto) com datas de venda
                for r in sorted(compras, key=lambda x: x["sale_dt"] or ""):
                    tree.insert(lot_iid, "end", text="", values=(
                        iso_to_br(r["sale_dt"]),
                        "",
                        "",
                        br_money(r["val"]),
                        "",
                        "",
                        "",
                        "",
                        "",
                        (r["descr"][:30] + "…") if len(r["descr"])>30 else r["descr"]
                    ))

            # totais do dia no text do pai
            tree.item(pay_iid, values=(iso_to_br(pay_dt), "", "", "", "", br_money(day_app), br_money(day_exp), br_money(day_app-day_exp), "", ""))

            tot_exp += day_exp
            tot_app += day_app

        diff_tot = tot_app - tot_exp
        try:
            app.s4_tot_var.set(f"Aplicado: {br_money(tot_app)} | Esperado Banco: {br_money(tot_exp)} | Diferença: {br_money(diff_tot)} | Tol: {br_money(tol)}")
        except Exception:
            pass

        return


    q = ("SELECT id, dt, pay_dt, bruto, liquido, autorizacao, raw_json "
         "FROM receb_tx "
         "WHERE provider=? AND is_deleted=0 AND dt BETWEEN ? AND ? "
         "ORDER BY dt ASC, autorizacao ASC, id ASC")
    rows = app.conn.execute(q, (prov, start.isoformat(), end.isoformat())).fetchall()

    try:
        app.s4_tree.delete(*app.s4_tree.get_children())
    except Exception:
        pass

    out = []
    tot_exp = Decimal("0")
    tot_app = Decimal("0")
    div_cnt = 0

    if prov == "TICKET":
        # Ticket: conciliação por LOTE (Número do reembolso). Tipo Cartão fixo = Voucher.
        groups: dict[str, dict] = {}
        for rid, dt_iso, pay_iso, bruto, liq, auth, raw in rows:
            reemb = (str(auth).strip() if auth else "") or extract_reembolso(raw) or str(rid)
            g = groups.setdefault(reemb, {
                "sale_dt": dt_iso,
                "pay_dt": pay_iso,
                "tipo": "VOUCHER",
                "reemb": reemb,
                "gross": Decimal("0"),
                "net": Decimal("0"),
                "qty": 0
            })
            bruto_d = Decimal(str(bruto or "0"))
            liq_d = Decimal(str(liq or "0"))
            g["gross"] += bruto_d
            g["net"] += liq_d
            g["qty"] += 1
            # manter a menor data de venda, e a maior data de pagamento (se aparecerem)
            if dt_iso and (not g["sale_dt"] or str(dt_iso) < str(g["sale_dt"])):
                g["sale_dt"] = dt_iso
            if pay_iso and (not g["pay_dt"] or str(pay_iso) > str(g["pay_dt"])):
                g["pay_dt"] = pay_iso

        rule = fee_rule_tx_match(rules, "VOUCHER")
        if not rule:
            # sem cadastro: mostra tudo como SEM CAD.
            for reemb, g in sorted(groups.items(), key=lambda kv: (kv[1]["pay_dt"] or "", kv[0])):
                gross = g["gross"].quantize(Decimal("0.01"))
                net = g["net"].quantize(Decimal("0.01"))
                fee_app = (gross - net).quantize(Decimal("0.01"))
                fee_exp = Decimal("0.00")
                transf = Decimal("0.00")
                esp_banco = (gross - fee_exp - transf).quantize(Decimal("0.01"))
                diff = (fee_app - fee_exp).quantize(Decimal("0.01"))
                out.append({
                    "sale_dt": iso_to_br(g["sale_dt"]),
                    "pay_dt": iso_to_br(g["pay_dt"]) if g["pay_dt"] else "",
                    "tipo": g["tipo"],
                    "nsu": g["reemb"],
                    "bruto": float(gross),
                    "liquido": float(net),
                    "taxa_aplic": float(fee_app),
                    "taxa_esp": float(fee_exp),
                    "transf": float(transf),
                    "esp_banco": float(esp_banco),
                    "dif": float(diff),
                    "status": "SEM CAD."
                })
                app.s4_tree.insert("", "end", values=(
                    iso_to_br(g["sale_dt"]), iso_to_br(g["pay_dt"]) if g["pay_dt"] else "", g["tipo"], g["reemb"],
                    br_money(gross), br_money(net),
                    br_money(fee_app), br_money(fee_exp), br_money(transf), br_money(esp_banco),
                    br_money(diff), "SEM CAD."
                ), tags=("div",))
            app._s4_last_df = pd.DataFrame(out)
            app.s4_tot_var.set(f"Taxa esperada: {br_money(Decimal('0'))} | Taxa aplicada: {br_money(sum(Decimal(str(r['taxa_aplic'])) for r in out))} | Diferença: {br_money(sum(Decimal(str(r['taxa_aplic'])) for r in out))}")
            app.s4_div_var.set(f"Divergências: {len(out)}")
            return

        mdr = Decimal(str(rule.get("mdr_percent", 0))).scaleb(-2)  # percent -> fração
        fee_fixed = Decimal(str(rule.get("fee_fixed", 0)))
        transf = Decimal(str(rule.get("transfer_fee", 0)))

        for reemb, g in sorted(groups.items(), key=lambda kv: (kv[1]["pay_dt"] or "", kv[0])):
            gross = g["gross"].quantize(Decimal("0.01"))
            net = g["net"].quantize(Decimal("0.01"))
            fee_app = (gross - net).quantize(Decimal("0.01"))

            fee_exp = (gross * mdr + (fee_fixed * Decimal(g["qty"]))).quantize(Decimal("0.01"))
            diff = (fee_app - fee_exp).quantize(Decimal("0.01"))

            esp_banco = (gross - fee_exp - transf).quantize(Decimal("0.01"))

            ok = abs(diff) <= tol
            status = "✅" if ok else "❌"
            tag = "" if ok else "div"
            if not ok:
                div_cnt += 1

            tot_exp += fee_exp
            tot_app += fee_app

            rowd = {
                "sale_dt": iso_to_br(g["sale_dt"]),
                "pay_dt": iso_to_br(g["pay_dt"]) if g["pay_dt"] else "",
                "tipo": g["tipo"],
                "nsu": g["reemb"],
                "bruto": float(gross),
                "liquido": float(net),
                "taxa_aplic": float(fee_app),
                "taxa_esp": float(fee_exp),
                "transf": float(transf),
                "esp_banco": float(esp_banco),
                "dif": float(diff),
                "status": status
            }
            out.append(rowd)
            app.s4_tree.insert("", "end", values=(
                rowd["sale_dt"], rowd["pay_dt"], rowd["tipo"], rowd["nsu"],
                br_money(gross), br_money(net),
                br_money(fee_app), br_money(fee_exp),
                br_money(transf), br_money(esp_banco),
                br_money(diff), status
            ), tags=(tag,) if tag else ())

    else:
        # Demais providers (Alelo): conciliação por transação (venda a venda)
        for rid, dt_iso, pay_iso, bruto, liq, auth, raw in rows:
            bruto_d = Decimal(str(bruto or "0"))
            liq_d = Decimal(str(liq or "0"))
            tipo = extract_tipo_cartao(raw)
            rule = fee_rule_tx_match(rules, tipo)

            taxa_app = (bruto_d - liq_d).quantize(Decimal("0.01"))
            taxa_exp = Decimal("0.00")
            transf = Decimal("0.00")
            status = "SEM CAD."
            tag = "div"

            if rule:
                mdr = Decimal(str(rule.get("mdr_percent", 0))).scaleb(-2)
                fee_fixed = Decimal(str(rule.get("fee_fixed", 0)))
                transf = Decimal(str(rule.get("transfer_fee", 0)))
                taxa_exp = (bruto_d * mdr + fee_fixed).quantize(Decimal("0.01"))
                diff = (taxa_app - taxa_exp).quantize(Decimal("0.01"))
                if abs(diff) <= tol:
                    status = "✅"
                    tag = ""
                else:
                    status = "❌"
                    tag = "div"
                    div_cnt += 1
            else:
                diff = (taxa_app - taxa_exp).quantize(Decimal("0.01"))
                div_cnt += 1

            tot_exp += taxa_exp
            tot_app += taxa_app

            esp_banco = (bruto_d - taxa_exp - transf).quantize(Decimal("0.01"))

            rowd = {
                "id": rid,
                "sale_dt": iso_to_br(dt_iso),
                "pay_dt": iso_to_br(pay_iso) if pay_iso else "",
                "tipo": tipo,
                "nsu": auth or "",
                "bruto": float(bruto_d),
                "liquido": float(liq_d),
                "taxa_aplic": float(taxa_app),
                "taxa_esp": float(taxa_exp),
                "transf": float(transf),
                "esp_banco": float(esp_banco),
                "dif": float((taxa_app - taxa_exp).quantize(Decimal("0.01"))),
                "status": status
            }
            out.append(rowd)

            app.s4_tree.insert("", "end", iid=str(rid), values=(
                rowd["sale_dt"], rowd["pay_dt"], rowd["tipo"], rowd["nsu"],
                br_money(Decimal(str(rowd["bruto"]))), br_money(Decimal(str(rowd["liquido"]))),
                br_money(Decimal(str(rowd["taxa_aplic"]))), br_money(Decimal(str(rowd["taxa_esp"]))),
                br_money(Decimal(str(rowd["transf"]))), br_money(Decimal(str(rowd["esp_banco"]))),
                br_money(Decimal(str(rowd["dif"]))), rowd["status"]
            ), tags=(tag,) if tag else ())

    # totais
    diff_tot = (tot_app - tot_exp).quantize(Decimal("0.01"))
    app.s4_tot_var.set(f"Taxa esperada: {br_money(tot_exp)} | Taxa aplicada: {br_money(tot_app)} | Diferença: {br_money(diff_tot)}")
    app.s4_div_var.set(f"Divergências: {div_cnt}")

    app._s4_last_df = pd.DataFrame(out)
    app._log(f"Etapa 4 (taxas/tarifas) atualizada na tela. Registros: {len(out)}")



# ==========================================================
# ETAPA 4 (Ticket) - Conferência dinâmica por arquivo importado
#   (recalcula taxas/tarifas por "Número do reembolso" e compara
#    com "Valor Líquido" do próprio relatório)
# ==========================================================
def run_step4_ticket_conferencia_dinamica(app):
    """Conferência dinâmica Ticket (Voucher) baseada nos XLSX importados (imports.kind='RECEB').

    - Busca caminhos das planilhas Ticket já importadas (DB -> tabela imports)
    - Reprocessa o layout "Extrato de Reembolso Detalhado" (skiprows=13)
    - Filtra por Data de crédito/débito (pay_dt) no período selecionado no topo
    - Agrupa por Número do reembolso (lote)
    - Recalcula:
        • subtotal_compra = soma das linhas COMPRA
        • qtd_compra = contagem de linhas COMPRA
        • taxa_tpe (Taxa TPE) / tarifa por transação / receita gestão (por lote)
        • líquido_calculado = subtotal - taxa_tpe - tarifa - receita_gestao
      (Ignora linha 'TPE REEMBOLSO' conforme regra do usuário)
    - Compara líquido_calculado x Valor Líquido do relatório e imprime no log.
    """
    if not hasattr(app, "conn") or app.conn is None:
        app.conn = connect(app.db_path.get()); init_db(app.conn)

    app_month = app._parse_month()
    if not app_month:
        return
    d1, d2 = app._parse_period()
    start, end = _get_period(app_month, d1, d2)

    prov = (app.provider.get() or "ALELO").strip().upper()
    if prov != "TICKET":
        app._log("Etapa 4 (Conferência Ticket): selecione a bandeira TICKET no topo.")
        return

    # parâmetros contratuais (regras da etapa 4)
    rules = fee_rules_tx_list(app.conn, prov)
    rule = fee_rule_tx_match(rules, "VOUCHER") or fee_rule_tx_match(rules, "TICKET") or (rules[0] if rules else None)
    mdr = Decimal("4.5")
    fee_fixed = Decimal("0.52")
    if rule is not None:
        try:
            # (id, prov, label, match, mdr_pct, fixed_fee, transfer_fee, active) ou variações
            if len(rule) >= 8:
                mdr = Decimal(str(rule[4] or "0"))
                fee_fixed = Decimal(str(rule[5] or "0"))
            elif len(rule) == 7:
                mdr = Decimal(str(rule[4] or "0"))
                fee_fixed = Decimal(str(rule[5] or "0"))
        except Exception:
            pass

    # localizar planilhas importadas
    cur = app.conn.cursor()
    cur.execute(
        "SELECT source_path, imported_at FROM imports WHERE provider=? AND kind='RECEB' ORDER BY imported_at DESC",
        (prov,),
    )
    rows = cur.fetchall() or []
    paths = []
    seen = set()
    for p, _ts in rows:
        p = str(p or "").strip()
        if not p or p in seen:
            continue
        seen.add(p)
        paths.append(p)

    if not paths:
        app._log("Etapa 4 (Conferência Ticket): não encontrei planilhas de RECEB importadas no DB (tabela imports).")
        return

    # tolerância
    tol_s = ui_get(app.conn, "fees_tolerance", "0.05")
    try:
        tol = Decimal(str(tol_s).replace(",", "."))
    except Exception:
        tol = Decimal("0.05")

    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip()).upper()

    # padrões (robustos para variações / encoding ruim)
    PAT_COMPRA = re.compile(r"\bCOMPRA\b", re.I)
    PAT_VALOR_LIQ = re.compile(r"VALOR\s*L[ÍI]QUIDO", re.I)
    PAT_TARIFA_TX = re.compile(r"TARIFA.*TRANSA", re.I)
    PAT_RECEITA_GESTAO = re.compile(r"RECEITA.*GEST", re.I)
    PAT_TAXA_TPE = re.compile(r"TAXA\s*TPE", re.I)
    PAT_TPE_REEMB = re.compile(r"TPE\s*REEMBOLSO", re.I)

    lotes = {}  # (pay_dt, reembolso) -> dict
    files_used = 0
    files_skipped = 0

    for xls_path in paths:
        try:
            if not Path(xls_path).exists():
                files_skipped += 1
                continue
            df = pd.read_excel(xls_path, skiprows=13, dtype=object)
        except Exception:
            files_skipped += 1
            continue

        if df is None or df.empty:
            files_skipped += 1
            continue

        cols = set([str(c).strip() for c in df.columns])
        if not (TICKET_REEMB_REQUIRED_COLS <= cols):
            # não é o layout esperado
            files_skipped += 1
            continue

        df = df[list(TICKET_REEMB_REQUIRED_COLS)].copy()
        df["Número do reembolso"] = df["Número do reembolso"].astype("string").fillna("").str.strip()
        df.loc[df["Número do reembolso"] == "", "Número do reembolso"] = pd.NA
        df["Número do reembolso"] = df["Número do reembolso"].ffill()
        df = df.dropna(subset=["Número do reembolso"]).copy()

        for c in ["Data de crédito/débito", "Data da transação"]:
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

        df = df.dropna(subset=["Data de crédito/débito"]).copy()
        df["pay_dt"] = df["Data de crédito/débito"].dt.date
        df = df[(df["pay_dt"] >= start) & (df["pay_dt"] <= end)].copy()
        if df.empty:
            continue

        df["reembolso_id"] = df["Número do reembolso"].astype("string").str.strip()
        df["descr"] = df["Descrição do lançamento"].astype("string").fillna("").map(_norm)
        df["val"] = df["Valor da transação"].apply(_ticket_safe_float_brl).astype(float)

        for (pay_dt, rid), g in df.groupby(["pay_dt", "reembolso_id"], dropna=False):
            rid = str(rid or "").strip()
            if not rid:
                continue
            key = (pay_dt, rid)
            lot = lotes.get(key)
            if lot is None:
                lot = {
                    "pay_dt": pay_dt,
                    "rid": rid,
                    "compras": [],  # list of dicts: sale_dt, bruto
                    "qtd": 0,
                    "subtotal": Decimal("0"),
                    "tarifa": Decimal("0"),
                    "receita_gestao": Decimal("0"),
                    "taxa_tpe": Decimal("0"),
                    "valor_liquido": None,
                }
                lotes[key] = lot

            # compras
            compras = g[g["descr"].str.contains(PAT_COMPRA, na=False)].copy()
            if not compras.empty:
                for _, rr in compras.iterrows():
                    sale_dt = rr.get("Data da transação")
                    sale_dt = sale_dt.date() if isinstance(sale_dt, pd.Timestamp) else None
                    v = Decimal(str(rr.get("val") or 0)).quantize(Decimal("0.01"))
                    lot["compras"].append({"sale_dt": sale_dt, "bruto": v})
                    lot["qtd"] += 1
                    lot["subtotal"] += v

            # tarifa por transação (linha do lote costuma vir acumulada)
            t = g[g["descr"].str.contains(PAT_TARIFA_TX, na=False) & (~g["descr"].str.contains(PAT_TPE_REEMB, na=False))].copy()
            if not t.empty:
                lot["tarifa"] += Decimal(str(float(t["val"].sum()))).quantize(Decimal("0.01"))

            # receita gestão de pagamento (taxa por transferência)
            rg = g[g["descr"].str.contains(PAT_RECEITA_GESTAO, na=False)].copy()
            if not rg.empty:
                lot["receita_gestao"] += Decimal(str(float(rg["val"].sum()))).quantize(Decimal("0.01"))

            # taxa tpe (adm)
            tt = g[g["descr"].str.contains(PAT_TAXA_TPE, na=False)].copy()
            if not tt.empty:
                lot["taxa_tpe"] += Decimal(str(float(tt["val"].sum()))).quantize(Decimal("0.01"))

            # valor líquido (aplicado)
            vl = g[g["descr"].str.contains(PAT_VALOR_LIQ, na=False)].copy()
            if not vl.empty:
                vls = Decimal(str(float(vl["val"].sum()))).quantize(Decimal("0.01"))
                lot["valor_liquido"] = vls

        files_used += 1

    if not lotes:
        app._log(f"Etapa 4 (Conferência Ticket): nenhuma linha no período {iso_to_br(start)} a {iso_to_br(end)} nas planilhas importadas.")
        return

    # montar relatórios
    rows_div = []
    rows_an = []

    tot_calc = Decimal("0")
    tot_file = Decimal("0")

    for (pay_dt, rid), lot in sorted(lotes.items(), key=lambda x: (x[0][0], x[0][1])):
        subtotal = lot["subtotal"]
        qtd = int(lot["qtd"])
        if subtotal <= 0 or qtd <= 0:
            continue

        tarifa = lot["tarifa"]
        # fallback: se o arquivo não trouxer tarifa, calcula por contrato (qtd * fee_fixed)
        if tarifa == 0 and fee_fixed > 0:
            tarifa = (fee_fixed * Decimal(qtd)).quantize(Decimal("0.01"))

        taxa_tpe = lot["taxa_tpe"]
        # fallback: se o arquivo não trouxer Taxa TPE, calcula por contrato (mdr% do subtotal)
        if taxa_tpe == 0 and mdr > 0:
            taxa_tpe = (subtotal * (mdr/Decimal("100"))).quantize(Decimal("0.01"))

        receita_gestao = lot["receita_gestao"]
        liq_calc = (subtotal - taxa_tpe - tarifa - receita_gestao).quantize(Decimal("0.01"))

        liq_file = lot["valor_liquido"]
        if liq_file is None:
            liq_file = Decimal("0.00")

        tot_calc += liq_calc
        tot_file += liq_file

        diff = (liq_file - liq_calc).quantize(Decimal("0.01"))
        if abs(diff) > tol:
            rows_div.append({
                "pay_dt": pay_dt,
                "reembolso": rid,
                "qtd": qtd,
                "subtotal": subtotal,
                "liq_file": liq_file,
                "liq_calc": liq_calc,
                "dif": diff,
                "tarifa": tarifa,
                "taxa_tpe": taxa_tpe,
                "receita_gestao": receita_gestao,
            })

        # analítico: % efetivo por compra, com rate (taxa_tpe+receita_gestao) proporcional + tarifa fixa por compra
        rate_prop = Decimal("0")
        if subtotal > 0:
            rate_prop = ((taxa_tpe + receita_gestao) / subtotal)
        tarifa_unit = (tarifa / Decimal(qtd)) if qtd else Decimal("0")

        for it in lot["compras"]:
            bruto = it["bruto"]
            if bruto <= 0:
                continue
            prop = (bruto * rate_prop).quantize(Decimal("0.01"))
            liq = (bruto - prop - tarifa_unit).quantize(Decimal("0.01"))
            perc = Decimal("0")
            if bruto != 0:
                perc = ((bruto - liq) / bruto * Decimal("100")).quantize(Decimal("0.01"))
            rows_an.append({
                "data_compra": it["sale_dt"],
                "reembolso": rid,
                "pay_dt": pay_dt,
                "vl_bruto": bruto,
                "vl_liq_calc": liq,
                "perc_efetivo": perc,
            })

    diff_tot = (tot_file - tot_calc).quantize(Decimal("0.01"))

    # imprimir no log (simulação solicitada)
    app._log("—"*72)
    app._log(f"ETAPA 4 | TICKET | CONFERÊNCIA DINÂMICA ({iso_to_br(start)} a {iso_to_br(end)})")
    app._log(f"Arquivos considerados: {files_used} | ignorados: {files_skipped}")
    app._log(f"Contrato (fallback): Taxa TPE/ADM = {mdr}% | Tarifa por transação = R$ {str(fee_fixed).replace('.',',')}")
    app._log(f"TOTAL (arquivo)  : {br_money(tot_file)}")
    app._log(f"TOTAL (calculado): {br_money(tot_calc)}")
    app._log(f"DIFERENÇA TOTAL  : {br_money(diff_tot)} | tolerância: ±{br_money(tol)}")
    app._log("—"*72)

    if rows_div:
        app._log(f"Divergências encontradas: {len(rows_div)} (mostrando até 30)")
        for r in rows_div[:30]:
            app._log(
                f"Reembolso {r['reembolso']} | Pgto {iso_to_br(r['pay_dt'])} | "
                f"Arquivo {br_money(r['liq_file'])} vs Calc {br_money(r['liq_calc'])} | Dif {br_money(r['dif'])} | "
                f"Sub {br_money(r['subtotal'])} | Qtd {r['qtd']} | "
                f"TaxaTPE {br_money(r['taxa_tpe'])} | Tarifa {br_money(r['tarifa'])} | Gestão {br_money(r['receita_gestao'])}"
            )
    else:
        app._log("Nenhuma divergência acima da tolerância no período. ✅")

    # analítico (média do período)
    if rows_an:
        df_an = pd.DataFrame(rows_an)
        try:
            pm = float(df_an["perc_efetivo"].astype(float).mean()) if not df_an.empty else 0.0
        except Exception:
            pm = 0.0
        app._log("—"*72)
        app._log("Analítico (primeiras 25 linhas): Data Compra | Bruto | Líq Calc | % Efetivo | Reembolso")
        for r in rows_an[:25]:
            app._log(
                f"{iso_to_br(r['data_compra']) if r['data_compra'] else '--/--/----'} | "
                f"{br_money(r['vl_bruto'])} | {br_money(r['vl_liq_calc'])} | {str(r['perc_efetivo']).replace('.',',')}% | {r['reembolso']}"
            )
        app._log(f"SOMA BRUTO (amostra) = {br_money(Decimal(str(df_an['vl_bruto'].astype(float).sum())).quantize(Decimal('0.01')))} | % médio (amostra) = {str(Decimal(str(pm)).quantize(Decimal('0.01'))).replace('.',',')}%")

    # guarda dataframes para export futuro, sem alterar outras etapas
    try:
        app._s4_ticket_dyn_div = pd.DataFrame(rows_div)
        app._s4_ticket_dyn_ana = pd.DataFrame(rows_an)
    except Exception:
        pass


class App(tk.Tk):
    def __init__(self, fixed_provider: str | None = None, fixed_db_path: str | None = None):
        super().__init__()
        self.title(APP_TITLE); self.geometry("1280x780")
        self.db_path = tk.StringVar(value=DEFAULT_DB)
        self.provider = tk.StringVar(value="ALELO")
        self._fixed_provider = (fixed_provider or "").strip().upper() or None
        if self._fixed_provider:
            self.provider.set(self._fixed_provider)
        self._fixed_db_path = (fixed_db_path or "").strip() or None
        if self._fixed_db_path:
            self.db_path.set(self._fixed_db_path)
        self.month_str = tk.StringVar(value=datetime.now().strftime("%m/%Y"))
        self.bank_keyword = tk.StringVar(value="")

        # restaura filtro banco persistido (se existir)
        try:
            _bk = ui_get(self.conn, "bank_keyword", None)
            if _bk and _bk.strip():
                self.bank_keyword.set(_bk.strip())
        except Exception:
            pass

        # persiste automaticamente ao editar
        def _persist_bank_kw(*_):
            try:
                ui_set(self.conn, "bank_keyword", (self.bank_keyword.get() or "").strip())
            except Exception:
                pass
        self.bank_keyword.trace_add("write", _persist_bank_kw)
        self.period_from = tk.StringVar(value=""); self.period_to = tk.StringVar(value="")
        self.year_view = tk.IntVar(value=datetime.now().year)
        self._build_ui()
        self.conn = connect(self.db_path.get()); init_db(self.conn)
        self._refresh_all()
    def _build_ui(self):
        top = ttk.Frame(self, padding=10); top.pack(fill="x")
        # (UI) Ocultar campo "Banco de dados" para liberar espaço no cabeçalho
        ttk.Button(top, text="Abrir/CRIAR DB", command=self._open_db).pack(side="left", padx=(0,10))
        ttk.Label(top, text="Bandeira:").pack(side="left")
        # Provider selector (hidden/locked when app is launched as a single-provider window)
        if self._fixed_provider:
            ttk.Label(top, text=self._fixed_provider).pack(side="left", padx=(5,10))
        else:
            self.cb_provider = ttk.Combobox(top, textvariable=self.provider, values=("ALELO","TICKET","FARMACIASAPP"), width=14, state="readonly")
            self.cb_provider.pack(side="left", padx=(5,10))

            # mudanças de bandeira (carrega palavras-chave, reconfigura grids)
            try:
                self.provider.trace_add("write", lambda *a: self._on_provider_change())
            except Exception:
                try:
                    self.provider.trace("w", lambda *a: self._on_provider_change())
                except Exception:
                    pass

        ttk.Label(top, text="Mês (mm/aaaa):").pack(side="left")
        ttk.Entry(top, textvariable=self.month_str, width=10).pack(side="left", padx=(5,10))
        ttk.Label(top, text="Período (dd/mm/aaaa):").pack(side="left")
        ttk.Entry(top, textvariable=self.period_from, width=11).pack(side="left", padx=(5,2))
        ttk.Label(top, text="a").pack(side="left")
        ttk.Entry(top, textvariable=self.period_to, width=11).pack(side="left", padx=(2,10))
        ttk.Label(top, text="Filtro Banco (token):").pack(side="left")
        ttk.Entry(top, textvariable=self.bank_keyword, width=16).pack(side="left", padx=(5,6))
        ttk.Button(top, text="Definir palavras-chave…", command=self._open_bank_keywords_editor).pack(side="left", padx=(0,8))
        self.s3_kw_lbl = ttk.Label(top, text=self._bank_kw_status())
        self.s3_kw_lbl.pack(side="left", padx=(0,10))
        # Botão Taxas/Tarifas (Etapa 4) - com ícone opcional (fallback para texto)
        try:
            if not hasattr(self, "_img_fee"):
                import base64 as _b64
                self._img_fee = tk.PhotoImage(data=_b64.b64decode(FEE_ICON_PNG_B64))
            ttk.Button(top, text="Cadastro de Taxas", command=self._open_fee_modal).pack(side="left", padx=(0,10))
        except Exception:
            ttk.Button(top, text="Cadastro de Taxas", command=self._open_fee_modal).pack(side="left", padx=(0,10))

        ttk.Button(top, text="Excluir em massa (período)", command=self._open_bulk_delete).pack(side="right")
        logwrap = ttk.PanedWindow(self, orient="vertical"); logwrap.pack(fill="x", padx=10, pady=(0,10))
        self.log_import = tk.Text(logwrap, height=5, wrap="word")
        self.log_error = tk.Text(logwrap, height=4, wrap="word")
        logwrap.add(self.log_import, weight=3); logwrap.add(self.log_error, weight=2)
        self._log("Pronto. Importar -> Rodar etapas -> Ajustar/Excluir se necessário.")
        self._err("Erros aparecerão aqui (se houver).")
        self.nb = ttk.Notebook(self); self.nb.pack(fill="both", expand=True, padx=10, pady=10)
        self.tab_import = ttk.Frame(self.nb); self.tab_s1 = ttk.Frame(self.nb); self.tab_s2 = ttk.Frame(self.nb)
        self.tab_s3 = ttk.Frame(self.nb); self.tab_close = ttk.Frame(self.nb); self.tab_bank = ttk.Frame(self.nb)
        self.tab_rep = ttk.Frame(self.nb)
        self.nb.add(self.tab_import, text="Importação")
        self.nb.add(self.tab_s1, text="Etapa 1 (ERP x Vendas)")
        self.nb.add(self.tab_s2, text="Etapa 2 (Vendas x Recebimentos)")
        self.nb.add(self.tab_s3, text="Etapa 3 (Fechamento Mensal)")
        self.nb.add(self.tab_close, text="Fechamentos")
        self.nb.add(self.tab_rep, text="Relatório de Divergências")
        self.nb.add(self.tab_bank, text="Banco - Pesquisa")
        self._build_import_tab(); self._build_step1_tab(); self._build_step2_tab(); self._build_step3_tab(); self._build_close_tab(); self._build_report_tab(); self._build_bank_tab()
        try:
            self._on_provider_change()
        except Exception:
            pass

    def _on_provider_change(self):
        """Atualiza UI e parâmetros dependentes da bandeira (sem quebrar outras)."""
        prov = (self.provider.get() if hasattr(self, "provider") else "").strip().upper() or "ALELO"
        # carregar palavras-chave persistidas para este provider
        if not hasattr(self, "conn") or self.conn is None:
            try:
                self.conn = connect(self.db_path.get()); init_db(self.conn)
            except Exception:
                self.conn = None
        kv_key = f"bank_user_memo_terms_{prov}"
        try:
            raw = ui_get(self.conn, kv_key, "") if self.conn is not None else ""
        except Exception:
            raw = ""
        terms = [x.strip() for x in str(raw).split("\n") if x.strip()]
        BANK_USER_MEMO_TERMS[:] = []
        for x in terms:
            if x not in BANK_USER_MEMO_TERMS:
                BANK_USER_MEMO_TERMS.append(x)

        # token padrão por provider
        token_std = BANK_TRANSF_TOKEN_FARM if prov=="FARMACIASAPP" else (BANK_TRANSF_TOKEN_TICKET if prov=="TICKET" else BANK_TRANSF_TOKEN_ALELO)
        try:
            all_terms = bank_all_memo_terms(prov)
            if hasattr(self, "bank_keyword"):
                self.bank_keyword.set(token_std if all_terms else "")
            if hasattr(self, "bank_term"):
                self.bank_term.set(token_std if all_terms else "")
        except Exception:
            pass

        # reconfigura grids que dependem do provider
        try:
            self._configure_step2_tree(prov)
        except Exception:
            pass
        try:
            if hasattr(self, "bank_kw_lbl"):
                self.bank_kw_lbl.configure(text=self._bank_kw_status())
            if hasattr(self, "s3_kw_lbl"):
                self.s3_kw_lbl.configure(text=self._bank_kw_status())
        except Exception:
            pass
        try:
            self._refresh_bank_search()
        except Exception:
            pass

    def _bank_kw_status(self) -> str:
        n = len(BANK_USER_MEMO_TERMS or [])
        return f"Ativo: {n} palavra(s)" if n else "Sem palavras personalizadas"

    def _open_bank_keywords_editor(self):
        """Editor simples de palavras-chave para busca no extrato bancário (memo).

        - Persistência por bandeira em ui_kv (SQLite)
        """
        # garante conexão
        if not hasattr(self, "conn") or self.conn is None:
            try:
                self.conn = connect(self.db_path.get()); init_db(self.conn)
            except Exception:
                self.conn = None

        prov = (self.provider.get() if hasattr(self, "provider") else "").strip().upper() or "ALELO"
        kv_key = f"bank_user_memo_terms_{prov}"
        # carrega termos persistidos
        try:
            raw = ui_get(self.conn, kv_key, "") if self.conn is not None else ""
        except Exception:
            raw = ""
        terms = [x.strip() for x in str(raw).split("\n") if x.strip()]
        BANK_USER_MEMO_TERMS[:] = []
        for x in terms:
            if x not in BANK_USER_MEMO_TERMS:
                BANK_USER_MEMO_TERMS.append(x)

        win = tk.Toplevel(self)
        win.title("Palavras-chave do banco (memo)")
        win.transient(self)
        win.grab_set()

        lb = tk.Listbox(win, width=72, height=10)
        lb.pack(padx=10, pady=10, fill="both", expand=True)

        def refresh():
            lb.delete(0, "end")
            for k in (BANK_USER_MEMO_TERMS or []):
                lb.insert("end", k)

        def add_kw():
            v = simpledialog.askstring("Adicionar", "Digite a palavra-chave (parte do memo):", parent=win)
            if v:
                vv = v.strip()
                if vv and vv not in BANK_USER_MEMO_TERMS:
                    BANK_USER_MEMO_TERMS.append(vv)
                refresh()

        def edit_kw():
            sel = lb.curselection()
            if not sel:
                return
            i = sel[0]
            cur = BANK_USER_MEMO_TERMS[i]
            v = simpledialog.askstring("Editar", "Edite a palavra-chave:", initialvalue=cur, parent=win)
            if v is None:
                return
            vv = v.strip()
            if not vv:
                return
            BANK_USER_MEMO_TERMS[i] = vv
            # dedup
            seen = []
            for x in BANK_USER_MEMO_TERMS:
                if x not in seen:
                    seen.append(x)
            BANK_USER_MEMO_TERMS[:] = seen
            refresh()

        def del_kw():
            sel = lb.curselection()
            if not sel:
                return
            del BANK_USER_MEMO_TERMS[sel[0]]
            refresh()

        frm = ttk.Frame(win)
        frm.pack(padx=10, pady=(0,10), fill="x")
        ttk.Button(frm, text="Adicionar", command=add_kw).pack(side="left", padx=5)
        ttk.Button(frm, text="Editar", command=edit_kw).pack(side="left", padx=5)
        ttk.Button(frm, text="Excluir", command=del_kw).pack(side="left", padx=5)
        ttk.Button(frm, text="Fechar", command=win.destroy).pack(side="right", padx=5)

        refresh()
        win.wait_window()

        # persiste por bandeira
        try:
            if self.conn is not None:
                ui_set(self.conn, kv_key, "\n".join([str(x).strip() for x in (BANK_USER_MEMO_TERMS or []) if str(x).strip()]))
        except Exception:
            pass

        # Após alterações, aplicar token padrão quando houver termos e atualizar labels
        try:
            token_std = BANK_TRANSF_TOKEN_FARM if prov=="FARMACIASAPP" else (BANK_TRANSF_TOKEN_TICKET if prov=="TICKET" else BANK_TRANSF_TOKEN_ALELO)
            # se não houver termos (fixos + usuário), não força token
            all_terms = bank_all_memo_terms(prov)
            self.bank_keyword.set(token_std if all_terms else "")
        except Exception:
            pass
        try:
            self.bank_term.set(token_std if all_terms else "")
        except Exception:
            pass
        if hasattr(self, "bank_kw_lbl"):
            self.bank_kw_lbl.configure(text=self._bank_kw_status())
        if hasattr(self, "s3_kw_lbl"):
            self.s3_kw_lbl.configure(text=self._bank_kw_status())
        # opcional: atualizar pesquisa bancária se a aba existir
        try:
            self._refresh_bank_search()
        except Exception:
            pass


    def _log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        line = f"[{ts}] {msg}"
        self.log_import.insert("end", line + "\n"); self.log_import.see("end")
        imp_path, _ = _app_log_paths()
        _append_line(imp_path, line)
    def _err(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        line = f"[{ts}] {msg}"
        self.log_error.insert("end", line + "\n"); self.log_error.see("end")
        _, err_path = _app_log_paths()
        _append_line(err_path, line)
    def _parse_month(self):
        m = month_start(self.month_str.get())
        if not m:
            messagebox.showerror("Erro","Mês inválido. Use mm/aaaa."); return None
        return m
    def _parse_period(self):
        d1 = parse_br_date_str(self.period_from.get()); d2 = parse_br_date_str(self.period_to.get())
        if (self.period_from.get().strip() or self.period_to.get().strip()) and (not d1 or not d2):
            messagebox.showerror("Erro","Período inválido. Use dd/mm/aaaa em ambos os campos."); return (None, None)
        return d1, d2
    def _open_db(self):
        p = filedialog.asksaveasfilename(title="Escolher/CRIAR banco de dados", defaultextension=".sqlite",
                                         filetypes=[("SQLite","*.sqlite"),("Todos","*.*")])
        if not p: return
        self.db_path.set(p)
        try:
            if hasattr(self, "conn") and self.conn: self.conn.close()
        except Exception:
            pass
        self.conn = connect(self.db_path.get()); init_db(self.conn)
        self._log(f"DB aberto: {self.db_path.get()}"); self._refresh_all()
    def _refresh_all(self):
        self._refresh_imports(); self._run_step1(silent=True); self._run_step2(silent=True); self._run_step3(silent=True)
        self._refresh_closings(); self._refresh_bank_search();
        try:
            self._refresh_diag()
        except Exception:
            pass
    def _build_import_tab(self):
        frm = ttk.Frame(self.tab_import, padding=10); frm.pack(fill="x")
        ttk.Button(frm, text="Importar ERP (Excel) - em massa", command=self._import_erp).pack(side="left", padx=5)
        ttk.Button(frm, text="Importar Vendas Portal (Excel) - em massa", command=self._import_sales).pack(side="left", padx=5)
        ttk.Button(frm, text="Importar Recebimentos (Excel) - em massa", command=self._import_receb).pack(side="left", padx=5)
        ttk.Button(frm, text="Importar Banco (OFX) - em massa", command=self._import_bank_ofx).pack(side="left", padx=5)
        ttk.Button(frm, text="Importar Banco (CSV Bradesco) - em massa", command=self._import_bank_csv).pack(side="left", padx=5)
        ttk.Button(frm, text="Exportar Imports (Excel)", command=self._export_imports).pack(side="right", padx=5)
        ttk.Separator(self.tab_import).pack(fill="x", pady=8)
        self.import_tree = ttk.Treeview(self.tab_import, columns=("quando","kind","arquivo"), show="headings", height=12)
        for c, w in (("quando",180),("kind",100),("arquivo",860)):
            self.import_tree.heading(c, text=c); self.import_tree.column(c, width=w, anchor="w")
        self.import_tree.pack(fill="both", expand=True, padx=10, pady=10)
    def _refresh_imports(self):
        for i in self.import_tree.get_children(): self.import_tree.delete(i)
        for r in self.conn.execute("SELECT imported_at, kind, source_path FROM imports ORDER BY id DESC LIMIT 500").fetchall():
            self.import_tree.insert("", "end", values=(r["imported_at"], r["kind"], r["source_path"]))
    def _export_imports(self):
        p = filedialog.asksaveasfilename(title="Salvar Excel (Imports)", defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not p: return
        try:
            export_tree_to_excel(self.import_tree, p); self._log(f"Exportado: {p}")
        except Exception as e:
            self._err(f"Falha ao exportar imports: {e}")
    def _import_erp(self):
        paths = filedialog.askopenfilenames(title="Selecionar arquivo(s) ERP (Excel)", filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not paths: return
        prov = (self.provider.get().strip().upper() or "ALELO")
        ins = drop = 0
        for p in paths:
            res = import_erp_by_provider(self.conn, p, provider=prov)
            ins += res.inserted; drop += res.dropped_dupe
            for w in res.warnings: self._log("AVISO: " + w)
            for e in res.errors: self._err(e)
        self._log(f"ERP importado: {ins} | Duplicados bloqueados: {drop}"); self._refresh_all()
    def _import_sales(self):
        paths = filedialog.askopenfilenames(title="Selecionar arquivo(s) Vendas (Portal) - Excel", filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not paths: return
        prov = (self.provider.get().strip().upper() or "ALELO")
        ins = drop = 0
        for p in paths:
            res = import_sales_by_provider(self.conn, p, provider=prov)
            ins += res.inserted; drop += res.dropped_dupe
            for w in res.warnings: self._log("AVISO: " + w)
            for e in res.errors: self._err(e)
        self._log(f"Vendas importadas: {ins} | Duplicados bloqueados: {drop}"); self._refresh_all()
    def _import_receb(self):
        paths = filedialog.askopenfilenames(title="Selecionar arquivo(s) Recebimentos (Portal) - Excel", filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not paths: return
        prov = (self.provider.get().strip().upper() or "ALELO")
        ins = drop = 0
        for p in paths:
            res = import_receb_by_provider(self.conn, p, provider=prov)
            ins += res.inserted; drop += res.dropped_dupe
            for w in res.warnings: self._log("AVISO: " + w)
            for e in res.errors: self._err(e)
        self._log(f"Recebimentos importados: {ins} | Duplicados bloqueados: {drop}"); self._refresh_all()
    def _import_bank_ofx(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivo(s) extrato bancário (OFX/PDF)",
            filetypes=[("Extrato bancário","*.ofx *.OFX *.pdf *.PDF"),("OFX","*.ofx *.OFX"),("PDF","*.pdf *.PDF"),("Todos","*.*")]
        )
        if not paths:
            return
        ins = drop = 0
        for p in paths:
            try:
                ext = os.path.splitext(p)[1].lower()
                if ext == ".pdf":
                    res = import_bank_pdf_caixa(self.conn, p, provider="BANCO")
                else:
                    res = import_bank_ofx(self.conn, p, provider="BANCO")
                if res is None:
                    self._err("Falha ao importar extrato (retorno None). Verifique o arquivo.")
                    continue
                ins += getattr(res, "inserted", 0)
                drop += getattr(res, "dropped_dupe", 0)
                for w in getattr(res, "warnings", []):
                    self._log("AVISO: " + str(w))
                for e in getattr(res, "errors", []):
                    self._err(str(e))
            except Exception as e:
                self._err(str(e))
        self._log(f"Banco (OFX/PDF) importado: {ins} | Duplicados bloqueados: {drop}")
        self._refresh_all()
    def _import_bank_csv(self):
        paths = filedialog.askopenfilenames(title="Selecionar arquivo(s) extrato bancário (CSV Bradesco)", filetypes=[("CSV","*.csv *.CSV"),("Todos","*.*")])
        if not paths: return
        ins = drop = 0
        for p in paths:
            res = import_bank_csv_bradesco(self.conn, p, provider="BANCO")
            ins += res.inserted; drop += res.dropped_dupe
            for w in res.warnings: self._log("AVISO: " + w)
            for e in res.errors: self._err(e)
        self._log(f"Banco (CSV) importado: {ins} | Duplicados bloqueados: {drop}"); self._refresh_all()
    def _build_step1_tab(self):
        top = ttk.Frame(self.tab_s1, padding=10); top.pack(fill="x")
        self.s1_lbl = ttk.Label(top, text="Totais do período (Etapa 1):", font=("Segoe UI",10,"bold")); self.s1_lbl.pack(side="left")
        ttk.Button(top, text="Rodar Etapa 1", command=self._run_step1).pack(side="right", padx=5)
        ttk.Button(top, text="Exportar Excel", command=lambda: self._export_tree(self.s1_tree, "etapa1.xlsx")).pack(side="right", padx=5)
        cols=("data","erp_id","erp_bruto","vendas_ref","vendas_bruto","delta_dias","status","diferenca")
        self.s1_tree = ttk.Treeview(self.tab_s1, columns=cols, show="headings", height=22)
        for c,w,anch,txt in [
            ("data",110,"w","DATA"),
            ("erp_id",80,"e","ERP ID"),
            ("erp_bruto",120,"e","ERP (BRUTO)"),
            ("vendas_ref",220,"w","NSU / TRANS."),
            ("vendas_bruto",130,"e","VENDAS (BRUTO)"),
            ("delta_dias",90,"e","Δ DIAS"),
            ("status",90,"w","STATUS"),
            ("diferenca",120,"e","DIFERENÇA"),
        ]:
            self.s1_tree.heading(c, text=txt)
            self.s1_tree.column(c, width=w, anchor=anch)
        self.s1_tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.s1_tree.tag_configure("div", background="#ffdddd")
        self.s1_tree.bind("<Double-1>", lambda e: self._open_day_selector(self.s1_tree, e, step=1))

    def _run_step1(self, silent=False):
        m = self._parse_month()
        if not m:
            return
        d1, d2 = self._parse_period()
        prov = (self.provider.get().strip().upper() or "ALELO")

        df = run_step1_capture(self.conn, prov, m, d1, d2, window_days=2)

        tot_erp = sum([parse_decimal(x) or Decimal("0") for x in (df["erp_bruto"].tolist() if not df.empty else [])], Decimal("0"))
        tot_vd = sum([parse_decimal(x) or Decimal("0") for x in (df["vendas_bruto"].tolist() if not df.empty else [])], Decimal("0"))
        self.s1_lbl.configure(text=f"Totais (Etapa 1 - Captura) | ERP bruto: {br_money(tot_erp)} | Vendas bruto: {br_money(tot_vd)} | Dif.: {br_money(tot_vd - tot_erp)}")

        for i in self.s1_tree.get_children():
            self.s1_tree.delete(i)

        if not df.empty:
            for _, r in df.iterrows():
                tag = "div" if str(r.get("status","")) != "✅" else ""
                self.s1_tree.insert(
                    "",
                    "end",
                    values=(
                        fmt_br_date(r["data"]),
                        ("" if r.get("erp_id","")=="" else str(r.get("erp_id"))),
                        br_money(r.get("erp_bruto")),
                        str(r.get("vendas_ref") or ""),
                        br_money(r.get("vendas_bruto")),
                        ("" if r.get("delta_dias","")=="" else str(r.get("delta_dias"))),
                        str(r.get("status") or ""),
                        br_money(r.get("diferenca")),
                    ),
                    tags=(tag,),
                )

        if not silent:
            self._log("Etapa 1 (captura) atualizada na tela.")
    
    
    def _build_step2_tab(self):
        top = ttk.Frame(self.tab_s2, padding=10); top.pack(fill="x")
        self.s2_lbl = ttk.Label(top, text="Totais do período (Etapa 2):", font=("Segoe UI",10,"bold"))
        self.s2_lbl.pack(side="left")
        self.s2_div_var = tk.StringVar(value="Divergências: 0")
        ttk.Label(top, textvariable=self.s2_div_var, foreground="#a00000").pack(side="left", padx=(12,0))
        ttk.Button(top, text="Rodar Etapa 2", command=self._run_step2).pack(side="right", padx=5)
        ttk.Button(top, text="Exportar Excel", command=lambda: self._export_tree(self.s2_tree, "etapa2.xlsx")).pack(side="right", padx=5)
    
        self.s2_tree_frame = ttk.Frame(self.tab_s2)
        self.s2_tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
        self.s2_tree = None
        self._configure_step2_tree(self.provider.get().strip().upper() or "ALELO")
    
    

    def _configure_step2_tree(self, prov: str):
            """Configura grid da Etapa 2. Padrão por evento; FarmaciasAPP por pedido."""
            # recria frame/tree
            for w in self.s2_tree_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            prov = (prov or "").strip().upper()

            if prov == "FARMACIASAPP":
                cols = (
                    "pedido","data_pedido","forma_pagamento",
                    "valor_pago_loja","valor_subsidio","valor_total_compra",
                    "valor_repasse_aplicado","valor_taxa_aplicada","perc_cobrado_aplicado",
                    "tx_adm_pct","tx_modal_pct",
                    "valor_repasse_esperado","valor_taxa_esperada","perc_cobrado_esperado",
                    "diferenca"
                )
                self.s2_tree = ttk.Treeview(self.s2_tree_frame, columns=cols, show="headings", height=22)

                display = {
                    "pedido": "Nº PEDIDO",
                    "data_pedido": "DATA",
                    "forma_pagamento": "FORMA\nPGTO",
                    "valor_pago_loja": "VALOR PAGO\nNA LOJA",
                    "valor_subsidio": "VALOR\nSUBSÍDIO",
                    "valor_total_compra": "VALOR TOTAL\nDA COMPRA",
                    "valor_repasse_aplicado": "REPASSE\n(APLIC.)",
                    "valor_taxa_aplicada": "TAXA\n(APLIC.)",
                    "perc_cobrado_aplicado": "%COBRADO\n(APLIC.)",
                    "tx_adm_pct": "TX ADM\n(%)",
                    "tx_modal_pct": "TX MODAL\n(%)",
                    "valor_repasse_esperado": "REPASSE\n(ESP.)",
                    "valor_taxa_esperada": "TAXA\n(ESP.)",
                    "perc_cobrado_esperado": "%COBRADO\n(ESP.)",
                    "diferenca": "DIFER.\n(R$)",
                }
                col_cfg = [
                    ("pedido",120,"w"),
                    ("data_pedido",90,"w"),
                    ("forma_pagamento",85,"w"),
                    ("valor_pago_loja",120,"e"),
                    ("valor_subsidio",110,"e"),
                    ("valor_total_compra",130,"e"),
                    ("valor_repasse_aplicado",120,"e"),
                    ("valor_taxa_aplicada",110,"e"),
                    ("perc_cobrado_aplicado",110,"e"),
                    ("tx_adm_pct",80,"e"),
                    ("tx_modal_pct",90,"e"),
                    ("valor_repasse_esperado",120,"e"),
                    ("valor_taxa_esperada",110,"e"),
                    ("perc_cobrado_esperado",110,"e"),
                    ("diferenca",110,"e"),
                ]
            else:
                # Padrão (Alelo/Ticket): layout unificado (aplicado x esperado)
                cols = ("vendas_ref","data","vendas_bruto","receb_bruto","receb_id","repasse_aplicado","taxa_aplicada","repasse_esperado","diferenca")
                self.s2_tree = ttk.Treeview(self.s2_tree_frame, columns=cols, show="headings", height=22)

                display = {
                    "vendas_ref": "NSU / TRANS.",
                    "data": "DATA",
                    "vendas_bruto": "BRUTO (REL.VENDAS)",
                    "receb_bruto": "BRUTO (REL.RECEB.)",
                    "receb_id": "RECEB ID",
                    "repasse_aplicado": "REPASSE (APLIC.)",
                    "taxa_aplicada": "TAXA (APLIC.)",
                    "repasse_esperado": "REPASSE (ESP.)",
                    "diferenca": "DIFER. (R$)",
                }
                col_cfg = [
                    ("vendas_ref",150,"w"),
                    ("data",90,"w"),
                    ("vendas_bruto",140,"e"),
                    ("receb_bruto",140,"e"),
                    ("receb_id",0,"e"),
                    ("repasse_aplicado",140,"e"),
                    ("taxa_aplicada",120,"e"),
                    ("repasse_esperado",140,"e"),
                    ("diferenca",120,"e"),
                ]

            for c in cols:
                self.s2_tree.heading(c, text=display.get(c, c))
                w, a = 120, "w"
                for cc, ww, aa in col_cfg:
                    if cc == c:
                        w, a = ww, aa
                        break
                self.s2_tree.column(c, width=w, anchor=a)

            # ocultar RECEB ID no layout padrão
            try:
                if "receb_id" in self.s2_tree["columns"]:
                    self.s2_tree.column("receb_id", width=0, minwidth=0, stretch=False)
            except Exception:
                pass

            self.s2_tree.pack(fill="both", expand=True, side="left")

            # tags (destaques)
            try:
                self.s2_tree.tag_configure("div", background="#ffdddd")
            except Exception:
                pass

            ysb = ttk.Scrollbar(self.s2_tree_frame, orient="vertical", command=self.s2_tree.yview)
            self.s2_tree.configure(yscroll=ysb.set)
            ysb.pack(fill="y", side="left")

    def _run_step2(self, silent=False):
        m = self._parse_month()
        if not m:
            return
        d1, d2 = self._parse_period()
        prov = (self.provider.get().strip().upper() or "ALELO")

        # garante grid configurada
        cur_cols = tuple(self.s2_tree["columns"]) if self.s2_tree is not None else ()
        need_cols_farm = (
            "pedido","data_pedido","forma_pagamento",
            "valor_pago_loja","valor_subsidio","valor_total_compra",
            "valor_repasse_aplicado","valor_taxa_aplicada","perc_cobrado_aplicado",
            "tx_adm_pct","tx_modal_pct",
            "valor_repasse_esperado","valor_taxa_esperada","perc_cobrado_esperado",
            "diferenca"
        )
        need_cols_std = ("vendas_ref","data","vendas_bruto","receb_bruto","receb_id","repasse_aplicado","taxa_aplicada","repasse_esperado","diferenca")

        if prov == "FARMACIASAPP":
            if cur_cols != need_cols_farm:
                self._configure_step2_tree(prov)
        else:
            if cur_cols != need_cols_std:
                self._configure_step2_tree(prov)

        provider = get_provider(prov, run_step2_recebiveis)
        df = provider.run_step2(self.conn, prov, m, d1, d2, window_days=0)

        for i in self.s2_tree.get_children():
            self.s2_tree.delete(i)

        if prov == "FARMACIASAPP":
            # Totais
            tot_compra = sum([parse_decimal(x) or Decimal("0") for x in (df["valor_total_compra"].tolist() if not df.empty else [])], Decimal("0"))
            tot_rep = sum([parse_decimal(x) or Decimal("0") for x in (df["valor_repasse_aplicado"].tolist() if not df.empty else [])], Decimal("0"))
            tot_esp = sum([parse_decimal(x) or Decimal("0") for x in (df["valor_repasse_esperado"].tolist() if not df.empty else [])], Decimal("0"))
            tot_diff = tot_rep - tot_esp
            divs = int((df["diferenca"].abs() > Decimal(str(ui_get(self.conn, "fees_tolerance", "0.05")))).sum()) if (not df.empty and "diferenca" in df.columns) else 0

            self.s2_lbl.configure(text=f"Totais (Etapa 2 - FarmaciasApp) | Total compra: {br_money(tot_compra)} | Repasse: {br_money(tot_rep)} | Repasse esp.: {br_money(tot_esp)} | Dif.: {br_money(tot_diff)}")
            self.s2_div_var.set(f"Divergências: {divs}")

            # Inserção
            if not df.empty:
                for _, r in df.iterrows():
                    vals = []
                    for c in need_cols_farm:
                        v = r.get(c, "")
                        # percentuais
                        if c in ("perc_cobrado_aplicado","perc_cobrado_esperado","tx_adm_pct","tx_modal_pct"):
                            if pd.isna(v) or v == "":
                                v = ""
                            else:
                                vv = v
                                if isinstance(vv, Decimal):
                                    vv = float(vv)
                                try:
                                    v = f"{float(vv):.2f}%"
                                except Exception:
                                    v = ""
                        else:
                            if isinstance(v, Decimal):
                                v = br_money(v)
                            elif isinstance(v, (float, int)) and c not in ("pedido", "data_pedido", "forma_pagamento"):
                                v = br_money(Decimal(str(v)))
                            elif pd.isna(v):
                                v = ""
                        vals.append(v)
                    diff_v = r.get("diferenca")
                    tol = Decimal(str(ui_get(self.conn, "fees_tolerance", "0.05")))
                    try:
                        diff_d = diff_v if isinstance(diff_v, Decimal) else (parse_decimal(diff_v) or Decimal("0"))
                    except Exception:
                        diff_d = Decimal("0")
                    tags = ("div",) if abs(diff_d) > tol else ()
                    self.s2_tree.insert("", "end", values=tuple(vals), tags=tags)
            if not silent:
                self._log("Etapa 2 (FarmaciasApp) atualizada na tela.")
            return

        # padrão (evento) — layout unificado
        tot_rb = sum([parse_decimal(x) or Decimal("0") for x in (df.get("receb_bruto", pd.Series([], dtype=object)).tolist() if not df.empty else [])], Decimal("0"))
        tot_rep_ap = sum([parse_decimal(x) or Decimal("0") for x in (df.get("repasse_aplicado", pd.Series([], dtype=object)).tolist() if not df.empty else [])], Decimal("0"))
        tot_rep_exp = sum([parse_decimal(x) or Decimal("0") for x in (df.get("repasse_esperado", pd.Series([], dtype=object)).tolist() if not df.empty else [])], Decimal("0"))
        tot_diff = sum([parse_decimal(x) or Decimal("0") for x in (df.get("diferenca", pd.Series([], dtype=object)).tolist() if not df.empty else [])], Decimal("0"))
        self.s2_lbl.configure(text=f"Totais (Etapa 2) | Bruto receb: {br_money(tot_rb)} | Repasse (aplic.): {br_money(tot_rep_ap)} | Repasse (esp.): {br_money(tot_rep_exp)} | Diferença: {br_money(tot_diff)}")
        # divergências por tolerância
        try:
            tol = Decimal(str(ui_get(self.conn, "fees_tolerance", "0.05")))
        except Exception:
            tol = Decimal("0.05")
        divs = 0
        if not df.empty and "diferenca" in df.columns:
            try:
                divs = int((df["diferenca"].apply(lambda x: abs(parse_decimal(x) or Decimal("0")) > tol)).sum())
            except Exception:
                divs = 0
        self.s2_div_var.set(f"Divergências: {divs}")

        if not df.empty:
            for _, r in df.iterrows():
                diff_d = parse_decimal(r.get("diferenca")) or Decimal("0")
                tags = ("div",) if abs(diff_d) > tol else ()
                self.s2_tree.insert("", "end", values=(
                    r.get("vendas_ref",""),
                    fmt_date(r.get("data")),
                    br_money(parse_decimal(r.get("vendas_bruto")) or Decimal("0")),
                    br_money(parse_decimal(r.get("receb_bruto")) or Decimal("0")),
                    r.get("receb_id",""),
                    br_money(parse_decimal(r.get("repasse_aplicado")) or Decimal("0")),
                    br_money(parse_decimal(r.get("taxa_aplicada")) or Decimal("0")),
                    br_money(parse_decimal(r.get("repasse_esperado")) or Decimal("0")),
                    br_money(diff_d),
                ), tags=tags)

        if not silent:
            self._log("Etapa 2 (recebíveis) atualizada na tela.")

    def _build_step3_tab(self):
        # ===== Topo (modelo "cards" + ação principal) =====
        top = tk.Frame(self.tab_s3, bg="#f6f7fb")
        top.pack(fill="x", padx=10, pady=(10, 0))

        # filtros (mantém padrão do app, sem duplicar campos)
        filt = ttk.Frame(top)
        filt.pack(fill="x", pady=(6, 8))

        ttk.Label(filt, text="Filtro Banco (memo/token):").pack(side="left", padx=(0, 6))
        self.s3_bank_term = self.bank_keyword  # alias (mesma variável do topo)
        ttk.Entry(filt, textvariable=self.s3_bank_term, width=26).pack(side="left", padx=(0, 6))
        ttk.Button(filt, text="Palavras-chave...", command=self._open_bank_keywords_editor).pack(side="left", padx=(0, 12))

        ttk.Label(filt, text="Janela (dias):").pack(side="left", padx=(0, 4))
        self.s3_window_days = tk.IntVar(value=3)
        ttk.Spinbox(filt, from_=0, to=10, width=4, textvariable=self.s3_window_days).pack(side="left")

        ttk.Label(filt, text="Spillover (dias):").pack(side="left", padx=(10, 4))
        self.s3_spill_days = tk.IntVar(value=3)
        ttk.Spinbox(filt, from_=0, to=60, width=4, textvariable=self.s3_spill_days).pack(side="left")

        # cards + botão
        cards = tk.Frame(top, bg="#f6f7fb")
        cards.pack(fill="x", pady=(0, 10))

        self.s3_audit_line = tk.Label(top, text="", bg="#f6f7fb", fg="#555", font=("Segoe UI", 9))
        self.s3_audit_line.pack(fill="x", pady=(0, 10))

        def _card(parent, title, bg):
            f = tk.Frame(parent, bg=bg, bd=0, highlightthickness=0)
            t = tk.Label(f, text=title, bg=bg, fg="#333", font=("Segoe UI", 9, "bold"))
            v = tk.Label(f, text="R$ -", bg=bg, fg="#111", font=("Segoe UI", 11, "bold"))
            t.pack(anchor="w", padx=12, pady=(8, 0))
            v.pack(anchor="w", padx=12, pady=(2, 10))
            return f, v

        self.s3_card_exp, self.s3_lbl_exp = _card(cards, "TOTAL ESPERADO", "#e9ecef")
        self.s3_card_bnk, self.s3_lbl_bnk = _card(cards, "TOTAL BANCO", "#d9edf7")
        self.s3_card_dif, self.s3_lbl_dif = _card(cards, "DIFERENÇA", "#f8d7da")
        self.s3_card_sal, self.s3_lbl_sal = _card(cards, "SALDO (MÊS)", "#d4edda")

        for i, fr in enumerate([self.s3_card_exp, self.s3_card_bnk, self.s3_card_dif, self.s3_card_sal]):
            fr.pack(side="left", padx=(0 if i == 0 else 12, 0), ipadx=4, ipady=2)

        # ação principal + confirmar saldo
        actions = tk.Frame(cards, bg="#f6f7fb")
        actions.pack(side="right", padx=6)

        self.btn_s3_run = ttk.Button(actions, text="Rodar Conciliação", command=self._run_step3)
        self.btn_s3_run.pack(side="top", pady=(2, 6))

        self.btn_s3_confirm = ttk.Button(actions, text="Confirmar saldo (próx. mês)", command=self._confirm_step3_carryover)
        self.btn_s3_confirm.pack(side="top")

        ttk.Button(actions, text="Exportar Excel", command=self._export_step3_excel).pack(side="top", pady=(6, 0))
        ttk.Button(actions, text="Finalizar Fechamento", command=self._finalize_closing).pack(side="top", pady=(6, 0))

        self.btn_s3_confirm.state(["disabled"])
        self.s3_last_saldo_mes = Decimal("0")


        # ===== Tabela (visão diária) =====
        cols = ("data", "esperado", "banco", "diferenca", "saldo_acum", "alocado", "saldo_banco", "status")
        self.s3_tree = ttk.Treeview(self.tab_s3, columns=cols, show="headings", height=22)

        display = {
            "data": "DATA",
            "esperado": "ESPERADO (PAY_DT)",
            "banco": "BANCO (CRÉDITO)",
            "diferenca": "DIF. DIA",
            "saldo_acum": "SALDO ACUM.",
            "alocado": "ALOCADO",
            "saldo_banco": "SALDO BANCO",
            "status": "STATUS",
        }
        cfg = [
            ("data", 110, "w"),
            ("esperado", 140, "e"),
            ("banco", 140, "e"),
            ("diferenca", 110, "e"),
            ("saldo_acum", 120, "e"),
            ("alocado", 110, "e"),
            ("saldo_banco", 120, "e"),
            ("status", 260, "w"),
        ]
        for c, w, a in cfg:
            self.s3_tree.heading(c, text=display[c])
            self.s3_tree.column(c, width=w, anchor=a)

        self.s3_tree.pack(fill="both", expand=True, padx=10, pady=10)

        # linhas alternadas (sem cor por divergência; divergência só na coluna status)
        self.s3_tree.tag_configure("odd", background="#ffffff")
        self.s3_tree.tag_configure("even", background="#f2f2f2")

        # duplo clique abre detalhes do dia (banco + recebíveis)
        self.s3_tree.bind("<Double-1>", self._open_step3_day_detail)

        # caches da etapa 3
        self.s3_comp = {}       # bank_id -> list[receb_id]
        self.s3_bank_df = None  # df detalhado por lançamento bancário


    def _run_step3(self, silent=False):
        m = self._parse_month()
        if not m:
            return

        prov = (self.provider.get().strip().upper() or "ALELO")

        # Filtro banco (fonte única)
        term = (self.bank_keyword.get() if hasattr(self, "bank_keyword") else "").strip()
        try:
            if hasattr(self, "s3_bank_term"):
                self.s3_bank_term.set(term)
        except Exception:
            pass

        wd = int(self.s3_window_days.get() if hasattr(self, "s3_window_days") else 2)
        sp = int(self.s3_spill_days.get() if hasattr(self, "s3_spill_days") else 15)

        df_day, df_bank, comp = run_step3_daily_view(
            self.conn, prov, m, window_days=wd, bank_term=term, spillover_days=sp
        )
        self.s3_comp = comp or {}
        self.s3_bank_df = df_bank if df_bank is not None else None

        # limpa grid
        for i in self.s3_tree.get_children():
            self.s3_tree.delete(i)

        # totais
        tot_exp = Decimal("0")
        tot_bnk = Decimal("0")
        saldo_mes = Decimal("0")
        used_sp = False

        if df_day is not None and not df_day.empty:
            tot_exp = sum([parse_decimal(x) or Decimal("0") for x in df_day["esperado"].tolist()], Decimal("0"))
            tot_bnk = sum([parse_decimal(x) or Decimal("0") for x in df_day["banco"].tolist()], Decimal("0"))
            used_sp = bool(getattr(df_day, "attrs", {}).get("used_spillover", False))
            saldo_mes = Decimal(str(getattr(df_day, "attrs", {}).get("saldo_mes", 0.0))).quantize(Decimal("0.01"))
            # linha auditável: Saldo inicial + Diferença (mês) = Saldo final
            try:
                end_m = getattr(df_day, "attrs", {}).get("month_end")
                saldo_ini = Decimal(str(getattr(df_day, "attrs", {}).get("saldo_inicial", 0.0))).quantize(Decimal("0.01"))
                dif_mes = Decimal("0")
                for _, rr2 in df_day.iterrows():
                    dd = rr2.get("data")
                    if isinstance(dd, date) and end_m and dd <= end_m:
                        dif_mes += (parse_decimal(rr2.get("diferenca")) or Decimal("0"))
                dif_mes = dif_mes.quantize(Decimal("0.01"))
                if hasattr(self, "s3_audit_var"):
                    self.s3_audit_var.set(
                        f"Saldo inicial: {fmt_money(saldo_ini)}  +  Diferença do mês: {fmt_money(dif_mes)}  =  Saldo final: {fmt_money(saldo_mes)}"
                    )
            except Exception:
                if hasattr(self, "s3_audit_var"):
                    self.s3_audit_var.set("")


            for idx, (_, r) in enumerate(df_day.iterrows()):
                st = str(r.get("status") or "")
                tag = "even" if (idx % 2 == 0) else "odd"
                self.s3_tree.insert(
                    "",
                    "end",
                    values=(
                        fmt_br_date(r["data"]),
                        br_money(r.get("esperado")),
                        br_money(r.get("banco")),
                        br_money(r.get("diferenca")),
                        br_money(r.get("saldo_acum")),
                        br_money(r.get("alocado")),
                        br_money(r.get("saldo_banco")),
                        st,
                    ),
                    tags=(tag,),
                )

        # Atualiza cards
        diff = (tot_bnk - tot_exp).quantize(Decimal("0.01"))
        self.s3_lbl_exp.configure(text=br_money(tot_exp))
        self.s3_lbl_bnk.configure(text=br_money(tot_bnk))

        # diferença: se positiva, verde; se negativa, vermelho; se zero, neutro
        if diff > 0:
            self.s3_lbl_dif.configure(text=f"+{br_money(diff)}")
        elif diff < 0:
            self.s3_lbl_dif.configure(text=f"-{br_money(abs(diff))}")
        else:
            self.s3_lbl_dif.configure(text=br_money(diff))

        self.s3_lbl_sal.configure(text=br_money(saldo_mes))

        # Linha auditável: saldo inicial + diferença do mês = saldo final
        try:
            saldo_ini = Decimal(str(getattr(df_day, "attrs", {}).get("saldo_inicial", 0.0))).quantize(Decimal("0.01"))
            dif_mes = (saldo_mes - saldo_ini).quantize(Decimal("0.01"))
            if hasattr(self, "s3_audit_line"):
                self.s3_audit_line.configure(text=f"Saldo inicial: {br_money(saldo_ini)}  +  Diferença do mês: {br_money(dif_mes)}  =  Saldo final: {br_money(saldo_mes)}")
        except Exception:
            if hasattr(self, "s3_audit_line"):
                self.s3_audit_line.configure(text="")

        # habilita confirmar saldo quando há saldo != 0
        self.s3_last_saldo_mes = saldo_mes
        if saldo_mes != Decimal("0"):
            self.btn_s3_confirm.state(["!disabled"])
        else:
            self.btn_s3_confirm.state(["disabled"])

        # pequeno aviso no título do botão quando spillover foi acionado
        try:
            if used_sp:
                self.btn_s3_run.configure(text="Rodar Conciliação (spillover usado)")
            else:
                self.btn_s3_run.configure(text="Rodar Conciliação")
        except Exception:
            pass

        if not silent:
            self._log("Etapa 3 (visão diária) atualizada na tela.")

    def _confirm_step3_carryover(self):
        """Confirma o saldo do mês para ser carregado no mês seguinte."""
        m = self._parse_month()
        if not m:
            return
        prov = (self.provider.get().strip().upper() or "ALELO")
        saldo_mes = self.s3_last_saldo_mes if hasattr(self, "s3_last_saldo_mes") else Decimal("0")
        if saldo_mes == Decimal("0"):
            messagebox.showinfo("Confirmar saldo", "Não há saldo para confirmar neste mês.")
            return

        if not messagebox.askyesno(
            "Confirmar saldo",
            f"Confirmar {br_money(saldo_mes)} para carregar no mês seguinte?\n\n"
            "Isso serve para sobras/faltas de repasse que devem ser perseguidas no próximo fechamento.",
        ):
            return

        try:
            set_confirmed_carryover(self.conn, prov, m, saldo_mes)
            self._log(f"Saldo confirmado para o próximo mês: {br_money(saldo_mes)}")
            messagebox.showinfo("Confirmar saldo", "Saldo confirmado com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não consegui confirmar o saldo: {e}")

    
    def _open_step3_day_detail(self, event=None):
        sel = self.s3_tree.selection()
        if not sel:
            return
        vals = self.s3_tree.item(sel[0], "values") or ()
        if not vals:
            return
        d = parse_br_date_str(str(vals[0]))
        if not d:
            return

        prov = (self.provider.get().strip().upper() or "ALELO")

        win = tk.Toplevel(self)
        win.title(f"Etapa 3 - Detalhes do dia {fmt_br_date(d)}")
        win.geometry("980x620")

        top = ttk.Frame(win, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text=f"DATA: {fmt_br_date(d)}", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(top, text=f"Esperado: {vals[1]}   |   Banco: {vals[2]}   |   Dif: {vals[3]}   |   Saldo acum: {vals[4]}").pack(anchor="w", pady=(2,0))

        body = ttk.Frame(win, padding=10)
        body.pack(fill="both", expand=True)

        # ---- Recebíveis do dia (EDITÁVEL) ----
        lf_r = ttk.Labelframe(body, text="Recebíveis (Etapa 2) — editável", padding=8)
        lf_r.pack(fill="both", expand=True, side="top")

        cols_r = ("id","dt","pay_dt","bruto","liquido","autorizacao")
        tr = ttk.Treeview(lf_r, columns=cols_r, show="headings", height=10)
        for c,w,anch,txt in [
            ("id",70,"e","ID"),
            ("dt",95,"w","DT (base)"),
            ("pay_dt",95,"w","PAY_DT"),
            ("bruto",120,"e","BRUTO"),
            ("liquido",120,"e","LÍQUIDO"),
            ("autorizacao",160,"w","AUTORIZAÇÃO"),
        ]:
            tr.heading(c, text=txt); tr.column(c, width=w, anchor=anch)
        tr.pack(fill="both", expand=True, side="left")

        sb_r = ttk.Scrollbar(lf_r, orient="vertical", command=tr.yview)
        tr.configure(yscrollcommand=sb_r.set)
        sb_r.pack(side="right", fill="y")

        # --- Banco (somente leitura) ---
        lf_b = ttk.LabelFrame(body, text="Banco — somente leitura", padding=8)
        lf_b.pack(fill="both", expand=False, side="top", pady=(6,0))

        bank_sum_var = tk.StringVar(value="Banco do dia (filtrado): R$ 0,00 | Lançamentos: 0")
        ttk.Label(lf_b, textvariable=bank_sum_var, font=("Segoe UI", 9)).pack(anchor="w", pady=(0,4))

        cols_b = ("id","dt","amount","memo","bank_id")
        tb = ttk.Treeview(lf_b, columns=cols_b, show="headings", height=6)
        for c,w,anch,txt in [
            ("id",70,"e","ID"),
            ("dt",95,"w","DT"),
            ("amount",120,"e","CRÉDITO"),
            ("memo",420,"w","MEMO"),
            ("bank_id",160,"w","BANK_ID"),
        ]:
            tb.heading(c, text=txt); tb.column(c, width=w, anchor=anch)
        tb.pack(fill="both", expand=True, side="left")

        sb_b = ttk.Scrollbar(lf_b, orient="vertical", command=tb.yview)
        tb.configure(yscrollcommand=sb_b.set)
        sb_b.pack(side="right", fill="y")

        def _memo_ok(memo: str) -> bool:
            token = (self.bank_keyword.get() if hasattr(self, "bank_keyword") else "").strip()
            return bank_memo_match(token, memo, self.provider.get() if hasattr(self,"provider") else None)

        # carrega lançamentos do banco do dia (somente créditos) respeitando token/memo
        # carrega lançamentos do banco do dia (somente créditos) respeitando token/memo
        # OBS: algumas importações podem gravar dt com sufixo de hora; por isso usamos LIKE (prefixo YYYY-MM-DD)
        dt_prefix = d.isoformat()
        b_rows = self.conn.execute(
            "SELECT id, dt, amount, memo, bank_id FROM bank_tx "
            "WHERE provider=? AND is_deleted=0 AND dt LIKE ? AND amount > 0",
            (prov, dt_prefix + "%"),
        ).fetchall()
        for b0 in b_rows:
            b = dict(b0)
            if not _memo_ok(b.get("memo") or ""):
                continue
            tb.insert("", "end", values=(
                b.get("id"),
                fmt_br_date(parse_date(b.get("dt")) or d),
                br_money(parse_decimal(b.get("amount")) or Decimal("0")),
                (b.get("memo") or "")[:500],
                b.get("bank_id") or "",
            ))

        # atualiza resumo do banco (filtrado)
        try:
            total_bank_day = Decimal("0")
            count_bank_day = 0
            for iid in tb.get_children(""):
                vals = tb.item(iid, "values")
                total_bank_day += parse_decimal(str(vals[2])) or Decimal("0")
                count_bank_day += 1
            bank_sum_var.set(f"Banco do dia (filtrado): {br_money(total_bank_day)} | Lançamentos: {count_bank_day}")
        except Exception:
            pass

        # carrega recebíveis cujo dt_efetiva = d
        rows = self.conn.execute(
            "SELECT id, dt, pay_dt, bruto, liquido, autorizacao FROM receb_tx "
            "WHERE provider=? AND is_deleted=0",
            (prov,),
        ).fetchall()
        for r0 in rows:
            r = dict(r0)
            dt_target = parse_date(r.get("pay_dt")) or parse_date(r.get("dt"))
            if not dt_target:
                continue
            dt_eff = next_business_day(dt_target)
            if dt_eff != d:
                continue
            tr.insert("", "end", values=(
                r["id"],
                fmt_br_date(parse_date(r.get("dt")) or dt_eff),
                fmt_br_date(parse_date(r.get("pay_dt")) or dt_target),
                fmt_money(parse_decimal(r.get("bruto")) or Decimal("0")),
                fmt_money(parse_decimal(r.get("liquido")) or Decimal("0")),
                str(r.get("autorizacao") or ""),
            ))

        btns = ttk.Frame(win, padding=(10,0,10,10))
        btns.pack(fill="x")

        def _edit_selected():
            sel2 = tr.selection()
            if not sel2:
                return
            v = tr.item(sel2[0], "values")
            rid = int(v[0])

            cur = self.conn.execute(
                "SELECT dt, pay_dt, bruto, liquido, autorizacao FROM receb_tx WHERE id=?",
                (rid,),
            ).fetchone()
            if not cur:
                return
            cur = dict(cur)

            dlg = tk.Toplevel(win)
            dlg.title(f"Editar recebível #{rid}")
            center_window(dlg, 420, 260)
            frm = ttk.Frame(dlg, padding=10); frm.pack(fill="both", expand=True)

            def _row(label, value):
                r = ttk.Frame(frm); r.pack(fill="x", pady=4)
                ttk.Label(r, text=label, width=14).pack(side="left")
                e = ttk.Entry(r); e.pack(side="left", fill="x", expand=True)
                e.insert(0, value)
                return e

            e_dt = _row("DT (base)", (fmt_br_date(parse_date(cur.get("dt"))) if parse_date(cur.get("dt")) else (cur.get("dt") or "")))
            e_pay = _row("PAY_DT", (fmt_br_date(parse_date(cur.get("pay_dt"))) if parse_date(cur.get("pay_dt")) else (cur.get("pay_dt") or "")))
            e_br = _row("BRUTO", str(cur.get("bruto") or "0"))
            e_liq = _row("LÍQUIDO", str(cur.get("liquido") or "0"))
            e_aut = _row("Autorização", str(cur.get("autorizacao") or ""))

            def _save():
                try:
                    ndt = parse_date(e_dt.get().strip()) or parse_date(cur.get("dt"))
                    npay = parse_date(e_pay.get().strip()) or None
                    nbr = (parse_decimal(e_br.get().strip()) or Decimal('0')).quantize(Decimal('0.01'))
                    nliq = (parse_decimal(e_liq.get().strip()) or Decimal('0')).quantize(Decimal('0.01'))
                    naut = e_aut.get().strip()
                    self.conn.execute(
                        "UPDATE receb_tx SET dt=?, pay_dt=?, bruto=?, liquido=?, autorizacao=? WHERE id=?",
                        (
                            (ndt.isoformat() if ndt else cur.get("dt")),
                            (npay.isoformat() if npay else None),
                            str(nbr),
                            str(nliq),
                            naut,
                            rid,
                        ),
                    )
                    self.conn.commit()
                    dlg.destroy()
                    # mantém a janela de detalhes aberta; recarrega recebíveis e banco
                    try:
                        for it in tr.get_children(""):
                            tr.delete(it)
                        rows2 = self.conn.execute(
                            "SELECT id, dt, pay_dt, bruto, liquido, autorizacao FROM receb_tx "
                            "WHERE provider=? AND is_deleted=0",
                            (prov,),
                        ).fetchall()
                        for r0 in rows2:
                            r = dict(r0)
                            dt_target = parse_date(r.get("pay_dt")) or parse_date(r.get("dt"))
                            if not dt_target:
                                continue
                            dt_eff = next_business_day(dt_target)
                            if dt_eff != d:
                                continue
                            tr.insert("", "end", values=(
                                r["id"],
                                fmt_br_date(parse_date(r.get("dt")) or d),
                                fmt_br_date(parse_date(r.get("pay_dt")) or d),
                                br_money(parse_decimal(r.get("bruto")) or Decimal("0")),
                                br_money(parse_decimal(r.get("liquido")) or Decimal("0")),
                                (r.get("autorizacao") or "")[:80],
                            ))
                    except Exception:
                        pass

                    try:
                        for it in tb.get_children(""):
                            tb.delete(it)
                        dt_prefix2 = d.isoformat()
                        b_rows2 = self.conn.execute(
                            "SELECT id, dt, amount, memo, bank_id FROM bank_tx "
                            "WHERE provider=? AND is_deleted=0 AND dt LIKE ? AND amount > 0",
                            (prov, dt_prefix2 + "%"),
                        ).fetchall()
                        total_bank_day2 = Decimal("0")
                        count_bank_day2 = 0
                        for b0 in b_rows2:
                            b = dict(b0)
                            if not _memo_ok(b.get("memo") or ""):
                                continue
                            amt = parse_decimal(b.get("amount")) or Decimal("0")
                            total_bank_day2 += amt
                            count_bank_day2 += 1
                            tb.insert("", "end", values=(
                                b.get("id"),
                                fmt_br_date(parse_date(b.get("dt")) or d),
                                br_money(amt),
                                (b.get("memo") or "")[:500],
                                b.get("bank_id") or "",
                            ))
                        try:
                            bank_sum_var.set(f"Banco do dia (filtrado): {br_money(total_bank_day2)} | Lançamentos: {count_bank_day2}")
                        except Exception:
                            pass
                    except Exception:
                        pass

                    try:
                        win.lift(); win.focus_force()
                    except Exception:
                        pass

                    self._run_step3(silent=True)

                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao salvar: {e}")

            ttk.Button(frm, text="Salvar", command=_save).pack(anchor="e", pady=(10,0))

        # duplo clique no recebível abre edição
        tr.bind("<Double-1>", lambda e: _edit_selected())

        def _delete_selected():
            sel2 = tr.selection()
            if not sel2:
                return
            v = tr.item(sel2[0], "values")
            rid = int(v[0])
            if not messagebox.askyesno("Confirmar", f"Excluir recebível #{rid}? (marcar como deletado)"):
                return
            try:
                self.conn.execute("UPDATE receb_tx SET is_deleted=1 WHERE id=?", (rid,))
                self.conn.commit()
                win.destroy()
                self._run_step3(silent=True)
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao excluir: {e}")


        def _edit_batch_paydt():
            # altera PAY_DT de TODOS os recebíveis exibidos neste dia (lote)
            ids = []
            for iid in tr.get_children(""):
                v = tr.item(iid, "values") or ()
                if v:
                    try:
                        ids.append(int(v[0]))
                    except Exception:
                        pass
            if not ids:
                messagebox.showinfo("Lote", "Nenhum recebível listado para este dia.")
                return

            new_s = simpledialog.askstring("Editar lote", f"Nova data de pagamento (dd/mm/aaaa) para {len(ids)} recebíveis do dia {fmt_br_date(d)}:")
            if not new_s:
                return
            nd = parse_br_date_str(new_s)
            if not nd:
                messagebox.showerror("Erro", "Data inválida. Use dd/mm/aaaa.")
                return

            try:
                qmarks = ",".join(["?"] * len(ids))
                self.conn.execute(
                    f"UPDATE receb_tx SET pay_dt=? WHERE id IN ({qmarks})",
                    (nd.isoformat(), *ids),
                )
                self.conn.commit()
                win.destroy()
                self._run_step3(silent=True)
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao editar lote: {e}")

        ttk.Button(btns, text="Editar recebível selecionado", command=_edit_selected).pack(side="left")
        ttk.Button(btns, text="Excluir recebível selecionado", command=_delete_selected).pack(side="left", padx=(8,0))
        ttk.Button(btns, text="Editar lote do dia (PAY_DT)", command=_edit_batch_paydt).pack(side="left", padx=(8,0))
        ttk.Button(btns, text="Fechar", command=win.destroy).pack(side="right")


    def _open_step3_composition(self, event):
        item = self.s3_tree.identify_row(event.y)
        if not item:
            return
        vals = self.s3_tree.item(item, "values")
        if not vals or len(vals) < 2:
            return
        try:
            bank_id = int(str(vals[1]))
        except Exception:
            return
        recebs = self.s3_comp.get(bank_id, []) or []

        win = tk.Toplevel(self); win.title(f"Composição do depósito (Bank ID {bank_id})")
        win.transient(self); win.grab_set(); win.focus_set()
        frm = ttk.Frame(win, padding=10); frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=f"Recebíveis que compõem o depósito {bank_id}:", font=("Segoe UI",10,"bold")).pack(anchor="w", pady=(0,8))

        cols = ("id","pay_dt","liq","autorizacao")
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for c,w,a,t in [("id",80,"e","ID"),("pay_dt",110,"w","DT PGTO"),("liq",140,"e","VALOR LÍQ"),("autorizacao",220,"w","NSU/CHAVE")]:
            tree.heading(c, text=t); tree.column(c, width=w, anchor=a)
        tree.pack(fill="both", expand=True)

        if not recebs:
            tree.insert("", "end", values=("-", "-", "-", "Nenhum recebível encontrado (não fechou soma)"))
        else:
            q = "SELECT id, COALESCE(pay_dt, dt) AS pdt, liquido, autorizacao FROM receb_tx WHERE id IN (%s)" % (",".join(["?"]*len(recebs)))
            rows = self.conn.execute(q, tuple(recebs)).fetchall()
            # ordena por data
            rows = sorted(rows, key=lambda r: (str(r["pdt"] or ""), int(r["id"])))
            for r in rows:
                tree.insert("", "end", values=(r["id"], fmt_br_date(parse_date(r["pdt"])), br_money(parse_decimal(r["liquido"])), (r["autorizacao"] or "")))

        ttk.Button(frm, text="Fechar", command=win.destroy).pack(anchor="e", pady=(8,0))


    
    def _export_step3_excel(self):
        """Exporta Etapa 3 (visão diária + lançamentos bancários) para Excel."""
        if not hasattr(self, "s3_tree"):
            return
        m = self._parse_month()
        if not m:
            return
        prov = (self.provider.get().strip().upper() or "ALELO")
        mm = f"{m.month:02d}-{m.year}"
        out = filedialog.asksaveasfilename(
            title="Salvar Etapa 3 (Excel)",
            defaultextension=".xlsx",
            initialfile=f"etapa3_{prov.lower()}_{mm}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out:
            return

        rows = []
        for iid in self.s3_tree.get_children():
            vals = self.s3_tree.item(iid, "values")
            rows.append({
                "DATA": vals[0],
                "ESPERADO": vals[1],
                "BANCO": vals[2],
                "DIF_DIA": vals[3],
                "SALDO_ACUM": vals[4],
                "ALOCADO": vals[5],
                "SALDO_BANCO": vals[6],
                "STATUS": vals[7],
            })
        df_day = pd.DataFrame(rows)
        df_bank = self.s3_bank_df if hasattr(self, "s3_bank_df") and self.s3_bank_df is not None else pd.DataFrame()

        try:
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                df_day.to_excel(writer, sheet_name="Etapa3_Diario", index=False)
                if df_bank is not None and not df_bank.empty:
                    df_bank.to_excel(writer, sheet_name="Banco_Detalhe", index=False)
            self._log(f"Etapa 3 exportada: {out}")
        except Exception as e:
            self._log_err(f"ERRO ao exportar Etapa 3: {e}")
            messagebox.showerror("Erro", f"Falha ao exportar: {e}")


    def _finalize_closing(self):
        m = self._parse_month()
        if not m: return
        prov = (self.provider.get().strip().upper() or "ALELO")
        mm = f"{m.month:02d}/{m.year}"
        close_month(self.conn, prov, mm)
        self._log(f"Fechamento finalizado: {mm} ({prov})")
        self._refresh_closings(); self._run_step3(silent=True)
    def _undo_closing(self):
        sel = self.close_tree.selection() if hasattr(self, 'close_tree') else ()
        if not sel:
            messagebox.showinfo('Info', 'Selecione um mês na lista de Fechamentos.')
            return
        vals = self.close_tree.item(sel[0], 'values')
        if not vals:
            return
        mm = str(vals[0])
        prov = (self.provider.get().strip().upper() or 'ALELO')
        if not messagebox.askyesno('Confirmar', f'Desfazer fechamento de {mm} ({prov})?'):
            return
        try:
            undo_month(self.conn, prov, mm)
            self._log(f'Fechamento desfeito: {mm} ({prov})')
            self._refresh_closings()
            self._run_step3(silent=True)
        except Exception as e:
            self._err(f'Falha ao desfazer fechamento: {e}')
    
    def _build_close_tab(self):
        top = ttk.Frame(self.tab_close, padding=10); top.pack(fill="x")
        ttk.Label(top, text="Ano:").pack(side="left")
        ttk.Spinbox(top, from_=2000, to=2100, textvariable=self.year_view, width=6, command=self._refresh_closings).pack(side="left", padx=6)
        ttk.Button(top, text="Atualizar", command=self._refresh_closings).pack(side="left", padx=6)
        ttk.Button(top, text="Desfazer fechamento (selecionado)", command=self._undo_closing).pack(side="left", padx=6)
        ttk.Label(top, text="Legenda: ").pack(side="left", padx=(20,0))
        tk.Label(top, text=" PENDENTE ", bg="#ffdddd").pack(side="left", padx=4)
        tk.Label(top, text=" FECHADO ", bg="#e7f7e7").pack(side="left", padx=4)
        cal = ttk.Frame(self.tab_close, padding=10); cal.pack(fill="x")
        self.month_boxes = {}
        months = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]
        for i, name in enumerate(months):
            lbl = tk.Label(cal, text=name, width=8, pady=6, relief="groove")
            lbl.grid(row=0, column=i, padx=4, pady=4, sticky="nsew")
            self.month_boxes[i+1] = lbl
        ttk.Separator(self.tab_close).pack(fill="x", pady=8)
        self.close_tree = ttk.Treeview(self.tab_close, columns=("mes","status","fechado_em"), show="headings", height=14)
        for c,w in (("mes",90),("status",100),("fechado_em",200)):
            self.close_tree.heading(c, text=c.upper()); self.close_tree.column(c, width=w, anchor="w")
        self.close_tree.pack(fill="both", expand=True, padx=10, pady=10)
        btm = ttk.Frame(self.tab_close, padding=10); btm.pack(fill="x")
        ttk.Button(btm, text="Exportar Excel (Fechamentos)", command=lambda: self._export_tree(self.close_tree, "fechamentos.xlsx")).pack(side="right")
    def _refresh_closings(self):
        prov = (self.provider.get().strip().upper() or "ALELO")
        first = first_imported_date(self.conn, prov)
        for i in self.close_tree.get_children(): self.close_tree.delete(i)
        if not first:
            for m, lbl in self.month_boxes.items(): lbl.configure(bg=self.cget("bg"))
            return
        year = int(self.year_view.get())
        for m, lbl in self.month_boxes.items():
            mm = f"{m:02d}/{year}"
            if is_month_closed(self.conn, prov, mm):
                lbl.configure(bg="#e7f7e7")
            else:
                if date(year, m, 1) >= date(first.year, first.month, 1) and date(year, m, 1) <= date(datetime.now().year, datetime.now().month, 1):
                    lbl.configure(bg="#ffdddd")
                else:
                    lbl.configure(bg=self.cget("bg"))
        months_list = iter_months(first, date(datetime.now().year, datetime.now().month, 1))
        for m0 in months_list:
            mm = f"{m0.month:02d}/{m0.year}"
            row = self.conn.execute("SELECT closed_at FROM closed_periods WHERE provider=? AND month=?", (prov, mm)).fetchone()
            st = "FECHADO" if row else "PENDENTE"
            closed_at = row["closed_at"] if row else ""
            self.close_tree.insert("", "end", values=(mm, st, closed_at))
    def _build_bank_tab(self):
        top = ttk.Frame(self.tab_bank, padding=10); top.pack(fill="x")
        ttk.Label(top, text="Período (dd/mm/aaaa):").pack(side="left")
        self.bank_from = tk.StringVar(value=""); self.bank_to = tk.StringVar(value="")
        ttk.Entry(top, textvariable=self.bank_from, width=11).pack(side="left", padx=(6,2))
        ttk.Label(top, text="a").pack(side="left")
        ttk.Entry(top, textvariable=self.bank_to, width=11).pack(side="left", padx=(2,10))
        ttk.Label(top, text="Termo (memo):").pack(side="left")
        self.bank_term = tk.StringVar(value="")
        ent = ttk.Entry(top, textvariable=self.bank_term, width=16)
        ent.pack(side="left", padx=(6,6))
        ttk.Button(top, text="Definir palavras-chave…", command=self._open_bank_keywords_editor).pack(side="left", padx=(0,8))
        self.bank_kw_lbl = ttk.Label(top, text=self._bank_kw_status())
        self.bank_kw_lbl.pack(side="left", padx=(0,10))
        ttk.Button(top, text="Pesquisar", command=self._refresh_bank_search).pack(side="left")
        ttk.Button(top, text="Exportar Excel", command=lambda: self._export_tree(self.bank_tree, "banco_pesquisa.xlsx")).pack(side="right", padx=5)
        self.bank_total_lbl = ttk.Label(top, text="Total do período: R$ -", font=("Segoe UI",10,"bold"))
        self.bank_total_lbl.pack(side="right", padx=10)
        cols=("dt","amount","memo","bank_id")
        self.bank_tree = ttk.Treeview(self.tab_bank, columns=cols, show="headings", height=18)
        for c,w in (("dt",110),("amount",140),("memo",760),("bank_id",140)):
            self.bank_tree.heading(c, text=c.upper()); self.bank_tree.column(c, width=w, anchor="w" if c in ("dt","memo","bank_id") else "e")
        self.bank_tree.pack(fill="both", expand=True, padx=10, pady=10)
    def _refresh_bank_search(self):
        if not hasattr(self, "bank_tree"):
            return
        # FIX: garantia de StringVar (algumas builds criam o label depois do refresh)
        if not hasattr(self, "bank_total_var") or self.bank_total_var is None:
            self.bank_total_var = tk.StringVar(value="")
        for i in self.bank_tree.get_children():
            self.bank_tree.delete(i)

        d1 = parse_br_date_str(self.bank_from.get())
        d2 = parse_br_date_str(self.bank_to.get())
        if (self.bank_from.get().strip() or self.bank_to.get().strip()) and (not d1 or not d2):
            self._err("Banco/Pesquisa: período inválido (dd/mm/aaaa).")
            return

        prov = (self.provider.get().strip().upper() or "ALELO")

        term = (self.bank_term.get() or "").strip()
        if not term:
            # Só pesquisa após digitar palavras-chave (evita puxar automaticamente lançamentos de outras bandeiras)
            self.bank_total_var.set("Total do período: " + br_money(Decimal("0")))
            return
        where = "WHERE provider='BANCO' AND is_deleted=0"
        params: list = []

        if d1 and d2:
            where += " AND dt>=? AND dt<=?"
            params += [d1.isoformat(), d2.isoformat()]

        if term:
            t = term.strip().upper()
            is_farm_token = (t == BANK_TRANSF_TOKEN_FARM or t.startswith(BANK_TRANSF_TOKEN_FARM) or BANK_TRANSF_TOKEN_FARM.startswith(t))
            is_alelo_token = (t == BANK_TRANSF_TOKEN_ALELO or t.startswith(BANK_TRANSF_TOKEN_ALELO) or BANK_TRANSF_TOKEN_ALELO.startswith(t))

            if (prov == "FARMACIASAPP" and is_farm_token) or (prov != "FARMACIASAPP" and is_alelo_token):
                terms = bank_all_memo_terms(prov)
                if terms:
                    ors = " OR ".join(["UPPER(COALESCE(memo,'')) LIKE ?"] * len(terms))
                    where += f" AND ({ors})"
                    params += [f"%{x.upper()}%" for x in terms]
            else:
                where += " AND UPPER(COALESCE(memo,'')) LIKE ?"
                params.append(f"%{t}%")

        rows = self.conn.execute(
            f"SELECT dt, amount, memo, bank_id FROM bank_tx {where} ORDER BY dt, id LIMIT 5000",
            tuple(params),
        ).fetchall()

        tot = Decimal("0")
        for r in rows:
            dt, amount, memo, bid = r
            a = parse_decimal(amount) or Decimal("0")
            tot += a
            self.bank_tree.insert("", "end", values=(fmt_date(dt), br_money(a), str(memo or ""), str(bid or "")))

        self.bank_total_var.set(f"Total do período: {br_money(tot)}")

    def _export_tree(self, tree: ttk.Treeview, default_name: str):
        p = filedialog.asksaveasfilename(title="Salvar Excel", initialfile=default_name, defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not p: return
        try:
            export_tree_to_excel(tree, p); self._log(f"Exportado: {p}")
        except Exception as e:
            self._err(f"Falha ao exportar: {e}")
    def _day_from_tree(self, tree: ttk.Treeview, event):
        item = tree.identify_row(event.y)
        if not item: return None
        vals = tree.item(item, "values")
        return parse_br_date_str(str(vals[0])) if vals else None
    def _open_day_selector(self, tree: ttk.Treeview, event, step: int):
        day = self._day_from_tree(tree, event)
        if not day: return
        prov = (self.provider.get().strip().upper() or "ALELO")
        win = tk.Toplevel(self); win.title(f"Abrir registros do dia {fmt_br_date(day)}")
        win.transient(self)
        win.grab_set()
        win.focus_set()
        frm = ttk.Frame(win, padding=10); frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Escolha a tabela para abrir:", font=("Segoe UI",10,"bold")).pack(anchor="w", pady=(0,8))
        def refresh_parent():
            self._run_step1(silent=True); self._run_step2(silent=True); self._run_step3(silent=True); self._refresh_closings()
        if step == 1:
            ttk.Button(frm, text="ERP (do dia)", command=lambda: UnderlyingByDayPopup(self, self.conn, prov, day, "erp_tx", refresh_parent)).pack(fill="x", pady=4)
            ttk.Button(frm, text="Vendas (Portal) (do dia)", command=lambda: UnderlyingByDayPopup(self, self.conn, prov, day, "sales_tx", refresh_parent)).pack(fill="x", pady=4)
        else:
            ttk.Button(frm, text="Vendas (Portal) (do dia)", command=lambda: UnderlyingByDayPopup(self, self.conn, prov, day, "sales_tx", refresh_parent)).pack(fill="x", pady=4)
            ttk.Button(frm, text="Recebimentos (do dia)", command=lambda: UnderlyingByDayPopup(self, self.conn, prov, day, "receb_tx", refresh_parent)).pack(fill="x", pady=4)
        win.resizable(False, False)
    def _open_bulk_delete(self):
        prov = (self.provider.get().strip().upper() or "ALELO")
        BulkDeletePopup(self, self.conn, prov, refresh_cb=self._refresh_all, log_cb=self._log, err_cb=self._err)
    
    
    # ------------------
    # Diagnóstico (DB)
    # ------------------
    
    
    # ------------------------------
    # Relatório de Divergências (Auditoria)
    # ------------------------------
    def _build_report_tab(self):
        top = ttk.Frame(self.tab_rep, padding=10); top.pack(fill="x")
        ttk.Label(top, text="Relatório de Divergências (auditoria / contestação)", font=("Segoe UI", 11, "bold")).pack(anchor="w")
    
        opts = ttk.Frame(self.tab_rep, padding=(10,0,10,10)); opts.pack(fill="x")
        self.rep_opt_s2 = tk.BooleanVar(value=True)
        self.rep_opt_s3 = tk.BooleanVar(value=True)
        ttk.Checkbutton(opts, text="Incluir Etapa 2 (Vendas x Recebimentos) - por Nº Autorização", variable=self.rep_opt_s2).pack(anchor="w", pady=2)
        ttk.Checkbutton(opts, text="Incluir Etapa 3 (Recebimentos x Banco) - diferença do fechamento", variable=self.rep_opt_s3).pack(anchor="w", pady=2)
    
        act = ttk.Frame(self.tab_rep, padding=(10,0,10,10)); act.pack(fill="x")
        ttk.Button(act, text="Gerar relatório (Excel + texto)", command=self._generate_divergence_report).pack(side="left")
        ttk.Label(act, text="Dica: use os filtros no topo (mês e/ou período) antes de gerar.", foreground="#555").pack(side="left", padx=10)
    
        body = ttk.Frame(self.tab_rep, padding=10); body.pack(fill="both", expand=True)
        ttk.Label(body, text="Texto para enviar à operadora (copiar e colar):", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,6))
        self.rep_text = tk.Text(body, height=18, wrap="word")
        self.rep_text.pack(fill="both", expand=True)
        self.rep_text.insert("1.0", "Gere o relatório para preencher este texto automaticamente.\n")
    
        btns = ttk.Frame(body); btns.pack(fill="x", pady=(8,0))
        ttk.Button(btns, text="Copiar texto", command=self._copy_report_text).pack(side="left")
    
    def _copy_report_text(self):
        try:
            txt = self.rep_text.get("1.0", "end-1c")
            self.clipboard_clear()
            self.clipboard_append(txt)
            self._log("Texto do relatório copiado para a área de transferência.")
        except Exception as e:
            self._err(f"Falha ao copiar texto: {e}")
    
    def _generate_divergence_report(self):
        include_s2 = bool(self.rep_opt_s2.get())
        include_s3 = bool(self.rep_opt_s3.get())
        if not include_s2 and not include_s3:
            messagebox.showwarning("Relatório", "Selecione pelo menos uma opção (Etapa 2 e/ou Etapa 3).")
            return
    
        m = self._parse_month()
        if not m:
            return
        d1, d2 = self._parse_period()  # pode ser None/None
        prov = (self.provider.get().strip().upper() or "ALELO")
        start, end = _get_period(m, d1, d2)
    
        df2 = pd.DataFrame()
        df3 = pd.DataFrame()
    
        if include_s2:
            df2 = self._build_divergences_step2_by_auth(prov, start, end)
    
        if include_s3:
            kw = (self.bank_keyword.get() or "").strip()
            totals = run_step3_monthly(self.conn, prov, m, bank_keyword=kw if kw else None)
            diff = totals.get("div_banco_vs_receb", Decimal("0")) or Decimal("0")
            if diff != 0:
                df3 = pd.DataFrame([{
                    "Mês": totals.get("mes",""),
                    "Total Recebimentos (líq)": totals.get("receb_liq", Decimal("0")),
                    "Total Banco": totals.get("banco", Decimal("0")),
                    "Diferença (Banco - Receb)": diff
                }])
            else:
                df3 = pd.DataFrame([{
                    "Mês": totals.get("mes",""),
                    "Total Recebimentos (líq)": totals.get("receb_liq", Decimal("0")),
                    "Total Banco": totals.get("banco", Decimal("0")),
                    "Diferença (Banco - Receb)": Decimal("0")
                }])
    
        # salvar Excel
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"relatorio_divergencias_{prov}_{ts}.xlsx"
        fpath = filedialog.asksaveasfilename(
            title="Salvar relatório (Excel)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not fpath:
            return
    
        with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
            # Resumo
            summary_rows = []
            summary_rows.append({"Item":"Bandeira", "Valor":prov})
            summary_rows.append({"Item":"Período", "Valor":f"{fmt_br_date(start)} a {fmt_br_date(end)}"})
            if include_s2:
                summary_rows.append({"Item":"Etapa 2 - divergências", "Valor":int(len(df2))})
            if include_s3:
                summary_rows.append({"Item":"Etapa 3 - divergência", "Valor":("" if df3.empty else br_money(parse_decimal(df3.iloc[0].get('Diferença (Banco - Receb)',0))))})
            pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Resumo")
    
            if include_s2:
                if df2.empty:
                    pd.DataFrame([{"Info":"Nenhuma divergência encontrada na Etapa 2 para o período."}]).to_excel(writer, index=False, sheet_name="Etapa2_Diverg")
                else:
                    out2 = df2.copy()
                    # formato monetário em texto, para leitura rápida
                    for c in ["Valor Bruto (Vendas)", "Valor Líquido (Vendas)", "Valor Bruto (Receb)", "Valor Líquido (Receb)", "Diferença Líq (Receb - Venda)", "Diferença Bruto (Receb - Venda)"]:
                        if c in out2.columns:
                            out2[c] = out2[c].apply(br_money)
                    out2.to_excel(writer, index=False, sheet_name="Etapa2_Diverg")
    
            if include_s3:
                if df3.empty:
                    pd.DataFrame([{"Info":"Nenhuma divergência encontrada na Etapa 3 para o mês selecionado."}]).to_excel(writer, index=False, sheet_name="Etapa3_Diverg")
                else:
                    out3 = df3.copy()
                    for c in ["Total Recebimentos (líq)", "Total Banco", "Diferença (Banco - Receb)"]:
                        if c in out3.columns:
                            out3[c] = out3[c].apply(br_money)
                    out3.to_excel(writer, index=False, sheet_name="Etapa3_Diverg")
    
        # texto corporativo
        lines = []
        lines.append(f"Prezados, bom dia.\n")
        lines.append(f"Segue para análise e contestação o relatório de divergências da bandeira {prov}, referente ao período {fmt_br_date(start)} a {fmt_br_date(end)}.")
        lines.append("")
        if include_s2:
            if df2.empty:
                lines.append("• Etapa 2 (Vendas x Recebimentos): não foram identificadas divergências por Nº de Autorização no período.")
            else:
                lines.append(f"• Etapa 2 (Vendas x Recebimentos): identificadas {len(df2)} ocorrência(s) em que o valor do recebimento diverge do valor registrado no relatório de vendas (por Nº de Autorização).")
        if include_s3:
            if df3.empty:
                lines.append("• Etapa 3 (Recebimentos x Banco): não foram identificadas divergências no fechamento do mês selecionado.")
            else:
                try:
                    diff_txt = br_money(parse_decimal(df3.iloc[0].get("Diferença (Banco - Receb)",0)))
                except Exception:
                    diff_txt = "-"
                lines.append(f"• Etapa 3 (Recebimentos x Banco): diferença apurada no fechamento: {diff_txt}.")
        lines.append("")
        lines.append("O arquivo Excel anexo contém o detalhamento para conferência e providências.\n")
        lines.append("Atenciosamente")
        lines.append("Daniel Del Bizogno | (31) 98862-4936,")
        lines.append("Drogaria Bem Estar | 12.070.242/0001-30")
    
        self.rep_text.delete("1.0", "end")
        self.rep_text.insert("1.0", "\n".join(lines))
    
        self._log(f"Relatório gerado: {fpath}")
    
    def _build_divergences_step2_by_auth(self, prov: str, start: date, end: date) -> pd.DataFrame:
        # Vendas no período (por dt) e recebimentos no período (por pay_dt se existir)
        q_sales = """SELECT dt, bruto, liquido, autorizacao, raw_json
                     FROM sales_tx
                     WHERE provider=? AND is_deleted=0 AND dt>=? AND dt<=? AND COALESCE(autorizacao,'')<>''"""
        sales = self.conn.execute(q_sales, (prov, start.isoformat(), end.isoformat())).fetchall()
    
        q_rec = """SELECT dt, pay_dt, bruto, liquido, autorizacao, raw_json
                     FROM receb_tx
                     WHERE provider=? AND is_deleted=0 AND COALESCE(autorizacao,'')<>''"""
        # Nota: para auditoria por "mês de venda", não filtramos recebimentos por pay_dt no período.
        # Assim, vendas do mês que pagaram no mês seguinte não viram "sem recebimento".
        recs = self.conn.execute(q_rec, (prov,)).fetchall()
    
        # Index por autorização (se vierem várias, mantém a primeira e registra o resto como observação)
        rec_by_auth = {}
        for r in recs:
            auth = (r["autorizacao"] or "").strip()
            if auth and auth not in rec_by_auth:
                rec_by_auth[auth] = r
    
        rows = []
        for s in sales:
            auth = (s["autorizacao"] or "").strip()
            if not auth:
                continue
            r = rec_by_auth.get(auth)
            sale_dt = parse_date(s["dt"])
            venda_br = parse_decimal(s["bruto"]) or Decimal("0")
            venda_lq = parse_decimal(s["liquido"]) or Decimal("0")
    
            # dados extras do raw_json (status / data pagamento, quando existir)
            status = ""
            data_pag = ""
            try:
                raw = json.loads(s["raw_json"]) if s["raw_json"] else {}
                status = raw.get("Status") or raw.get("STATUS") or raw.get("Situação") or raw.get("Situacao") or ""
                data_pag = raw.get("Data de Pagamento") or raw.get("Data Pagamento") or raw.get("Data do Pagamento") or raw.get("Data de crédito/débito") or ""
            except Exception:
                pass
    
            # normaliza data de pagamento (vendas) para evitar notação científica no Excel
            dp = fmt_br_date(parse_any_date(data_pag)) if data_pag else ""

            if r is None:
                # sem recebimento correspondente no período -> útil para auditoria também
                rows.append({
                    "Nº Autorização": auth,
                    "Data da Venda": fmt_br_date(sale_dt),
                    "Valor Bruto (Vendas)": venda_br,
                    "Valor Líquido (Vendas)": venda_lq,
                    "Status (Vendas)": status,
                    "Data de Pagamento (Vendas)": dp,
                    "Data de Pagamento (Receb)": "",
                    "Valor Bruto (Receb)": Decimal("0"),
                    "Valor Líquido (Receb)": Decimal("0"),
                    "Diferença Líq (Receb - Venda)": (Decimal("0") - venda_lq),
                    "Diferença Bruto (Receb - Venda)": (Decimal("0") - venda_br),
                    "Situação": "SEM RECEBIMENTO (BASE)"
                })
                continue
    
            receb_pay = parse_date(r["pay_dt"] or r["dt"])
            rec_br = parse_decimal(r["bruto"]) or Decimal("0")
            rec_lq = parse_decimal(r["liquido"]) or Decimal("0")
    
            diff_lq = rec_lq - venda_lq
            diff_br = rec_br - venda_br
    
            if diff_lq != 0 or diff_br != 0:
                situ = "DIVERGENTE"
                if receb_pay and (receb_pay < start or receb_pay > end):
                    situ = "DIVERGENTE (PAGTO FORA DO PERÍODO)"
                rows.append({
                    "Nº Autorização": auth,
                    "Data da Venda": fmt_br_date(sale_dt),
                    "Valor Bruto (Vendas)": venda_br,
                    "Valor Líquido (Vendas)": venda_lq,
                    "Status (Vendas)": status,
                    "Data de Pagamento (Vendas)": dp,
                    "Data de Pagamento (Receb)": fmt_br_date(receb_pay),
                    "Valor Bruto (Receb)": rec_br,
                    "Valor Líquido (Receb)": rec_lq,
                    "Diferença Líq (Receb - Venda)": diff_lq,
                    "Diferença Bruto (Receb - Venda)": diff_br,
                    "Situação": situ
                })
    
        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values(by=["Data da Venda","Nº Autorização"], ascending=True)
        return df
    
    
    def _build_diag_tab(self):
        # Etapa 4 - Taxas/Tarifas (Contratada x Aplicada) sobre recebimentos
        top = ttk.Frame(self.tab_diag, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Etapa 4 - Taxas/Tarifas (Esperada x Aplicada)", font=("Segoe UI", 10, "bold")).pack(side="left")

        ttk.Button(top, text="Cadastro (Taxas/Bandeiras)…", command=self._open_fee_modal).pack(side="right", padx=(6,0))
        ttk.Button(top, text="Tolerância…", command=self._open_fee_tolerance).pack(side="right", padx=(6,0))
        ttk.Button(top, text="Exportar Excel", command=self._export_step4).pack(side="right", padx=(6,0))
        ttk.Button(top, text="Rodar Etapa 4", command=lambda: run_step4(self)).pack(side="right", padx=(6,0))
        ttk.Button(top, text="Conferência Ticket (dinâmica)", command=lambda: run_step4_ticket_conferencia_dinamica(self)).pack(side="right", padx=(6,0))

        info = ttk.Frame(self.tab_diag, padding=(10,0,10,8))
        info.pack(fill="x")
        self.s4_tot_var = tk.StringVar(value="Taxa esperada: R$ 0,00 | Taxa aplicada: R$ 0,00 | Diferença: R$ 0,00")
        self.s4_div_var = tk.StringVar(value="Divergências: 0")
        ttk.Label(info, textvariable=self.s4_tot_var, font=("Segoe UI", 10, "bold")).pack(side="left")
        ttk.Label(info, textvariable=self.s4_div_var, foreground="#b00020").pack(side="left", padx=(15,0))

        cols = ("sale_dt","pay_dt","tipo","nsu","bruto","liquido","taxa_aplic","taxa_esp","transf","esp_banco","dif","status")
        self.s4_tree = ttk.Treeview(self.tab_diag, columns=cols, show="headings", height=20)
        headers = {
            "sale_dt":"DATA VENDA",
            "pay_dt":"DATA PAGTO",
            "tipo":"TIPO CARTÃO",
            "nsu":"NSU/AUT.",
            "bruto":"BRUTO",
            "liquido":"LÍQUIDO",
            "taxa_aplic":"TAXA APLIC.",
            "taxa_esp":"TAXA ESPER.",
            "transf":"TX TRANSF.",
            "esp_banco":"ESP. BANCO",
            "dif":"DIF (R$)",
            "status":"STATUS"
        }
        widths = {
            "sale_dt":90, "pay_dt":90, "tipo":150, "nsu":100,
            "bruto":90, "liquido":90, "taxa_aplic":90, "taxa_esp":90, "transf":90, "esp_banco":95,
            "dif":85, "status":80
        }
        for c in cols:
            self.s4_tree.heading(c, text=headers.get(c,c))
            self.s4_tree.column(c, width=widths.get(c,90), anchor="w" if c in ("tipo",) else "center")
        self.s4_tree.pack(fill="both", expand=True, padx=10, pady=(0,8))
        self.s4_tree.tag_configure("div", background="#ffdddd")

        self.s4_tree.bind("<Double-1>", self._on_step4_dblclick)

        # armazenamento do último DF para exportação
        self._s4_last_df = pd.DataFrame()

    def _open_fee_tolerance(self):
        # abre o mesmo modal (campo fica no topo)
        self._open_fee_modal()

    def _open_fee_modal(self):
        if not hasattr(self, "conn") or self.conn is None:
            try:
                self.conn = connect(self.db_path.get()); init_db(self.conn)
            except Exception:
                pass
        win = tk.Toplevel(self)
        win.title("Taxas/Tarifas - Cadastro")
        win.geometry("860x560")
        win.grab_set()

        # tolerância
        tol_frame = ttk.Labelframe(win, text="Configuração", padding=10)
        tol_frame.pack(fill="x", padx=10, pady=(10,6))

        cur_tol = ui_get(self.conn, "fees_tolerance", "0.05") if getattr(self, "conn", None) else "0.05"
        tol_var = tk.StringVar(value=str(cur_tol))
        ttk.Label(tol_frame, text="Tolerância (R$):").pack(side="left")
        tol_ent = ttk.Entry(tol_frame, textvariable=tol_var, width=10)
        tol_ent.pack(side="left", padx=(6,10))
        ttk.Label(tol_frame, text="(Diferenças <= tolerância serão OK)").pack(side="left")

        def save_tol():
            try:
                v = (tol_var.get() or "").replace(",", ".").strip()
                d = Decimal(v)
                if d < 0:
                    raise ValueError
                ui_set(self.conn, "fees_tolerance", str(d))
                messagebox.showinfo("Ok", "Tolerância salva.")
            except Exception:
                messagebox.showerror("Erro", "Tolerância inválida. Ex.: 0,05")

        ttk.Button(tol_frame, text="Salvar tolerância", command=save_tol).pack(side="right")

        # regras
        rules_frame = ttk.Labelframe(win, text="Regras por Transação (Taxa ADM + Taxa Pix/Crédito + Tarifa fixa)", padding=10)
        rules_frame.pack(fill="both", expand=True, padx=10, pady=(0,10))

        cols = ("id","label","match","mdr","fixed","transf","active")
        tree = ttk.Treeview(rules_frame, columns=cols, show="headings", height=16)
        headers = {"id":"ID","label":"Bandeira/Regra","match":"Match (Forma Pgto contém)","mdr":"Tx ADM (%)","fixed":"Tx Créd/Pix (%)","transf":"Tarifa fixa (R$)","active":"Ativo"}
        widths = {"id":60,"label":190,"match":220,"mdr":90,"fixed":90,"transf":110,"active":70}
        for c in cols:
            tree.heading(c, text=headers[c])
            tree.column(c, width=widths[c], anchor="w" if c in ("label","match") else "center")
        tree.pack(fill="both", expand=True, side="left")
        ysb = ttk.Scrollbar(rules_frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=ysb.set)
        ysb.pack(fill="y", side="left")

        btns = ttk.Frame(rules_frame)
        btns.pack(fill="y", side="right", padx=(10,0))

        def load_rules():
            tree.delete(*tree.get_children())
            prov = (self.provider.get() or "ALELO").strip().upper()
            rules = fee_rules_tx_list(self.conn, prov)
            for r in rules:
                tree.insert("", "end", iid=str(r["id"]), values=(
                    r["id"],
                    r["label"],
                    r["match_text"],
                    f"{r['mdr_percent']}",
                    f"{r['fee_fixed']}",
                    f"{r.get('transfer_fee',0)}",
                    "SIM" if r["is_active"] else "NÃO"
                ))

        def selected_rule_id():
            sel = tree.selection()
            return int(sel[0]) if sel else None

        def open_rule_editor(rule_id=None):
            prov = (self.provider.get() or "ALELO").strip().upper()
            data = {"label":"","match_text":"","mdr":"0","fixed":"0","transf":"0","active":1}
            if rule_id is not None:
                row = self.conn.execute("SELECT label, match_text, mdr_percent, fee_fixed, transfer_fee, is_active FROM fee_rules_tx WHERE id=?",
                                        (int(rule_id),)).fetchone()
                if row:
                    data = {"label":row[0], "match_text":row[1], "mdr":str(row[2] or "0"), "fixed":str(row[3] or "0"), "transf":str(row[4] or "0"), "active":int(row[5] or 0)}

            ew = tk.Toplevel(win)
            ew.title("Regra de taxa")
            ew.geometry("420x260")
            ew.grab_set()

            v_label = tk.StringVar(value=data["label"])
            v_match = tk.StringVar(value=data["match_text"])
            v_mdr = tk.StringVar(value=str(data["mdr"]).replace(".", ","))
            v_fix = tk.StringVar(value=str(data["fixed"]).replace(".", ","))
            v_trf = tk.StringVar(value=str(data.get("transf","0")).replace(".", ","))
            v_act = tk.IntVar(value=data["active"])

            frm = ttk.Frame(ew, padding=12); frm.pack(fill="both", expand=True)
            ttk.Label(frm, text="Nome (ex.: Alelo Alimentação):").grid(row=0, column=0, sticky="w")
            ttk.Entry(frm, textvariable=v_label, width=40).grid(row=1, column=0, sticky="we", pady=(0,8))
            ttk.Label(frm, text="Match Tipo Cartão contém (ex.: Alimentação):").grid(row=2, column=0, sticky="w")
            ttk.Entry(frm, textvariable=v_match, width=40).grid(row=3, column=0, sticky="we", pady=(0,8))

            row2 = ttk.Frame(frm); row2.grid(row=4, column=0, sticky="we")
            ttk.Label(row2, text="% Adm:").pack(side="left")
            ttk.Entry(row2, textvariable=v_mdr, width=10).pack(side="left", padx=(6,12))
            ttk.Label(row2, text="Tarifa (R$):").pack(side="left")
            ttk.Entry(row2, textvariable=v_fix, width=10).pack(side="left", padx=(6,12))
            ttk.Label(row2, text="Tx transf (R$):").pack(side="left")
            ttk.Entry(row2, textvariable=v_trf, width=10).pack(side="left", padx=(6,0))

            ttk.Checkbutton(frm, text="Ativo", variable=v_act).grid(row=5, column=0, sticky="w", pady=(10,0))

            def save():
                try:
                    label = v_label.get().strip()
                    match = v_match.get().strip()
                    mdr = Decimal((v_mdr.get() or "0").replace(",", "."))
                    fix = Decimal((v_fix.get() or "0").replace(",", "."))
                    trf = Decimal((v_trf.get() or "0").replace(",", "."))
                    fee_rule_tx_upsert(self.conn, prov, label, match, mdr, fix, trf, int(v_act.get()), rule_id=rule_id)
                    ew.destroy()
                    load_rules()
                except Exception as e:
                    messagebox.showerror("Erro", f"Não foi possível salvar: {e}")

            btm = ttk.Frame(frm); btm.grid(row=6, column=0, sticky="e", pady=(12,0))
            ttk.Button(btm, text="Cancelar", command=ew.destroy).pack(side="right")
            ttk.Button(btm, text="Salvar", command=save).pack(side="right", padx=(0,8))

            frm.columnconfigure(0, weight=1)

        def add_rule():
            open_rule_editor(None)

        def edit_rule():
            rid = selected_rule_id()
            if rid is None:
                return
            open_rule_editor(rid)

        def del_rule():
            rid = selected_rule_id()
            if rid is None:
                return
            if not messagebox.askyesno("Excluir", "Excluir a regra selecionada?"):
                return
            try:
                fee_rule_tx_delete(self.conn, rid)
                load_rules()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao excluir: {e}")

        ttk.Button(btns, text="Adicionar", command=add_rule).pack(fill="x", pady=(0,6))
        ttk.Button(btns, text="Editar", command=edit_rule).pack(fill="x", pady=(0,6))
        ttk.Button(btns, text="Excluir", command=del_rule).pack(fill="x", pady=(0,6))
        ttk.Separator(btns).pack(fill="x", pady=10)
        ttk.Button(btns, text="Fechar", command=win.destroy).pack(fill="x")


        # Seed automático (apenas se não houver regras)
        try:
            prov_now = (self.provider.get() or "ALELO").strip().upper()
            existing = fee_rules_tx_list(self.conn, prov_now)
            if not existing:
                if prov_now == "ALELO":
                    if messagebox.askyesno("Taxas/Tarifas", "Não encontrei regras de taxa para ALELO.\nDeseja criar automaticamente (Alimentação e Multibenefícios)?"):
                        seed_rules_alelo_if_empty(self.conn)
                elif prov_now == "TICKET":
                    if messagebox.askyesno("Taxas/Tarifas", "Não encontrei regras de taxa para TICKET.\nDeseja criar automaticamente (Voucher)?"):
                        seed_rules_ticket_if_empty(self.conn)
                elif prov_now == "FARMACIASAPP":
                    if messagebox.askyesno("Taxas/Tarifas", "Não encontrei regras de taxa para FARMACIASAPP.\nDeseja criar automaticamente (PIX/CREDIT 5% + 1%)?"):
                        seed_rules_farmaciasapp_if_empty(self.conn)
        except Exception:
            pass

        load_rules()



        # (FIX) _run_step4 moved to module-level run_step4(app)


    def _export_step4(self):
        if not hasattr(self, "_s4_last_df") or self._s4_last_df is None or self._s4_last_df.empty:
            messagebox.showinfo("Exportar", "Nada para exportar. Rode a Etapa 4 primeiro.")
            return
        try:
            from tkinter import filedialog
            prov = (self.provider.get() or "ALELO").strip().upper()
            fn = f"taxas_tarifas_{prov}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=fn,
                                                filetypes=[("Excel","*.xlsx")])
            if not path:
                return
            self._s4_last_df.to_excel(path, index=False)
            messagebox.showinfo("Exportar", "Arquivo gerado com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar: {e}")

    def _on_step4_dblclick(self, _evt):
        sel = self.s4_tree.selection()
        if not sel:
            return
        iid = sel[0]

        # Linhas agregadas (Ticket): não abrem editor
        if isinstance(iid, str) and (iid.startswith("P:") or iid.startswith("L:")):
            return

        # Alguns modos podem prefixar o ID
        if isinstance(iid, str) and iid.startswith("R:"):
            iid_num = iid[2:]
        else:
            iid_num = iid

        if not str(iid_num).isdigit():
            return

        rid = int(iid_num)

        def refresh():
            run_step4(self)

        try:
            EditRecordPopup(self, self.conn, "receb_tx", rid, refresh)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir editor: {e}")

def main():
    import argparse, sys
    p = argparse.ArgumentParser(description="Concilia (multi-bandeiras)")
    p.add_argument("--provider", default="", help="Força uma bandeira (ALELO/TICKET/FARMACIASAPP)")
    p.add_argument("--db", default="", help="Caminho do banco SQLite (opcional)")
    args = p.parse_args()

    prov = (args.provider or "").strip().upper() or None
    dbp = (args.db or "").strip() or None

    App(fixed_provider=prov, fixed_db_path=dbp).mainloop()

if __name__ == "__main__":
    main()